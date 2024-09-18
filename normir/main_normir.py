import json
import sqlite3
import sys
import well_data
import psycopg2

from PyQt5 import QtCore, QtWidgets
from PyQt5.QtGui import QIntValidator, QDoubleValidator
from PyQt5.QtWidgets import QApplication, QMessageBox, QWidget, QLabel, QComboBox, QLineEdit, QGridLayout, QTabWidget, \
    QMainWindow, QPushButton
from PyQt5.QtCore import Qt
from datetime import datetime
from data_base.work_with_base import connect_to_db

from openpyxl.reader.excel import load_workbook


class TabPage_normir(QWidget):
    def __init__(self, work_plan, tableWidget):
        super().__init__()
        self.validator_int = QIntValidator(0, 8000)

        self.tableWidget = tableWidget

        self.work_plan = work_plan

        self.well_number_label = QLabel('номер скважины')
        self.well_number_edit = QLineEdit(self)
        self.well_number_edit.setValidator(self.validator_int)

        self.well_area_label = QLabel('площадь скважины')
        self.well_area_edit = QLineEdit(self)

        self.repair_summary_label = QLabel('Сводка ремонта')
        self.repair_summary_line = QLineEdit(self)

        self.button = QPushButton("Загрузить сводку")
        self.button.clicked.connect(self.load_repair_summary)

        self.date_begin_label = QLabel('Дата начала ремонта')
        self.date_begin_datetime = QLineEdit(self)

        self.date_cancel_label = QLabel('Дата окончания ремонта')
        self.date_end_datetime = QLineEdit(self)

        self.brigade_number_label = QLabel('Номер бригады')
        self.brigade_number_line = QLineEdit(self)

        self.lifting_unit_label = QLabel('Подьемный агрегат')
        self.lifting_unit_combo = QComboBox(self)
        lifting_list = ['', 'СУРС-40', 'АЗИНМАШ-37А (Оснастка 2×3)', 'АПРС-32 (Оснастка 2×3)',
                         'АПРС-40 (Оснастка 3×4)', 'АПРС-40 (Оснастка 2×3)', 'АПРС-50 (Оснастка 3×4)',
                        'АПР60/80 (Оснастка 3×4)', 'УПА-60/80 (Оснастка 3×4)', 'УПТ-32 (Оснастка 3×4)',
                        'БАРС 60/80 (Оснастка 3×4)']

        self.lifting_unit_combo.addItems(lifting_list)

        self.type_kr_label = QLabel('Вид ГТМ (плановый)')
        self.type_kr_combo = QComboBox(self)
        self.type_kr_combo.addItems([x.split(' ')[0] for x in well_data.type_kr_list])

        self.type_kr_after = QLabel('Вид ГТМ (фактический)')
        self.type_kr_combo_after = QComboBox(self)
        self.type_kr_combo_after.addItems([x.split(' ')[0] for x in well_data.type_kr_list])

        self.current_after_label = QLabel('Забой после ремонта')
        self.current_after_line = QLineEdit(self)

        self.surname_master_label = QLabel('Фамилия мастера')
        self.surname_master_line = QLineEdit(self)

        self.availabity_of_des_label = QLabel('Наличие ДЭС')
        self.availabity_of_des_combo = QComboBox(self)
        self.availabity_of_des_combo.addItems(['нет', 'да'])

        self.commissioning_at_repair_label = QLabel('запуск в работу при бригаде в конце ремонта')
        self.commissioning_at_repair_combo = QComboBox(self)
        self.commissioning_at_repair_combo.addItems(['Да', 'Нет'])

        self.table_name = ''

        self.grid = QGridLayout(self)

        self.grid.addWidget(self.well_number_label, 2, 1)
        self.grid.addWidget(self.well_number_edit, 3, 1)

        self.grid.addWidget(self.well_area_label, 2, 2)
        self.grid.addWidget(self.well_area_edit, 3, 2)

        self.grid.addWidget(self.type_kr_label, 4, 1)
        self.grid.addWidget(self.type_kr_combo, 5, 1)

        self.grid.addWidget(self.type_kr_after, 4, 2)
        self.grid.addWidget(self.type_kr_combo_after, 5, 2)

        self.grid.addWidget(self.current_after_label, 4, 3)
        self.grid.addWidget(self.current_after_line, 5, 3)

        self.grid.addWidget(self.commissioning_at_repair_label, 4, 4)
        self.grid.addWidget(self.commissioning_at_repair_combo, 5, 4)

        self.grid.addWidget(self.button, 10, 1, 1, 2)
        self.grid.addWidget(self.repair_summary_label, 11, 1, 1, 5)
        self.grid.addWidget(self.repair_summary_line, 12, 1, 1, 5)

        self.grid.addWidget(self.brigade_number_label, 8, 1)
        self.grid.addWidget(self.brigade_number_line, 9, 1)

        self.grid.addWidget(self.surname_master_label, 13, 1)
        self.grid.addWidget(self.surname_master_line, 14, 1)

        self.grid.addWidget(self.availabity_of_des_label, 13, 2)
        self.grid.addWidget(self.availabity_of_des_combo, 14, 2)

        self.grid.addWidget(self.date_begin_label, 8, 2)
        self.grid.addWidget(self.date_begin_datetime, 9, 2)

        self.grid.addWidget(self.date_cancel_label, 8, 3)
        self.grid.addWidget(self.date_end_datetime, 9, 3)

        self.grid.addWidget(self.lifting_unit_label, 8, 4)
        self.grid.addWidget(self.lifting_unit_combo, 9, 4)

        self.well_area_edit.setText(f"{well_data.well_area._value}")
        # self.well_area_edit.textChanged.connect(self.update_well)
        self.well_number_edit.editingFinished.connect(self.update_well)

        if well_data.work_plan not in ['dop_plan_in_base']:
            self.well_number_edit.setText(f'{well_data.well_number._value}')

        if well_data.data_in_base:
            self.table_in_base_label = QLabel('данные по скважине')
            self.table_in_base_combo = QComboBox()
            self.table_in_base_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)

            self.well_data_label = QLabel('файл excel')
            self.well_data_in_base_combo = QComboBox()
            self.well_data_in_base_combo.setSizeAdjustPolicy(QComboBox.SizeAdjustPolicy.AdjustToContents)

            table_list = self.get_tables_starting_with(self.well_number_edit.text(), self.well_area_edit.text())
            if table_list:
                self.table_in_base_combo.clear()
                self.table_in_base_combo.addItems(table_list[::-1])

            self.grid.addWidget(self.table_in_base_label, 2, 5)
            self.grid.addWidget(self.table_in_base_combo, 3, 5)
            self.grid.addWidget(self.well_data_label, 2, 6)
            self.grid.addWidget(self.well_data_in_base_combo, 3, 6)

            self.table_in_base_combo.currentTextChanged.connect(self.update_area)

    def load_repair_summary(self):
        self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Выберите файл', '.',
                                                         "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)")
        self.repair_summary_line.setText(self.fname)
        # if well_data.well_area._value in fname:

        index_type = 0
        a = well_data.type_kr.split(' ')[0]

        for index, kr in enumerate(well_data.type_kr_list):
            if well_data.type_kr.split(' ')[0] in kr:
                index_type = index

        if index_type != 0:
            self.type_kr_combo.setCurrentIndex(index_type)
            self.type_kr_combo_after.setCurrentIndex(index_type)

        self.type_kr_combo.currentText()

        if self.fname:
            wb_summary = load_workbook(self.fname, data_only=True)
            number_brigade = self.fname[self.fname.index('№') + 1:self.fname.index('(')].strip()
            self.brigade_number_line.setText(number_brigade)
            ws_summary = wb_summary.active
            work_list_in_ois = []
            for row_ind, row in enumerate(ws_summary.iter_rows(values_only=True, max_col=4)):
                work_list_in_ois.append(row)

                if row_ind == 1:
                    date_begin = row[0].split('\n')[0]
                    if '>' in row[1]:
                        time_begin = row[1].split('>')[0]
                        if '-' in time_begin:
                            date_begin = date_begin + ' ' + time_begin.split('-')[0]
                elif row_ind != 1:
                    date_cancel = row[0].split('\n')[0]
                    if '>' in row[1]:
                        time_cancel = row[1].split('>')[0]
                        date_cancel = date_cancel + ' ' + time_cancel.split('-')[1]

                    value_str = row[1].upper().strip()
                    if 'СУРС-40' in value_str:
                        unit_lifting = 'СУРС-40'
                        self.lifting_unit_combo.setCurrentIndex(1)
                    elif 'ДЭС' in value_str:
                        self.availabity_of_des_combo.setCurrentIndex(1)

                    elif 'АЗИНМАШ' in value_str:
                        unit_lifting = 'АЗИНМАШ-37А'
                        self.lifting_unit_combo.setCurrentIndex(2)
                    elif 'АПРС-32' in value_str:
                        unit_lifting = 'АПРС-32'
                        self.lifting_unit_combo.setCurrentIndex(3)
                    elif 'АПРС-40' in value_str:
                        unit_lifting = 'АПРС-40'
                        self.lifting_unit_combo.setCurrentIndex(4)
                    elif 'АПРС-50' in value_str:
                        unit_lifting = 'АПРС-50'
                        self.lifting_unit_combo.setCurrentIndex(5)
                    elif 'АПР60/80' in value_str or 'АПР-60/80' in value_str:
                        unit_lifting = 'АПР60/80'
                        self.lifting_unit_combo.setCurrentIndex(7)
                    elif 'УПА-60/80' in value_str:
                        unit_lifting = 'УПА-60/80'
                        self.lifting_unit_combo.setCurrentIndex(8)
                    elif 'БАРС60/80' in value_str or 'БАРС-60/80' in value_str:
                        unit_lifting = 'БАРС60/80'
                        self.lifting_unit_combo.setCurrentIndex(10)
            well_data.work_list_in_ois = work_list_in_ois
            self.date_begin_datetime.setText(date_begin)
            self.date_end_datetime.setText(date_cancel)

    def update_area(self):
        table_in_base_combo = self.table_in_base_combo.currentText()
        if len(table_in_base_combo.split(" ")) > 1:
            well_area = table_in_base_combo.split(" ")[1]
            self.well_area_edit.setText(well_area)
            self.well_number_edit.setText(table_in_base_combo.split(" ")[0])

            well_list = self.check_in_database_well_data(self.well_number_edit.text())
            if well_list:
                self.well_data_in_base_combo.clear()
                self.well_data_in_base_combo.addItems(well_list)

    def check_in_database_well_data(self, number_well):
        table_in_base_combo = self.table_in_base_combo.currentText()
        if ' от' in table_in_base_combo:
            table_in_base_combo = table_in_base_combo[:-14]

        if number_well and len(table_in_base_combo.split(" ")) > 3:
            well_number, well_area = table_in_base_combo.split(" ")[:2]
            self.well_number_edit.setText(well_number)
            self.well_area_edit.setText(well_area)

            if well_area != ' ':
                if well_data.connect_in_base:
                    try:
                        conn = psycopg2.connect(**well_data.postgres_params_data_well)
                        cursor = conn.cursor()

                        # Запрос для извлечения всех скважин с наличием данных
                        cursor.execute(
                            "SELECT well_number, area_well, contractor, costumer, today, work_plan FROM wells "
                            "WHERE well_number=(%s) AND area_well=(%s)",
                            (str(number_well), well_area))


                    except psycopg2.Error as e:
                        # Выведите сообщение об ошибке
                        mes = QMessageBox.warning(None, 'Ошибка',
                                                  f'Ошибка подключения к базе данных, Скважина не добавлена в базу: \n {type(e).__name__}\n\n{str(e)}')
                else:
                    try:
                        db_path = connect_to_db('well_data.db', 'data_base_well/')

                        conn = sqlite3.connect(f'{db_path}')
                        cursor = conn.cursor()

                        cursor.execute(
                            "SELECT  well_number, area_well, contractor, costumer, today, work_plan FROM wells "
                            "WHERE well_number = ? AND area_well = ? "
                            "AND contractor = ? AND costumer = ?",
                            (str(well_number), well_area, well_data.contractor, well_data.costumer))



                    except sqlite3.Error as e:
                        # Выведите сообщение об ошибке
                        mes = QMessageBox.warning(None, 'Ошибка',
                                                  f'Ошибка подключения к базе данных, Скважина не добавлена '
                                                  f'в базу: \n {type(e).__name__}\n\n{str(e)}')
                # Получение всех результатов
                wells_with_data = cursor.fetchall()
                # Проверка, есть ли данные
                if wells_with_data:
                    well_list = []
                    for well in wells_with_data:
                        try:
                            if 'Ойл' in well[2]:
                                contractor = 'Ойл'
                            elif 'РН' in well[2]:
                                contractor = 'РН'
                        except:
                            contractor = 'Ойл'
                        date_string = well[4]
                        try:
                            # Преобразуем строку в объект datetime
                            datetime_object = datetime.strptime(date_string, "%Y-%m-%d %H:%M:%S.%f")

                            # Форматируем объект datetime в нужный формат
                            formatted_date = datetime_object.strftime("%d.%m.%Y")
                        except:
                            formatted_date = well[4]

                        # Формируем список скважин
                        well_list.append(f'{well[0]} {well[1]} {contractor} {well[5]} от {formatted_date}')

                        self.grid.setColumnMinimumWidth(5, self.table_in_base_combo.sizeHint().width())
                        self.grid.setColumnMinimumWidth(6, self.well_data_in_base_combo.sizeHint().width())

                    return well_list[::-1]
                else:
                    return False

    def update_well(self):

        self.table_name = str(self.well_number_edit.text()) + self.well_area_edit.text()
        if well_data.data_in_base:

            table_list = self.get_tables_starting_with(self.well_number_edit.text(), self.well_area_edit.text())

            if table_list:
                table_list = table_list[::-1]
                self.table_in_base_combo.clear()
                self.table_in_base_combo.addItems(table_list)

            well_list = self.check_in_database_well_data(self.well_number_edit.text())
            if well_list:
                self.well_data_in_base_combo.clear()
                self.well_data_in_base_combo.addItems(well_list)

    def update_table_in_base_combo(self):

        number_dp = self.number_DP_Combo.currentText()

        table_in_base_combo = self.table_in_base_combo.currentText()
        if ' от' in table_in_base_combo:
            table_in_base_combo = table_in_base_combo[:-14]

        well_number, well_area = table_in_base_combo.split(" ")[:2]
        self.well_number_edit.setText(well_number)
        self.well_area_edit.setText(well_area)

    def get_tables_starting_with(self, well_number, well_area):
        from data_base.work_with_base import connect_to_db, get_table_creation_time
        """
        Возвращает список таблиц, имена которых начинаются с заданного префикса.
        """
        prefix = well_number
        if 'Ойл' in well_data.contractor:
            contractor = 'ОЙЛ'
        elif 'РН' in well_data.contractor:
            contractor = 'РН'

        if prefix != '':
            if well_data.connect_in_base:
                conn = psycopg2.connect(**well_data.postgres_conn_work_well)
                cursor = conn.cursor()
                cursor.execute("""
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                AND table_name LIKE %s
                """, (prefix + '%',))
                tables = []

                for row in cursor.fetchall():
                    data_in = get_table_creation_time(conn, row[0])

                    tables.append(row[0] + data_in)
                # tables.insert(0, '')

                cursor.close()

                tables_filter = list(filter(lambda x: contractor in x, tables))
                if len(tables_filter) == 0:
                    tables_filter = tables.insert(0, ' ')
                try:
                    tables_filter = tables_filter[::-1]
                    return tables_filter
                except:
                    return
            else:
                try:
                    # Формируем полный путь к файлу базы данных
                    db_path = connect_to_db('databaseWell.db', 'data_base_well')
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()

                    cursor.execute("""SELECT name FROM sqlite_master WHERE type='table'""")

                    tables = []
                    table_in_base = cursor.fetchall()[1:]

                    for table_name in table_in_base:
                        if prefix in table_name[0].split(' ')[0]:
                            data_in = get_table_creation_time(conn, table_name[0])
                            tables.append(table_name[0] + data_in)
                    # tables.insert(0, '')

                    # Фильтруем таблицы по префиксу и подрядчику
                    tables_filter = list(filter(lambda x: contractor in x, tables))

                    # Сортируем таблицы в обратном порядке
                    tables_filter = tables_filter[::-1]

                    return tables_filter

                except sqlite3.Error as e:
                    print(f"Ошибка получения списка таблиц: {type(e).__name__}\n\n{str(e)}")
                finally:
                    if cursor:
                        cursor.close()
                    if conn:
                        conn.close()
        else:
            return []


class TabWidget(QTabWidget):
    def __init__(self, work_plan, tableWidget=0, old_index=0):
        super().__init__()
        self.addTab(TabPage_normir(work_plan, tableWidget), 'Нормирование')


class NormirWindow(QMainWindow):

    def __init__(self, table_widget, work_plan, ws=None, parent=None):
        super(NormirWindow, self).__init__()
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        # self.ins_ind = ins_ind
        # self.table_widget = table_widget
        self.work_plan = work_plan
        self.dict_perforation = []
        self.date_begin_datetime = ''
        self.date_end_datetime = ''
        self.brigade_number_line = ''
        self.type_kr_combo_after = ''
        self.lifting_unit_combo = ''
        self.type_kr_combo = ''
        self.availabity_of_des_combo = ''
        self.table_widget = table_widget
        self.current_after = ''
        self.commissioning_at_repair_combo = ''
        self.repair_summary_line = ''

        self.tabWidget = TabWidget(self.work_plan)

        self.buttonadd_work = QPushButton('Загрузить план работ')
        self.buttonadd_work.clicked.connect(self.add_work, Qt.QueuedConnection)

        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tabWidget, 0, 0, 1, 2)

        vbox.addWidget(self.buttonadd_work, 3, 0, 1, 2)

    def add_work(self):
        from data_base.work_with_base import check_in_database_well_data, insert_data_well_dop_plan, round_cell

        current_widget = self.tabWidget.currentWidget()

        well_number = current_widget.well_number_edit.text()
        well_area = current_widget.well_area_edit.text()
        self.repair_summary_line = current_widget.repair_summary_line.text()

        self.commissioning_at_repair_combo = current_widget.commissioning_at_repair_combo.currentText()
        self.type_kr_combo = current_widget.type_kr_combo.currentText()
        self.type_kr_combo_after = current_widget.type_kr_combo_after.currentText()
        self.current_after = current_widget.current_after_line.text()
        self.date_begin_datetime = current_widget.date_begin_datetime.text()
        self.date_end_datetime = current_widget.date_end_datetime.text()
        self.brigade_number_line = current_widget.brigade_number_line.text()
        self.brigade_number_line = current_widget.brigade_number_line.text()
        self.lifting_unit_combo = current_widget.lifting_unit_combo.currentText()
        self.availabity_of_des_combo = current_widget.availabity_of_des_combo.currentText()
        self.surname_master = current_widget.surname_master_line.text()
        well_data.lifting_unit_combo = self.lifting_unit_combo


        if well_number not in self.repair_summary_line:
            QMessageBox.warning(self, 'Ошибка', 'Номер скважины не совпадает с названием сводки')
            return
        well_data.date_work = self.date_begin_datetime.split(' ')[0]
        work_list = self.add_work_excel()
        self.populate_normir(work_list, self.table_widget)

        well_data.pause = False
        self.close()

    def populate_normir(self, work_list, table_widget):
        for row_index, row_data in enumerate(work_list):
            # table_widget.insertRow(row_index)
            for column, data in enumerate(row_data):
                item = QtWidgets.QTableWidgetItem(str(data))
                item.setFlags(item.flags() | Qt.ItemIsEditable)
                if not data is None:
                    table_widget.setItem(row_index, column, item)
                else:
                    table_widget.setItem(row_index, column, QtWidgets.QTableWidgetItem(str('')))

    def delete_data(self, number_well, area_well, work_plan):
        if well_data.connect_in_base:
            try:
                conn = psycopg2.connect(**well_data.postgres_params_data_well)
                cursor = conn.cursor()

                cursor.execute("""
                DELETE FROM wells 
                WHERE well_number = %s AND area_well = %s AND contractor = %s AND costumer = %s AND work_plan= %s """,
                               (str(number_well), area_well, well_data.contractor, well_data.costumer, work_plan)
                               )

                conn.commit()
                cursor.close()
                conn.close()

            except psycopg2.Error as e:
                # Выведите сообщение об ошибке
                mes = QMessageBox.warning(None, 'Ошибка',
                                          f'Ошибка удаления {type(e).__name__}\n\n{str(e)}')
        else:
            try:
                db_path = connect_to_db('well_data.db', 'data_base_well/')

                conn = sqlite3.connect(f'{db_path}')
                cursor = conn.cursor()

                cursor.execute("DELETE FROM wells  WHERE well_number = ? AND area_well = ? "
                               "AND contractor = ? AND costumer = ? AND work_plan=?",
                               (str(number_well._value), area_well._value, well_data.contractor, well_data.costumer,
                                work_plan))

                conn.commit()
                cursor.close()
                conn.close()

            except sqlite3.Error as e:
                # Выведите сообщение об ошибке
                mes = QMessageBox.warning(None, 'Ошибка',
                                          f'Ошибка удаления {type(e).__name__}\n\n{str(e)}')

    def add_work_excel(self):

        if '40' in self.lifting_unit_combo:
            lifting = 40
        elif '50' in self.lifting_unit_combo:
            lifting = 50
        elif '60' in self.lifting_unit_combo:
            lifting = 60
        elif '80' in self.lifting_unit_combo:
            lifting = 80

        region = {
            'ТГМ': 'Туймазинский регион',
            'КГМ': 'Краснохолмский регион',
            'ЧГМ': 'Чекмагушевский регион',
            'АГМ': 'Арланский регион',
            'ИГМ': 'Ишимбайский регион'
        }
        if well_data.curator == 'КГМ':
            region = 'Краснохолмский регион'

        work_list = [
            [None, None, ' АКТ                                                       ', None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'v.2024.2/2', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, 'проверяет ЦДНГ'],
            [None, None, 'на сдачу скважины из капитального ремонта', None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'бр/час руб.', None,
             '№ договора',
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, '9609,18 - КРС; 9128,72 - ПРС; ', None, 'БНД/у/8/53/24/БУР',
             None, None],
            [None, 'Регион', None, region[well_data.region], None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, 'Исполнитель: ', None, None, f'{well_data.contractor}', None, None, None, None, None, None,
             None,
             None, None,
             None],
            [None, 'ЦДНГ', None, well_data.cdng._value, None, 'Начало работ:', None, self.date_begin_datetime, None,
             'уникальный номер ремонта', None, None, 'причины разделения', None, 'примечание', None, '№ Скважина', None,
             'Бригада:',
             None, self.brigade_number_line, None, 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', None, 'оттяжки ПА',
             'макс, в тн (с оттяжками)', lifting,
             'Подъемник по нормативу на СПО', None, None],
            [None, 'Месторождение:', None, well_data.well_oilfield._value, None, 'Окончание работ:', None,
             self.date_end_datetime,
             None, '222222222222222222', None, None, None, None, None, None, '=D11', None, 'Мастер:', None,
             self.surname_master, None,
             'ГКШ-1500 / ГКШ 300', '111', 'АПРС-40', None, None, None, None, self.lifting_unit_combo, None, None],
            [None, 'Площадь:', None, well_data.well_area._value, None, 'Начало работ:', None, None, None,
             'уникальный номер ремонта', None,
             None, 'причины объединения', None, 'примечание', None, '№ Скважина', None, 'Бригада:', None,
             self.brigade_number_line, None,
             'Вид ГКШ',
             'гос № ПА', 'Подъемник по факту', None, None, 'макс в тн (РЭ без оттяжек)', lifting,
             'Подъемник по нормативу на СПО',
             None, None],
            [None, 'Куст:', None, None, None, 'Окончание работ:', None, None, None, None, None, None, None, None, None,
             None, None,
             None, 'Мастер:', None, None, None, 'ГКШ-1500 / ГКШ 300', None, None, None, None, None, None,
             None, None,
             None],
            [None, '№ Скважины:', None, well_data.well_number._value, None, 'Начало работ:', None, None, None,
             'уникальный номер ремонта', None,
             None,
             'причины объединения', None, 'примечание', None, '№ Скважина', None, 'Бригада:', None, None, None,
             'Вид ГКШ',
             'гос № ПА', 'Подъемник по факту', None, None, None, None, 'Подъемник по нормативу на СПО', None, None],
            [None, 'Признак отказности по отчётности ОРМФ', None, None, 'нет', 'Окончание работ:', None, None, None,
             None, None,
             None, None, None, None, None, None, None, 'Мастер:', None, None, None, None, None, None, None, None, None,
             None, None,
             None, None],
            [None, 'Причины отказов', None, None, None, 'Начало работ:', None, None, None, 'уникальный номер ремонта',
             None, None,
             'причины объединения', None, 'примечание', None, '№ Скважина', None, 'Бригада:', None, None, None,
             'Вид ГКШ',
             'гос № ПА', 'Подъемник по факту', None, None, None, None, 'Подъемник по нормативу на СПО', None, None],
            [None, None, None, None, None, 'Окончание работ:', None, None, None, None, None, None, None, None, None,
             None, None,
             None, 'Мастер:', None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, 'Начало работ:', None, None, None, 'уникальный номер ремонта', None, None,
             'причины объединения', None, 'примечание', None, '№ Скважина', None, 'Бригада:', None, None, None,
             'Вид ГКШ',
             'гос № ПА', 'Подъемник по факту', None, None, None, None, 'Подъемник по нормативу на СПО', None, None],
            [None, 'Дата отказа', None, None, None, 'Окончание работ:', None, None, None, None, None, None, None, None,
             None, None,
             None, None, 'Мастер:', None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Дата окон. предыдущего ремонта', None, None, None, 'Начало работ:', None, None, None,
             'уникальный номер ремонта', None, None, 'причины объединения', None, 'примечание', None, '№ Скважина',
             None,
             'Бригада:', None, None, None, 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', None, None, 'факт макс. в тн',
             '20',
             'Подъемник по нормативу на СПО', None, None],
            [None, 'ННО', None, '=IF((D16-D17)<=0,"-",(D16-D17))', None, 'Окончание работ:', None, None, None, None,
             None, None,
             None, None, None, None, None, None, 'Мастер:', None, None, None, None, None, None, None, None, None, None,
             None, None,
             None],
            [None, '=IF(D18<365,"ДК","-")', None, None, None, 'Время  календ.:', None,
             '=(H8-H7)*24+(H10-H9)*24+(H12-H11)*24+(H14-H13)*24+(H16-H15)*24+(H18-H17)*24', None, 'оплата по УИН:',
             None, None,
             '=J8', None, '№ Скважина', None, '=D11', None, 'Дизельная эл.станция', None, None, None,
             self.availabity_of_des_combo,
             'Работа МЭС в часах',
             None, None, None, '=IF(W19="да",J1292-J1287,"0")', None, None, None, None],
            [None, 'запуск в работу при бригаде в конце ремонта', None, None, None, 'Категория скважины', None, None,
             '=MIN(G24:K24)', 'ЗАВЕРШЕНИЕ РЕМОНТА', None, None, 'да', None, 'инвентарный №', None,
             well_data.inv_number._value, None, None,
             None, None,
             None, None, 'Ст-ть часа работы МЭС, руб.', None, None, None, None, None, None, None, None],
            [None, self.commissioning_at_repair_combo, None, None, None, 'ПЛАН', 'по Pпл', None, 'по H2S',
             'по газовому фактору', None,
             'ВИД РЕМОНТА ', None,
             '=VLOOKUP(X21,ЦИКЛ!AE3:AX245,2,0)', None, None, None, None, None, None, None, None, 'Шифр \nКР/ТР',
             self.type_kr_combo,
             'Вид ГТМ (плановый)', None, 'прочие', None, 'Вид фонда', well_data.appointment, None, None],
            [None, 'Причины не запуска', None, None, None, None, well_data.category_pressuar, None,
             well_data.category_h2s, well_data.category_gf, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, 'ФАКТ', 'по Pпл', None, 'по H2S', 'по газовому фактору', None, None, None,
             '=VLOOKUP(X23,ЦИКЛ!AE3:AX245,2,0)', None, None, None, None, None, None, None, None, None,
             self.type_kr_combo_after,
             'Вид ГТМ (фактический)', None, 'прочие', None, None, well_data.appointment, None, None],
            [None, 'Примечание', None, None, None, None, well_data.category_pressuar, None, well_data.category_h2s,
             well_data.category_gf, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, 'Изменение категории скважины с пункта', None, None, None, None, None,
             'ВИД РЕМОНТА  \nпри ТР 4-8', None, None, None, None, None, None, None, None, None, None, 'Хоз.процессы',
             None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'О С Н О В Н Ы Е    Д А Н Н Ы Е', None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ДАННЫЕ ПО СКВАЖИНЕ', None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Искусственный забой', None, None, None, well_data.bottomhole_artificial._value, None, None,
             'данные по УА до ремонта', None, None,
             None, None,
             None, 'данные по УА после ремонта', None, None, None, None, None, None, None,
             None, None, None, None, None,
             None, None,
             None, None, None],
            [None, 'Текущий забой до ремонта', None, None, None, well_data.current_bottom, None, None, 'типоразмер',
             None, None, 'сост.',
             'зав.№', None,
             'типоразмер', None, None, 'сост.', 'зав.№', None, None, None, None, None, None, None, None, None, None,
             None, None,
             None],
            [None, 'Текущий забой после ремонта', None, None, None, self.current_after, None, None,
             well_data.wellhead_fittings, None,
             None, 'удовл',
             None, None, well_data.wellhead_fittings, None, None, 'удовл', None, None, None, None, None, None, None,
             None,
             None, None,
             None, None, None, None],
            [None, 'П О Д З Е М Н О Е  О Б О Р У Д О В А Н И Е', None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Тип поднятого оборудования, марка, мощность напора /\n номер насоса', None, None, None, None, None,
             'ЭЦН', None,
             None, None, None, 'Тип спущенного оборудования, марка, мощность напора / \nномер насоса', None, None, None,
             'ШГН',
             None, None, None, 'Кол-во крепежных поясов, шт.', None, None, None, None, None, None, 'Переводник, №',
             None, None,
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             'Тип протекторов', None, None, None, None, None, None, 'Переводник, №', None, None, None, None],
            [None, 'Глубина спуска, м.', None, None, None, None, None, None, None, None, None, None,
             'Глубина спуска, м.', None,
             None, None, None, None, None, None, 'Кол-во протекто-ров', None, '60мм', None, None, None, None,
             'Подвес. патрубок,№',
             None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, '73мм', None, None, None, None, 'Противополетка', None, None, None, None],
            [None, 'Доп.оборудование поднято ', None, None, None, None, None, None, None, None, None, None,
             'Длина спущенного кабеля, м.', None, None, None, None, None, None, None, 'Тип сбивного клапана', None,
             None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             'Тип обратного клапана', None, None, None, None, None, None, None, None, None, None, None],
            [None, ' интервал установки, м', None, None, None, None, None, None, None, None, None, None,
             'Доп.оборудование   после ТКРС                                                         интервал установки',
             None, None,
             None, None, None, None, None,
             'Доп.оборудование   после ТКРС                                                         интервал установки',
             None, None,
             None, None, None, None, 'противополётное оборудование', None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            ['О б ъ е м    в ы п о л н е н н ы х    р а б о т', None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None],
            ['№', 'Дата', None, 'Цикл ремонта', 'Разбивка цикла ремонта', 'Вид работ', None, None, None, None, None,
             None, None,
             None, 'Акт', 'Химия', 'реагенты', 'тн/шт', 'Источник норм: ', 'Простои/глушение', 'Ед.изм.', 'Кол-во',
             'Норма на\nедин.(час)', 'Коэф.', 'Время по норме, \nчас.', 'Время к оплате', 'Исключенное Время', None,
             None, None,
             None, 'ПРИМЕЧАНИЕ '],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, 'ОТКРСиО', 'УСРСиСТ', 'ПТО РУДНГ', 'ОГТМ/ОР', 'Причина снятия', None],
            ['Объём работ по основному плану', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None]]
        well_data.brigade_number = self.brigade_number_line
        well_data.date_end = self.date_end_datetime.split(' ')[0]
        return work_list


def extraction_data(self, table_name, paragraph_row=0):
    from data_base.work_with_base import connect_to_db

    if well_data.connect_in_base:
        try:
            # Устанавливаем соединение с базой данных
            conn1 = psycopg2.connect(**well_data.postgres_conn_work_well)
            cursor1 = conn1.cursor()

            # Проверяем наличие таблицы с определенным именем
            result_table = 0

            if well_data.work_plan in ['krs', 'plan_change']:
                # print(cursor1.execute(f"SELECT EXISTS (SELECT FROM information_schema.tables").fetchall())
                cursor1.execute(
                    f"SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = '{table_name}')")
                result_table = cursor1.fetchone()

            elif well_data.work_plan in ['dop_plan', 'dop_plan_in_base']:

                cursor1.execute(f"SELECT EXISTS (SELECT FROM information_schema.tables "
                                f"WHERE table_name = '{table_name}')")

                result_table = cursor1.fetchone()

            if result_table[0]:
                well_data.data_in_base = True
                cursor2 = conn1.cursor()

                cursor2.execute(f'SELECT * FROM "{table_name}"')
                result = cursor2.fetchall()

                well_data.data_well_is_True = True

            else:
                well_data.data_in_base = False
                mes = QMessageBox.warning(self, 'Проверка наличия таблицы в базе данных',
                                          f"Таблицы '{table_name}' нет в базе данных.")


        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
            QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных,')
        finally:
            # Закройте курсор и соединение
            if cursor1:
                cursor1.close()
            if conn1:
                conn1.close()
    else:
        try:
            # Формируем полный путь к файлу базы данных
            db_path = connect_to_db('databaseWell.db', 'data_base_well')
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()
            result_table = 0

            if well_data.work_plan in ['krs', 'plan_change']:
                cursor.execute(
                    f"SELECT name FROM sqlite_master WHERE type='table' AND name=? ",
                    (table_name,))
                result_table = cursor.fetchone()

            elif well_data.work_plan in ['dop_plan', 'dop_plan_in_base']:
                cursor.execute(
                    f"SELECT name FROM sqlite_master WHERE type='table' AND name=?",
                    (table_name,))
                result_table = cursor.fetchone()

            if result_table:
                well_data.data_in_base = True
                cursor2 = conn.cursor()

                cursor2.execute(f'SELECT * FROM "{table_name}"')
                result = cursor2.fetchall()

                well_data.data_well_is_True = True

            else:
                well_data.data_in_base = False
                mes = QMessageBox.warning(self, 'Проверка наличия таблицы в базе данных',
                                          f"Таблицы '{table_name}' нет в базе данных.")

        except sqlite3.Error as e:
            # Выведите сообщение об ошибке
            mes = QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных.')

        finally:
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()

    return


def insert_data_dop_plan(self, result, paragraph_row):
    try:
        paragraph_row = paragraph_row - 1
    except:
        paragraph_row = 1
    if len(result) < paragraph_row:
        mes = QMessageBox.warning(self, 'Ошибка', f'В плане работ только {len(result)} пункта')
        return

    well_data.current_bottom = result[paragraph_row][1]

    well_data.dict_perforation = json.loads(result[paragraph_row][2])
    well_data.plast_all = json.loads(result[paragraph_row][3])
    well_data.plast_work = json.loads(result[paragraph_row][4])
    well_data.leakage = json.loads(result[paragraph_row][5])
    if result[paragraph_row][6] == 'true':
        well_data.column_additional = True
    else:
        well_data.column_additional = False

    well_data.fluid_work = result[paragraph_row][7]

    well_data.category_pressuar = result[paragraph_row][8]
    well_data.category_h2s = result[paragraph_row][9]
    well_data.category_gf = result[paragraph_row][10]
    try:
        well_data.template_depth, well_data.template_lenght, well_data.template_depth_addition, \
        well_data.template_lenght_addition = json.loads(result[paragraph_row][11])
    except:
        well_data.template_depth = result[paragraph_row][11]
    well_data.skm_interval = json.loads(result[paragraph_row][12])
    well_data.problemWithEk_depth = result[paragraph_row][13]
    well_data.problemWithEk_diametr = result[paragraph_row][14]
    well_data.dict_perforation_short = json.loads(result[paragraph_row][2])


def insert_data_plan(self, result):
    well_data.data_list = []
    for row in result:
        data_list = []
        for index, data in enumerate(row[:-1]):
            if index == 6:
                if data == 'false' or data == 0 or data == '0':
                    data = False
                else:
                    data = True
            data_list.append(data)
        well_data.data_list.append(data_list)

    well_data.fluid_work_short = well_data.fluid_work_short


def work_list(self, work_earlier):
    krs_begin = [[None, None,
                  f' Ранее проведенные работ: \n {work_earlier}',
                  None, None, None, None, None, None, None,
                  'Мастер КРС', None]]

    return krs_begin


if __name__ == "__main__":
    # app3 = QApplication(sys.argv)

    app = QApplication(sys.argv)
    window = NormirWindow('normir', 2)
    window.show()
    sys.exit(app.exec_())
