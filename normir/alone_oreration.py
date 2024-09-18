import json
import logging
from collections import namedtuple
from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QInputDialog, QMessageBox
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import well_data


def kot_select(self, current_bottom):
    if well_data.column_additional is False \
            or (
            well_data.column_additional is True and well_data.current_bottom <= well_data.head_column_additional._value):
        kot_select = f'КОТ-50 (клапан обратный тарельчатый) +НКТ{well_data.nkt_diam}мм 10м + репер '

    elif well_data.column_additional is True and well_data.column_additional_diametr._value < 110 and \
            current_bottom >= well_data.head_column_additional._value:
        kot_select = f'КОТ-50 (клапан обратный тарельчатый) +НКТ{60}мм 10м + репер + ' \
                     f'НКТ60мм L- {round(current_bottom - well_data.head_column_additional._value, 0)}м'
    elif well_data.column_additional is True and well_data.column_additional_diametr._value > 110 and \
            current_bottom >= well_data.head_column_additional._value:
        kot_select = f'КОТ-50 (клапан обратный тарельчатый) +НКТ{73}мм со снятыми фасками 10м + репер + ' \
                     f'НКТ{well_data.nkt_diam}мм со снятыми фасками' \
                     f' L- {round(current_bottom - well_data.head_column_additional._value, 0)}м'

    return kot_select


def check_h2s(self, plast=0, fluid_new=0, expected_pressure=0):
    if len(well_data.plast_project) != 0:
        if len(well_data.plast_project) != 0:
            plast = well_data.plast_project[0]
        else:
            plast, ok = QInputDialog.getText(self, 'выбор пласта для расчета ЖГС ', 'введите пласт для перфорации')
            well_data.plast_project.append(plast)
        try:
            fluid_new = list(well_data.dict_perforation_project[plast]['рабочая жидкость'])[0]
        except:
            fluid_new, ok = QInputDialog.getDouble(self, 'Новое значение удельного веса жидкости',
                                                   'Введите значение удельного веса жидкости', 1.02, 1, 1.72, 2)
        if len(well_data.dict_category) != 0:
            expected_pressure = well_data.dict_category[well_data.plast_project[0]]['по давлению'].data_pressuar
        else:
            expected_pressure, ok = QInputDialog.getDouble(self, 'Ожидаемое давление по пласту',
                                                           'Введите Ожидаемое давление по пласту', 0, 0, 300, 1)

    else:
        fluid_new, ok = QInputDialog.getDouble(self, 'Новое значение удельного веса жидкости',
                                               'Введите значение удельного веса жидкости', 1.02, 1, 1.72, 2)
        plast, ok = QInputDialog.getText(self, 'выбор пласта для расчета ЖГС ', 'введите пласт для перфорации')

        expected_pressure, ok = QInputDialog.getDouble(self, 'Новое значение удельного веса жидкости',
                                                       'Введите значение удельного веса жидкости', 1.02, 1, 1.72, 2)

    return fluid_new, plast, expected_pressure


def konte(self):
    konte_list = [
        [f'Скважина согласована на проведение работ по технологии контейнерно-канатных технологий', None,
         f'Скважина согласована на проведение работ по технологии контейнерно-канатных технологий по '
         f'технологическому плану Таграс-РС.'
         f'Вызвать геофизическую партию. Заявку оформить за 24 часов сутки через '
         f'геологическую службу {well_data.contractor}. '
         f'Произвести монтаж ПАРТИИ ГИС согласно утвержденной главным инженером от'
         f' {well_data.dict_contractor[well_data.contractor]["Дата ПВО"]}. Предварительно нужно заявить вставку №6',
         None, None, None, None, None, None, None,
         'мастер КРС', 1.25],
        [None, None, f'Произвести работы указанные в плане работ силами спец подрядчика, при выполнении '
                     f'из основного плана работ работы исключить',
         None, None, None, None, None, None, None,
         'мастер КРС', 12]
    ]
    return konte_list


def definition_Q(self):
    definition_Q_list = [
        [f'Насыщение 5м3 определение Q при 80-120атм', None,
         f'Произвести насыщение скважины до стабилизации давления закачки не менее 5м3. Опробовать  '
         f' на приемистость в трех режимах при Р=80-120атм в '
         f'присутствии представителя супервайзерской службы или подрядчика по РИР. '
         f'Составить акт. (Вызов представителя осуществлять телефонограммой за 12 часов, '
         f'с подтверждением за 2 часа до '
         f'начала работ). ',
         None, None, None, None, None, None, None,
         'мастер КРС', 0.17 + 0.2 + 0.2 + 0.2 + 0.15 + 0.52]]
    return definition_Q_list


def privyazkaNKT(self):
    priv_list = [[f'ГИС Привязка по ГК и ЛМ', None,
                  f'Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через ЦИТС {well_data.contractor}". '
                  f'Произвести  монтаж ПАРТИИ ГИС согласно схемы  №8а утвержденной главным инженером '
                  f'{well_data.dict_contractor[well_data.contractor]["Дата ПВО"]}г. '
                  f'ЗАДАЧА 2.8.1 Привязка технологического оборудования скважины',
                  None, None, None, None, None, None, None,
                  'Мастер КРС, подрядчик по ГИС', 4]]
    return priv_list

def definition_plast_work(self):
    # Определение работающих пластов
    plast_work = set()
    perforation_roof = well_data.current_bottom
    perforation_sole = 0

    for plast, value in well_data.dict_perforation.items():
        for interval in value['интервал']:

            if well_data.dict_perforation[plast]["отключение"] is False:
                plast_work.add(plast)
            roof = min(list(map(lambda x: x[0], list(well_data.dict_perforation[plast]['интервал']))))
            sole = max(list(map(lambda x: x[1], list(well_data.dict_perforation[plast]['интервал']))))
            well_data.dict_perforation[plast]["кровля"] = roof
            well_data.dict_perforation[plast]["подошва"] = sole
            if well_data.dict_perforation[plast]["отключение"] is False:

                if perforation_roof >= roof and well_data.current_bottom > roof:
                    perforation_roof = roof
                if perforation_sole < sole and well_data.current_bottom > sole:
                    perforation_sole = sole

    well_data.perforation_roof = perforation_roof
    well_data.perforation_sole = perforation_sole
    well_data.dict_perforation = dict(
        sorted(well_data.dict_perforation.items(), key=lambda item: (not item[1]['отключение'],
                                                                     item[0])))
    well_data.plast_all = list(well_data.dict_perforation.keys())
    well_data.plast_work = list(plast_work)

def definitionBottomGKLM(self):
    priv_list = [
        [f'Отбить забой по ГК и ЛМ', None,
         f'Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через ЦИТС {well_data.contractor}". '
         f'Произвести  монтаж ПАРТИИ ГИС согласно схемы  №8а утвержденной главным инженером '
         f'{well_data.dict_contractor[well_data.contractor]["Дата ПВО"]}г. '
         f'ЗАДАЧА 2.8.2 Отбить забой по ГК и ЛМ',
         None, None, None, None, None, None, None,
         'Мастер КРС, подрядчик по ГИС', 4]]
    return priv_list


def pressuar_gis(self):
    priv_list = [[f'Замер Рпл', None,
                  f'Вызвать геофизическую партию. Заявку оформить за 16 часов сутки через ЦИТС {well_data.contractor}". '
                  f'Произвести  монтаж ПАРТИИ ГИС согласно схемы  №8а утвержденной главным инженером '
                  f'{well_data.dict_contractor[well_data.contractor]["Дата ПВО"]}г. '
                  f'Произвести замер Рпл в течении 4часов. При необходимости согласовать с заказчиком смену категории',
                  None, None, None, None, None, None, None,
                  'Мастер КРС, подрядчик по ГИС', 8]]
    return priv_list


def pvo_cat1(self):
    if 'Ойл' in well_data.contractor:
        date_str = 'от 07.03.2024г'
    elif 'РН' in well_data.contractor:
        date_str = 'от 28.02.2024г'

    pvo_1 = f'Установить ПВО по схеме №2 утвержденной главным инженером {well_data.contractor} {date_str} ' \
            f'(тип плашечный сдвоенный ПШП-2ФТ-160х21Г Крестовина КР160х21Г, ' \
            f'задвижка ЗМС 65х21 (3шт), Шарового крана 1КШ-73х21, авар. трубы (патрубок НКТ73х7-7-Е, ' \
            f' (при необходимости произвести монтаж переводника' \
            f' П178х168 или П168 х 146 или ' \
            f'П178 х 146 в зависимости от типоразмера крестовины и колонной головки). Спустить и посадить ' \
            f'пакер на глубину 10м. Опрессовать ПВО (трубные плашки превентора) на Р-{well_data.max_admissible_pressure._value}атм ' \
            f'(на максимально допустимое давление опрессовки ' \
            f'эксплуатационной колонны в течении 30мин), сорвать и извлечь пакер. \n' \
            f'- Обеспечить о обогрев превентора, станции управления ПВО оборудовать теплоизоляционными ' \
            f'материалом в зимней период. \n Получить разрешение на производство работ в присутствии представителя ПФС'

    pvo_list = [
        [None, None,
         "На скважинах первой категории Подрядчик обязан пригласить представителя ПАСФ " \
         "для проверки качества м/ж и опрессовки ПВО, документации и выдачи разрешения на производство " \
         "работ по ремонту скважин. При обнаружении нарушений, которые могут повлечь за собой опасность" \
         " для жизни людей"
         " и/или возникновению ГНВП и ОФ, дальнейшие работы должны быть прекращены. Представитель "
         "ПАСФ приглашается за 24 часа до проведения "
         "проверки монтажа ПВО телефонограммой. произвести практическое обучение по команде ВЫБРОС. "
         "Пусковой комиссией составить акт готовности "
         "подъёмного агрегата для ремонта скважины.",
         None, None, None, None, None, None, None,
         'Мастер КРС', None],
        [f'монтаж ПВО по схеме № 2 c гидроПВО', None,
         pvo_1, None, None,
         None, None, None, None, None,
         'Мастер КРС, представ-ли ПАСФ и Заказчика, Пуск. ком', 4.67]]
    well_data.kat_pvo = 1
    return pvo_list


def update_fluid(index_plan, fluid_str, table_widget):
    row_index = index_plan - well_data.count_row_well
    aaa = well_data.data_list
    for index_row, data in enumerate(well_data.data_list):
        if index_row == row_index:
            fluid_str_old = well_data.data_list[index_row][7]
        if row_index <= index_row:
            aaad = well_data.data_list[index_row][7]
            if well_data.data_list[index_row][7] == fluid_str_old:
                well_data.data_list[index_row][7] = fluid_str
                ccc = well_data.data_list[index_row][7]
                for column in range(table_widget.columnCount()):
                    if column == 2 or column == 0:
                        row_change = index_row + well_data.count_row_well
                        value = table_widget.item(row_change, column).text()
                        if value != None or value != '':
                            if fluid_str_old in value:
                                new_value = value.replace(fluid_str_old, fluid_str)
                                new_value = QtWidgets.QTableWidgetItem(f'{new_value}')
                                table_widget.setItem(row_change, column, new_value)


def calculationFluidWork(vertical, pressure):
    if (isinstance(vertical, float) or isinstance(vertical, int)) and (
            isinstance(pressure, float) or isinstance(pressure, int)):

        # print(vertical, pressure)
        stockRatio = 0.1 if float(vertical) <= 1200 else 0.05

        fluidWork = round(float(str(pressure)) * (1 + stockRatio) / float(vertical) / 0.0981, 2)
        # print(fluidWork < 1.02 , (well_data.region == 'КГМ' or well_data.region == 'АГМ'))
        if fluidWork < 1.02 and (well_data.region == 'КГМ' or well_data.region == 'АГМ'):
            fluidWork = 1.02
        elif fluidWork < 1.02 and (
                well_data.region == 'ИГМ' or well_data.region == 'ТГМ' or well_data.region == 'ЧГМ'):
            fluidWork = 1.01

        return fluidWork
    else:
        return None


def pvo_gno(kat_pvo):
    if 'Ойл' in well_data.contractor:
        date_str = 'от 07.03.2024г'
    elif 'РН' in well_data.contractor:
        date_str = ''
    # print(f' ПВО {kat_pvo}')
    pvo_2 = f'Установить ПВО по схеме №2 утвержденной главным инженером {well_data.contractor} {date_str} (тип плашечный ' \
            f'сдвоенный ПШП-2ФТ-152х21) и посадить пакер. ' \
            f'Спустить пакер на глубину 10м. Опрессовать ПВО (трубные плашки превентора) и линии манифольда до концевых ' \
            f'задвижек на Р-{well_data.max_admissible_pressure._value}атм на максимально допустимое давление ' \
            f'опрессовки эксплуатационной колонны в течении ' \
            f'30мин), сорвать пакер. ' \
        # f'В случае невозможности опрессовки по ' \
    # f'результатам определения приемистости и по согласованию с заказчиком  опрессовать трубные плашки ПВО на ' \
    # f'давление поглощения, но не менее 30атм. '

    pvo_1 = f'Установить ПВО по схеме №2 утвержденной главным инженером {well_data.contractor} {date_str} ' \
            f'(тип плашечный сдвоенный ПШП-2ФТ-160х21Г Крестовина КР160х21Г, ' \
            f'задвижка ЗМС 65х21 (3шт), Шарового крана 1КШ-73х21, авар. трубы (патрубок НКТ73х7-7-Е, ' \
            f' (при необходимости произвести монтаж переводника' \
            f' П178х168 или П168 х 146 или ' \
            f'П178 х 146 в зависимости от типоразмера крестовины и колонной головки). Спустить и посадить ' \
            f'пакер на глубину 10м. Опрессовать ПВО (трубные плашки превентора) на ' \
            f'Р-{well_data.max_admissible_pressure._value}атм ' \
            f'(на максимально допустимое давление опрессовки ' \
            f'эксплуатационной колонны в течении 30мин), сорвать и извлечь пакер. Опрессовать ' \
            f'выкидную линию после концевых задвижек на ' \
            f'Р - 50 кгс/см2 (5 МПа) - для противовыбросового оборудования, рассчитанного на' \
            f'давление до 210 кгс/см2 ((21 МПа)\n' \
            f'- Обеспечить обогрев превентора и СУП в зимнее время . \n Получить разрешение на производство работ в ' \
            f'присутствии представителя ПФС'
    if kat_pvo == 1:
        return pvo_1, f'Монтаж ПВО по схеме №2 + ГидроПревентор'
    else:
        # print(pvo_2)
        return pvo_2, f'Монтаж ПВО по схеме №2'


def lifting_unit(self):
    aprs_40 = f'Установить подъёмный агрегат на устье не менее 40т.\n' \
              f' Пусковой комиссией составить акт готовности  подьемного агрегата и бригады для проведения ремонта скважины.' \
              f'ПРИМЕЧАНИЕ:  ПРИ ИСПОЛЬЗОВАНИИ ПОДЪЕМНОГО АГРЕТАТА АПРС-50, А5-40, АПРС-50 ДОПУСКАЕТСЯ РАБОТА БЕЗ ' \
              f'ПРИМЕНЕНИЯ ВЕТРОВЫХ ОТТЯЖЕК ПРИ НАГРУЗКАХ НЕ БОЛЕЕ 25ТН. ПРИ НЕОБХОДИМОСТИ УВЕЛИЧЕНИЯ НАГРУЗКИ ТРЕБУЕТСЯ ' \
              f'ОСНАСТИТЬ ПОДЪЕМНЫЙ АГРЕГАТ ВЕТРОВЫМИ ОТТЯЖКАМИ. ПРИ ЭТОМ МАКСИМАЛЬНУЮ НАГРУЗКА НЕ ДОЛЖНА ПРЕВЫШАТЬ 80% ОТ' \
              f' СТРАГИВАЮЩЕЙ НАГРУЗКИ НА НКТ.ПРИ ИСПОЛЬЗОВАНИИ ПОДЬЕМНОГО АГРЕГАТА  УПА-60/80, БАРС, А-50, АПР 60/80 ' \
              f'РАБОТАТЬ ТОЛЬКО С ПРИМЕНЕНИЕМ  ОТТЯЖЕК МАКСИМАЛЬНУЮ НАГРУЗКА НЕ ДОЛЖНА ПРЕВЫШАТЬ 80% ОТ СТРАГИВАЮЩЕЙ ' \
              f'НАГРУЗКИ НА НКТ. После монтажа подъёмника якоря ветровых оттяжек должны быть испытаны на нагрузки, ' \
              f'установленные инструкцией по эксплуатации завода - изготовителя в присутствии супервайзера Заказчика. ' \
              f'Составить акт готовности подъемного агрегата. Пусковой комиссией составить акт готовности  подьемного ' \
              f'агрегата и бригады для проведения ремонта скважины. Дальнейшие работы продолжить после проведения пусковой ' \
              f'комиссии заполнения пусковой документации. '
    upa_60 = f'Установить подъёмный агрегат на устье не менее 60т. Пусковой комиссией составить ' \
             f'акт готовности  подьемного агрегата и бригады для проведения ремонта скважины.'

    return upa_60 if well_data.bottomhole_artificial._value >= 2300 else aprs_40


def volume_vn_ek(current):
    if well_data.column_additional is False or well_data.column_additional is True and current < well_data.head_column_additional._value:
        volume = round(
            (well_data.column_diametr._value - 2 * well_data.column_wall_thickness._value) ** 2 * 3.14 / 4 / 1000, 2)
    else:
        volume = round(
            (well_data.column_additional_diametr._value - 2 * well_data.column_additional_wall_thickness._value
             ) ** 2 * 3.14 / 4 / 1000, 2)

    return round(volume, 1)


def volume_vn_nkt(dict_nkt):  # Внутренний объем одного погонного местра НКТ
    # print(dict_nkt)
    for nkt, lenght_nkt in dict_nkt.items():
        volume_vn_nkt = 0
        nkt = ''.join(c for c in str(nkt) if c.isdigit())
        if '60' in str(nkt):
            t_nkt = 5
            volume_vn_nkt += round(3.14 * (int(nkt) - 2 * t_nkt) ** 2 / 4000000 * lenght_nkt, 5)
        elif '73' in str(nkt):
            t_nkt = 5.5
            volume_vn_nkt += round(3.14 * (int(nkt) - 2 * t_nkt) ** 2 / 4000000 * lenght_nkt, 5)
        elif '89' in str(nkt):
            t_nkt = 6
            volume_vn_nkt += round(3.14 * (int(nkt) - 2 * t_nkt) ** 2 / 4000000 * lenght_nkt, 5)

        elif '48' in str(nkt):
            t_nkt = 4.5
            volume_vn_nkt += round(3.14 * (int(nkt) - 2 * t_nkt) ** 2 / 4000000 * lenght_nkt * 1.1, 5)

    return round(volume_vn_nkt, 1)


def volume_rod(self, dict_sucker_rod):  # Объем штанг

    from find import FindIndexPZ
    volume_rod = 0
    # print(dict_sucker_rod)
    for diam_rod, lenght_rod in dict_sucker_rod.items():
        if diam_rod:
            volume_rod += (3.14 * (lenght_rod * (
                    FindIndexPZ.check_str_None(self, diam_rod) / 1000) / lenght_rod) ** 2) / 4 * lenght_rod
    return round(volume_rod, 5)


def volume_nkt(dict_nkt):  # Внутренний объем НКТ по фондовым НКТ
    volume_nkt = 0

    for nkt, length_nkt in dict_nkt.items():
        if nkt:
            volume_nkt += (float(nkt) - 2 * 7.6) ** 2 * 3.14 / 4 / 1000000 * length_nkt
    # print(f'объем НКТ {volume_nkt}')
    return volume_nkt


def weigth_pipe(dict_nkt):
    weigth_pipe = 0
    for nkt, lenght_nkt in dict_nkt.items():
        if '73' in str(nkt):
            weigth_pipe += lenght_nkt * 9.2 / 1000
        elif '60' in str(nkt):
            weigth_pipe += lenght_nkt * 7.5 / 1000
        elif '89' in str(nkt):
            weigth_pipe += lenght_nkt * 16 / 1000
        elif '48' in str(nkt):
            weigth_pipe += lenght_nkt * 4.3 / 1000
    return weigth_pipe


def volume_nkt_metal(dict_nkt):  # Внутренний объем НКТ железа по фондовым
    volume_nkt_metal = 0
    for nkt, length_nkt in dict_nkt.items():
        if '73' in str(nkt):
            volume_nkt_metal += 1.17 * length_nkt / 1000
        elif '60' in str(nkt):
            volume_nkt_metal += 0.87 * length_nkt / 1000
        elif '89' in str(nkt):
            volume_nkt_metal += 1.7 * length_nkt / 1000
        elif '48' in str(nkt):
            volume_nkt_metal += 0.55 * length_nkt / 1000
    return round(volume_nkt_metal, 1)


def well_volume(self, current_bottom):
    # print(well_data.column_additional)
    if well_data.column_additional is False:
        # print(well_data.column_diametr._value, well_data.column_wall_thickness._value, current_bottom)
        volume_well = 3.14 * (
                well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000000 * (
                          current_bottom)

    else:
        # print(f' ghb [{well_data.column_additional_diametr._value, well_data.column_additional_wall_thickness._value}]')
        volume_well = (3.14 * (
                well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                               current_bottom - float(well_data.head_column_additional._value)) / 1000) + (
                              3.14 * (
                              well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                  float(well_data.head_column_additional._value)) / 1000)
    # print(f'Объем скважины {volume_well}')
    return round(volume_well, 1)


def volume_pod_NKT(self):  # Расчет необходимого объема внутри НКТ и между башмаком НКТ и забоем

    nkt_l = round(sum(list(well_data.dict_nkt.values())), 1)
    if well_data.column_additional is False:
        v_pod_gno = 3.14 * (int(well_data.column_diametr._value) - int(
            well_data.column_wall_thickness._value) * 2) ** 2 / 4 / 1000 * (
                            float(well_data.current_bottom) - int(nkt_l)) / 1000

    elif round(sum(list(well_data.dict_nkt.values())), 1) > float(well_data.head_column_additional._value):
        v_pod_gno = 3.14 * (
                well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                            float(well_data.head_column_additional._value) - nkt_l) / 1000 + 3.14 * (
                            well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                            well_data.current_bottom - float(well_data.head_column_additional._value)) / 1000
    elif nkt_l <= float(well_data.head_column_additional._value):
        v_pod_gno = 3.14 * (
                well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                            well_data.current_bottom - nkt_l) / 1000
    volume_in_nkt = v_pod_gno + volume_vn_nkt(well_data.dict_nkt) - volume_rod(self, well_data.dict_sucker_rod)
    # print(f'Внутренный объем + Зумпф{volume_in_nkt, v_pod_gno, volume_vn_nkt(well_data.dict_nkt)}, ')
    return round(volume_in_nkt, 1)


def volume_jamming_well(self, current_bottom):  # объем глушения скважины

    volume_jamming_well = round(
        (well_volume(self, current_bottom) - volume_nkt_metal(well_data.dict_nkt) - volume_rod(self,
                                                                                               well_data.dict_sucker_rod)) * 1.1,
        1)
    # print(f' объем глушения {well_volume(self, well_data.current_bottom), volume_jamming_well}')
    # print(f' объем {volume_nkt_metal(well_data.dict_nkt)} , {volume_rod(self, well_data.dict_sucker_rod)}')
    return volume_jamming_well


def well_jamming(self, without_damping, lift_key, volume_well_jaming):
    # print(f' выбранный {lift_key}')

    # print(f'расстояние ПВР {abs(sum(list(well_data.dict_nkt.values())) - well_data.perforation_roof), volume_jamming_well(self), volume_nkt_metal(well_data.dict_nkt), volume_rod(well_data.dict_sucker_rod)}')
    well_jamming_list2 = f'Вести контроль плотности на  выходе в конце глушения. В случае отсутствия  на последнем кубе глушения  жидкости ' \
                         f'уд.веса равной удельному весу ЖГ, дальнейшие промывки и удельный вес жидкостей промывок согласовать с Заказчиком,' \
                         f' при наличии Ризб - произвести замер, перерасчет ЖГ и повторное глушение с корректировкой удельного веса жидкости' \
                         f' глушения. В СЛУЧАЕ ОТСУТСТВИЯ ЦИРКУЛЯЦИИ ПРИ ГЛУШЕНИИ СКВАЖИНЫ, А ТАКЖЕ ПРИ ГАЗОВОМ ФАКТОРЕ БОЛЕЕ 200м3/сут ' \
                         f'ПРОИЗВЕСТИ ЗАМЕР СТАТИЧЕСКОГО УРОВНЯ В ТЕЧЕНИИ ЧАСА С ОТБИВКОЙ УРОВНЯ В СКВАЖИНЕ С ИНТЕРВАЛОМ 15 МИНУТ.' \
                         f'ПО РЕЗУЛЬТАТАМ ЗАМЕРОВ ПРИНИМАЕТСЯ РЕШЕНИЕ ОБ ПРОДОЛЖЕНИИ ОТБИВКИ УРОВНЯ В СКВАЖИНЕ ДО КРИТИЧЕСКОЙ ГЛУБИНЫ ЗА ' \
                         f'ПРОМЕЖУТОК ВРЕМЕНИ.'

    # print(f' Глушение {volume_jamming_well(self, well_data.current_bottom), volume_nkt_metal(well_data.dict_nkt), volume_rod(well_data.dict_sucker_rod)}')
    # print(well_data.well_volume_in_PZ)

    if without_damping is True:
        well_jamming_str = f'Скважина состоит в перечне скважин ООО Башнефть-Добыча, на которых допускается проведение ТКРС без предварительного глушения на текущий квартал'
        well_jamming_short = f'Скважина без предварительного глушения'
        well_jamming_list2 = f'В случае наличия избыточного давления необходимость повторного глушения скважины дополнительно согласовать со специалистами ПТО  и ЦДНГ.'
    elif without_damping is False and lift_key in ['НН с пакером', 'НВ с пакером', 'ЭЦН с пакером', 'ОРЗ']:

        well_after = f'Произвести закачку на поглощение не более {well_data.max_admissible_pressure._value}атм тех жидкости в ' \
                     f'объеме {round(volume_well_jaming - well_volume(self, sum(list(well_data.dict_nkt_po.values()))), 1)}м3.' if round(
            volume_well_jaming - well_volume(self, sum(list(well_data.dict_nkt_po.values()))), 1) > 0.1 else ''
        well_jamming_str = f'Произвести закачку в трубное пространство тех жидкости уд.весом {well_data.fluid_work} в ' \
                           f'объеме {round(well_volume(self, sum(list(well_data.dict_nkt.values()))) - volume_pod_NKT(self), 1)}м3 на циркуляцию. ' \
                           f'{well_after} Закрыть затрубное пространство. ' \
                           f' Закрыть скважину на  стабилизацию не менее 2 часов. (согласовать ' \
                           f'глушение в коллектор, в случае отсутствия на желобную емкость)'
        well_jamming_short = f'Глушение в НКТ уд.весом {well_data.fluid_work_short} ' \
                             f'объеме {round(well_volume(self, sum(list(well_data.dict_nkt.values()))) - volume_pod_NKT(self), 1)}м3 ' \
                             f'на циркуляцию. {well_after} '
    elif without_damping is False and lift_key in ['ОРД']:
        well_jamming_str = f'Произвести закачку в затрубное пространство тех жидкости уд.весом {well_data.fluid_work_short}в ' \
                           f'объеме {round(well_volume(self, well_data.current_bottom) - well_volume(self, well_data.depth_fond_paker_do["do"]), 1)}м3 ' \
                           f'на поглощение при давлении не более {well_data.max_admissible_pressure._value}атм. Закрыть ' \
                           f'затрубное пространство. Закрыть скважину на стабилизацию не менее 2 часов. (согласовать ' \
                           f'глушение в коллектор, в случае отсутствия на желобную емкость)'
        well_jamming_short = f'Глушение в затруб уд.весом {well_data.fluid_work_short} в ' \
                             f'объеме {round(well_volume(self, well_data.current_bottom) - well_volume(self, well_data.depth_fond_paker_do["do"]), 1)}м3 '
    elif without_damping is False and lift_key in ['НН', 'НВ', 'ЭЦН']:
        well_jamming_str = f'Произвести глушение скважины в объеме {volume_well_jaming}м3 тех ' \
                           f'жидкостью уд.весом {well_data.fluid_work}' \
                           f' на циркуляцию в следующим алгоритме: \n Произвести закачку в затрубное пространство ' \
                           f'тех жидкости в ' \
                           f'объеме {round(well_volume(self, sum(list(well_data.dict_nkt.values()))), 1)}м3 на ' \
                           f'циркуляцию. Закрыть трубное пространство. ' \
                           f'Произвести закачку на поглощение не более {well_data.max_admissible_pressure._value}атм ' \
                           f'тех жидкости в ' \
                           f'объеме {round(volume_well_jaming - well_volume(self, sum(list(well_data.dict_nkt.values()))), 1)}м3. Закрыть скважину на ' \
                           f'стабилизацию не менее 2 часов. (согласовать глушение в коллектор, в случае ' \
                           f'отсутствия на желобную емкость'
        well_jamming_short = f'Глушение в затруб в объеме {volume_well_jaming}м3 тех ' \
                             f'жидкостью уд.весом {well_data.fluid_work_short}'
    elif abs(sum(list(well_data.dict_nkt.values())) - well_data.perforation_roof) > 150:
        well_jamming_str = f'Произвести глушение скважины объеме {volume_well_jaming}м3 тех ' \
                           f'жидкостью уд.весом {well_data.fluid_work}' \
                           f' на циркуляцию в следующим алгоритме: \n Произвести закачку в затрубное пространство ' \
                           f'тех жидкости в ' \
                           f'объеме {round(well_volume(self, sum(list(well_data.dict_nkt.values()))), 1)}м3 на ' \
                           f'циркуляцию. Закрыть трубное пространство. ' \
                           f'Произвести закачку на поглощение не более {well_data.max_admissible_pressure._value}атм ' \
                           f'тех жидкости в ' \
                           f'объеме {round(volume_well_jaming - well_volume(self, sum(list(well_data.dict_nkt.values()))), 1)}м3. Закрыть скважину на ' \
                           f'стабилизацию не менее 2 часов. (согласовать глушение в коллектор, в случае ' \
                           f'отсутствия на желобную емкость'
        well_jamming_short = f'Глушение в затруб в объеме {volume_well_jaming}м3 тех ' \
                             f'жидкостью уд.весом {well_data.fluid_work_short}'
    elif abs(sum(list(well_data.dict_nkt.values())) - well_data.perforation_roof) <= 150:
        well_jamming_str = f'Произвести глушение скважины  в объеме {volume_well_jaming}м3 тех ' \
                           f'жидкостью уд.весом {well_data.fluid_work}' \
                           f' на циркуляцию. Закрыть скважину на ' \
                           f'стабилизацию не менее 2 часов. (согласовать глушение в коллектор, в случае отсутствия ' \
                           f'на желобную емкость)'
        well_jamming_short = f'Глушение в затруб в объеме {volume_well_jaming}м3 уд.весом {well_data.fluid_work_short}'

        # print([well_jamming_str, well_jamming_list2, well_jamming_short])
    return [well_jamming_str, well_jamming_list2, well_jamming_short]



def count_row_height(ws2, work_list,  merged_cells_dict):


    ind_ins = 46

    stop_str = len(work_list)
    for i in range(1, stop_str + 1):  # Добавлением работ
        for j in range(1, 31):
            cell = ws2.cell(row=i, column=j)
            if cell and str(cell) != str(work_list[i - 1][j - 1]):
                if work_list[i - 1][j - 1]:
                    cell.value = is_num(work_list[i - 1][j - 1])
                    if i >= ind_ins:

                        if j == 11:
                            cell.font = Font(name='Times New Roman', size=11, bold=False)
                        else:
                            cell.font = Font(name='Times New Roman', size=16, bold=False)
                        ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                        ws2.cell(row=i, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                         vertical='center')
                        ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')


    ws2.merge_cells(start_row=77, start_column=2, end_row=77, end_column=3)
    ws2.merge_cells(start_row=77, start_column=6, end_row=77, end_column=14)
    for key, row in merged_cells_dict.items():
        ws2.merge_cells(start_row=row[1], start_column=row[0], end_row=row[3], end_column=row[2])

    return 'Высота изменена'

def is_num(num):
    try:
        if isinstance(num, datetime):
            return num.strftime('%d.%m.%Y')
        elif str(round(float(num), 6))[-1] != 0:
            return round(float(num), 6)
        elif str(round(float(num), 5))[-1] != 0:
            return round(float(num), 5)
        elif str(round(float(num), 4))[-1] != 0:
            return round(float(num), 4)
        elif str(round(float(num), 3))[-1] != 0:
            return round(float(num), 3)
        elif str(round(float(num), 2))[-1] != 0:
            return round(float(num), 2)
        elif str(round(float(num), 1))[-1] != 0:
            return round(float(num), 1)
        elif str(round(float(num), 0))[-1] != 0:
            return int(float(num))
    except:
        return num
def is_number(num):
    if num is None:
        return 0
    try:
        float(str(num).replace(",", "."))
        return True
    except ValueError or TypeError:
        return False


# Определение трех режимов давлений при определении приемистости
def pressure_mode(mode, plast):
    mode = float(mode) / 10 * 10
    if mode > well_data.max_admissible_pressure._value and (plast != 'D2ps' or plast.lower() != 'дпаш'):
        mode_str = f'{float(mode)}, {float(mode) - 10}, {float(mode) - 20}'
    elif (plast == 'D2ps' or plast.lower() == 'дпаш') and well_data.region == 'ИГМ':
        mode_str = f'{120}, {140}, {160}'
    else:
        mode_str = f'{float(mode) - 10}, {float(mode)}, {float(mode) + 10}'
    return mode_str


