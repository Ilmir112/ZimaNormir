import base64

import well_data
from datetime import datetime
from PyQt5.QtWidgets import QInputDialog, QMessageBox, QMainWindow
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment
from block_name import region, razdel_1, curator_sel, pop_down


from find import ProtectedIsNonNone


class CreatePZ(QMainWindow):
    def __init__(self, wb, ws, data_window, perforation_correct_window2, parent=None):
        super(CreatePZ, self).__init__(parent)

        self.wb = wb
        self.ws = ws
        self.data_window = data_window
        self.perforation_correct_window2 = perforation_correct_window2

    def open_excel_file(self, ws, work_plan):
        from find import FindIndexPZ

        from category_correct import CategoryWindow
        from main import MyWindow
        from find import WellNkt, Well_perforation, WellCondition, WellHistory_data, Well_data, Well_Category, \
            WellFond_data, WellSucker_rod, Well_expected_pick_up, WellData
        from data_base.work_with_base import check_in_database_well_data

        well_data.work_plan = work_plan

        well_data.dict_category = CategoryWindow.dict_category
        # Запуск основного класса и всех дочерних классов в одной строке
        well_pz = FindIndexPZ(ws)

        well_data.region = region(well_data.cdng._value)

        WellData.read_well(WellData, ws, well_data.cat_well_max._value, well_data.data_pvr_min._value)
        well_data.region = region(well_data.cdng._value)



        if well_data.data_well_is_True is False:
            WellNkt.read_well(self, ws, well_data.pipes_ind._value, well_data.condition_of_wells._value)
            if well_data.work_plan not in ['application_pvr', 'application_gis']:
                WellSucker_rod.read_well(self, ws, well_data.sucker_rod_ind._value, well_data.pipes_ind._value)
                WellFond_data.read_well(self, ws, well_data.data_fond_min._value, well_data.condition_of_wells._value)
            WellHistory_data.read_well(self, ws, well_data.data_pvr_max._value, well_data.data_fond_min._value)
            WellCondition.read_well(self, ws, well_data.condition_of_wells._value, well_data.data_well_max._value)

            Well_expected_pick_up.read_well(self, ws, well_data.data_x_min._value, well_data.data_x_max._value)
            Well_data.read_well(self, ws, well_data.cat_well_max._value, well_data.data_pvr_min._value)

            Well_perforation.read_well(self, ws, well_data.data_pvr_min._value, well_data.data_pvr_max._value + 1)
            Well_Category.read_well(self, ws, well_data.cat_well_min._value, well_data.data_well_min._value)




        if well_data.inv_number._value == 'не корректно' or well_data.inv_number is None:
            mes = QMessageBox.warning(self, 'Инвентарный номер отсутствует',
                                      'Необходимо уточнить наличие инвентарного номера')
            return



        if work_plan not in ['application_pvr', 'application_gis', 'gnkt_bopz', 'gnkt_opz', 'gnkt_after_grp',
                             'gnkt_frez']:
            if work_plan != 'plan_change':
                for row_ind, row in enumerate(ws.iter_rows(values_only=True, max_col=13)):
                    ws.row_dimensions[row_ind].hidden = False

                    if any(['ПЛАН РАБОТ' in str(col).upper() for col in row]) \
                            and work_plan == 'dop_plan':
                        ws.cell(row=row_ind + 1, column=2).value = f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {well_data.number_dp}'


                    elif 'План-заказ' in row:
                        # print(row)

                        ws.cell(row=row_ind + 1, column=2).value = 'ПЛАН РАБОТ'

                    for col, value in enumerate(row):
                        if not value is None and col <= 12:
                            if 'гипс' in str(value).lower() or 'гидратн' in str(value).lower():
                                well_data.gipsInWell = True

                if well_data.emergency_well is True:
                    emergency_quest = QMessageBox.question(self, 'Аварийные работы ',
                                                           'Программа определила что в скважине'
                                                           f' авария - {well_data.emergency_count}, верно ли?')
                    if emergency_quest == QMessageBox.StandardButton.Yes:
                        well_data.emergency_well = True
                        well_data.emergency_bottom, ok = QInputDialog.getInt(self, 'Аварийный забой',
                                                                             'Введите глубину аварийного забоя',
                                                                             0, 0,
                                                                             int(well_data.bottomhole_artificial._value))
                    else:
                        well_data.emergency_well = False
                if well_data.problemWithEk is True:
                    problemWithEk_quest = QMessageBox.question(self, 'ВНИМАНИЕ НЕПРОХОД ',
                                                               f'Программа определила что в скважине '
                                                               f'ссужение в ЭК -, верно ли?')
                    if problemWithEk_quest == QMessageBox.StandardButton.Yes:
                        well_data.problemWithEk = True
                        well_data.problemWithEk_depth, ok = QInputDialog.getInt(self, 'Глубина сужения',
                                                                                "ВВедите глубину cсужения", 0, 0,
                                                                                int(well_data.current_bottom))
                        well_data.problemWithEk_diametr = QInputDialog.getInt(self, 'диаметр внутренний cсужения',
                                                                              "ВВедите внутренний диаметр cсужения", 0,
                                                                              0,
                                                                              int(well_data.current_bottom))[0]
                    else:
                        well_data.problemWithEk = ProtectedIsNonNone(False)

                if well_data.gipsInWell is True:
                    gips_true_quest = QMessageBox.question(self, 'Гипсовые отложения',
                                                           'Программа определила что скважина осложнена гипсовыми отложениями '
                                                           'и требуется предварительно определить забой на НКТ, верно ли это?')

                    if gips_true_quest == QMessageBox.StandardButton.Yes:
                        well_data.gipsInWell = True
                    else:
                        well_data.gipsInWell = False

            try:
                # Копирование изображения
                image_loader = SheetImageLoader(ws)
            except:
                mes = QMessageBox.warning(None, 'Ошибка', 'Ошибка в копировании изображений')


            well_data.image_data = []
            for row in range(1, well_data.data_well_max._value):
                for col in range(1, 12):
                    try:
                        image = image_loader.get(f'{get_column_letter(col)}{row}')
                        image.save(
                            f'{well_data.path_image}imageFiles/image_work/image{get_column_letter(col)}{row}.png')
                        image_size = image.size
                        image_path = f'{well_data.path_image}imageFiles/image_work/image{get_column_letter(col)}{row}.png'

                        coord = f'{get_column_letter(col)}{row + 17 - well_data.cat_well_min._value}'

                        well_data.image_list.append((image_path, coord, image_size))
                        # Чтение изображения в байты
                        with open(image_path, "rb") as f:
                            image_bytes = f.read()
                        # Преобразование в Base64
                        image_base64 = base64.b64encode(image_bytes).decode("utf-8")

                        # Создание словаря для изображения
                        image_info = {
                            "coord": coord,
                            "width": image_size[0],
                            "height": image_size[1],
                            "data": image_base64
                        }
                        # Сохранение Base64 данных в файл (для проверки)
                        with open("image_base64.txt", "w", encoding="utf-8") as f:
                            f.write(image_base64)
                        # Добавление информации в список
                        well_data.image_data.append(image_info)

                    except:
                        pass
            if work_plan != 'plan_change':
                for j in range(well_data.data_x_min._value,
                               well_data.data_x_max._value):  # Ожидаемые показатели после ремонта
                    lst = []
                    for i in range(0, 12):
                        lst.append(ws.cell(row=j + 1, column=i + 1).value)
                    well_data.row_expected.append(lst)



            return ws


            return ws



    def add_itog(self, ws, ins_ind, work_plan):
        if ws.merged_cells.ranges:
            merged_cells_copy = list(ws.merged_cells.ranges)  # Создаем копию множества объединенных ячеек
            for merged_cell in merged_cells_copy:
                if merged_cell.min_row > ins_ind + 5:
                    try:
                        ws.unmerge_cells(str(merged_cell))
                    except:
                        pass



        curator_s = curator_sel(well_data.curator, well_data.region)
        # print(f'куратор {curator_sel, well_data.curator}')
        podp_down = pop_down(self, well_data.region, curator_s)

        for i in range(1 + ins_ind, 1 + ins_ind + len(podp_down)):

            # Добавлением подписантов внизу
            for j in range(1, 13):
                ws.cell(row=i, column=j).value = podp_down[i - 1 - ins_ind][j - 1]
                ws.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=False)

            if i in range(ins_ind + 7, 1 + ins_ind + 15):
                ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical='center', horizontal='left')
            else:
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical='center', horizontal='left')
        ws.row_dimensions[ins_ind + 7].height = 30
        ws.row_dimensions[ins_ind + 9].height = 25

        ins_ind += len(podp_down)
        aaa = ws.max_row

        ws.delete_rows(ins_ind, aaa - ins_ind)

    def is_valid_date(date):
        try:
            datetime.strptime(date, '%Y-%m-%d')
            return True
        except ValueError:
            return False

#
