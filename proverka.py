from PyQt5 import QtWidgets
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

fname = 'property_excel/template_normir_2.xlsm'
wb2 = load_workbook(fname, keep_vba=True)
name_list = wb2.sheetnames
# print(name_list)
ws2 = wb2['АВР']
# if fname:
#     normir_list = []
#     for row_ind, row in enumerate(ws2.iter_rows(max_col=31, min_row=1249, max_row=1348)):
#         list = []
#
#
#         for cell in row[:32]:
#             # Получение значения и стилей
#             value = cell.value
#
#             font = cell.font
#             fill = cell.fill
#             # Преобразуем RGB в строковый формат
#             rgb_string = f"RGB({fill.fgColor.rgb})"
#
#             borders = cell.border
#             alignment = cell.alignment
#             protection = cell.protection
#
#
#             list.append({
#
#                 'value': value,
#                 'font': {
#                     'name': font.name,
#                     'size': font.size,
#                     'bold': font.bold,
#                     'italic': font.italic
#                 },
#                 'fill': {
#                     'color': rgb_string
#                 },
#                 'borders': {
#                     'left': borders.left.style,
#                     'right': borders.right.style,
#                     'top': borders.top.style,
#                     'bottom': borders.bottom.style
#                 },
#                 'alignment': {
#                     'horizontal': alignment.horizontal,
#                     'vertical': alignment.vertical,
#                     'wrap_text': alignment.wrap_text
#                 },
#                 'protection': {
#                     'hidden': protection.hidden,
#                     'locked': protection.locked},
#             })
#         normir_list.append(list)
#     a = normir_list
#
#     print(normir_list)
a = [
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'Очистка от снега в сутки 2 раза',
     'Очистка от снега в сутки 2 раза', None, None, None, None, None, None, None, None, None, None, None, None,
     '§12разд.1', None, 'шт', '=(H19/22)*2', 0.07, 1, '=V1249*W1249*X1249', '=Y1249-AA1249-AB1249-AC1249-AD1249', None,
     None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'ПЗР в начале и конце смены с заполнением вахтового журнала',
     'ПЗР в начале и конце смены с заполнением вахтового журнала', None, None, None, None, None, None, None, None, None,
     None, None, None, '§13разд.1', None, 'шт', '=(H19/22)*2', 0.31, 1, '=V1250*W1250*X1250',
     '=Y1250-AA1250-AB1250-AC1250-AD1250', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'проведение УТЗ', 'проведение УТЗ', None, None, None, None, None,
     None, None, None, None, None, None, None, '§300разд.1', None, 'шт', '=(H19/82.5)', 0.06, 1, '=V1251*W1251*X1251',
     '=Y1251-AA1251-AB1251-AC1251-AD1251', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'Открытие закрытие скв', 'Открытие закрытие скв', None, None,
     None, None, None, None, None, None, None, None, None, None, '§300разд.1', None, 'шт', '=(H19/22)*2', 0.18, 1,
     '=V1252*W1252*X1252', '=Y1252-AA1252-AB1252-AC1252-AD1252', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'Заправка ПА', 'Заправка ПА', None, None, None, None, None, None,
     None, None, None, None, None, None, '§15разд.1', None, 'шт', '=H19/22', 0.13, 1, '=V1253*W1253*X1253',
     '=Y1253-AA1253-AB1253-AC1253-AD1253', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'Усредненный поправочный зимний коэффициент',
     'Усредненный поправочный зимний коэффициент (декабрь-март)', None, None, None, None, None, None, None, 0.05, None,
     None, None, None, 'П2-05.01 Н-0002 ЮЛ-305 в 4', None, None, None, None, None,
     '=(SUMIFS($Y47:$Y1232,B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254)-SUMIFS($Y47:$Y1232,$D47:$D1232,"переезд",$S47:$S1232,"§302разд1",B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254)-SUMIFS($Y47:$Y1232,$S47:$S1232,ЦИКЛ!$U$21,B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254)-SUMIFS($Y47:$Y1232,$S47:$S1232,ЦИКЛ!$U$20,B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254)-SUMIFS($Y47:$Y1232,$S47:$S1232,ЦИКЛ!$U$18,B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254)-SUMIFS($Y47:$Y1232,$S47:$S1232,ЦИКЛ!$U$19,B47:B1232,">="&AI1254,B47:B1232,"<="&AJ1254))*$N1254',
     '=Y1254-AA1254-AB1254-AC1254-AD1254', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, None, None, 'Нормативное время (без учета мелких ремонтных работ)', None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     '=ROUND(SUM(Y47:Y1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$Y$47:$Y$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$Y$47:$Y$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$Y$47:$Y$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$Y$47:$Y$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$Y$47:$Y$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$Y$47:$Y$1254),2)',
     '=ROUND(SUM(Z47:Z1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$Z$47:$Z$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$Z$47:$Z$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$Z$47:$Z$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$Z$47:$Z$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$Z$47:$Z$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$Z$47:$Z$1254),2)',
     None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Тех.операции', 'Мелкие ремонтные работы', 'Мелкие ремонтные работы', None, None,
     None, None, None, None, None, None, None, None, None, None, 'П2-05.01Н-0002ЮЛ-305 вер.4', None, None, None, None,
     None,
     '=ROUND((Y1255-SUMIF(D47:D1253,ЦИКЛ!T17,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T18,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T19,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T20,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T21,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T22,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T23,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T24,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T25,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T26,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T27,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T28,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T29,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T30,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T32,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T33,Y47:Y1253)-SUMIF(S47:S1253,ЦИКЛ!T34,Y47:Y1253))*(F30/100*0.004),2)',
     '=ROUND(Y1256-AA1256-AB1256-AC1256,2)', None, None, None, None, None],
    ['=ROW()-ROW($A$46)', None, None, 'Нормативное время', None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, '=ROUND(Y1255+Y1256,2)', '=ROUND(Z1255+Z1256,2)',
     '=ROUND(SUM(AA47:AA1256)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AA$47:$AA$1254),2)',
     '=ROUND(SUM(AB46:AB1256)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AB$47:$AB$1254),2)',
     '=ROUND(SUM(AC46:AC1256)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AC$47:$AC$1254),2)',
     '=ROUND(SUM(AD46:AD1256)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AD$47:$AD$1254),2)',
     None],
    ['А н а л и з    н е    п р о и з в о д и т е л ь н о г о    в р е м е н и', None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None],
    ['Химия и метериалы использованные в процессе ремонта', None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None],
    [None, '№', 'Наименование', None, None, None, 'Ед.изм', 'Кол-во', 'Стоимость', None, 'Итого', '№', 'Наименование',
     None, None, None, 'Ед.изм', 'Кол-во', 'Стоимость', None, 'Итого', '№', 'Наименование', None, None, None, 'Ед.изм',
     'Кол-во', 'Стоимость', None, 'Итого'],
    [None, 1, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1261,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1261="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1261,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1261*I1261,""),2),"")', 11,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1261,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1261="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1261,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1261*S1261,"")', 21,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1261,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1261="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1261,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1261*AC1261,"")'],
    [None, 2, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1262,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1262="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1262,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1262*I1262,""),2),"")', 12,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1262,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1262="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1262,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1262*S1262,"")', 22,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1262,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1262="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1262,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1262*AC1262,"")'],
    [None, 3, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1263,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1263="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1263,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1263*I1263,""),2),"")', 13,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1263,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1263="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1263,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1263*S1263,"")', 23,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1263,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1263="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1263,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1263*AC1263,"")'],
    [None, 4, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1264,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1264="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1264,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1264*I1264,""),2),"")', 14,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1264,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1264="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1264,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1264*S1264,"")', 24,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1264,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1264="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1264,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1264*AC1264,"")'],
    [None, 5, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1265,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1265="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1265,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1265*I1265,""),2),"")', 15,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1265,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1265="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1265,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1265*S1265,"")', 25,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1265,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1265="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1265,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1265*AC1265,"")'],
    [None, 6, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1266,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1266="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1266,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1266*I1266,""),2),"")', 16,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1266,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1266="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1266,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1266*S1266,"")', 26,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1266,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1266="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1266,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1266*AC1266,"")'],
    [None, 7, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1267,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1267="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1267,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1267*I1267,""),2),"")', 17,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1267,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1267="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1267,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1267*S1267,"")', 27,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1267,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1267="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1267,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1267*AC1267,"")'],
    [None, 8, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1268,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1268="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1268,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1268*I1268,""),2),"")', 18,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1268,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1268="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1268,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1268*S1268,"")', 28,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1268,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1268="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1268,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1268*AC1268,"")'],
    [None, 9, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1269,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1269="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1269,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1269*I1269,""),2),"")', 19,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1269,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1269="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1269,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1269*S1269,"")', 29,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1269,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1269="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1269,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1269*AC1269,"")'],
    [None, 10, '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(B1270,$AG$47:$AG$1255,0)),"")', None, None, None,
     '=IF(C1270="","","тн/шт")', '=IFERROR(INDEX($R$47:$R$1256,MATCH(B1270,$AG$47:$AG$1255,0)),"")', None, None,
     '=IFERROR(ROUND(IFERROR(H1270*I1270,""),2),"")', 20,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(L1270,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(M1270="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(L1270,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(R1270*S1270,"")', 30,
     '=IFERROR(INDEX($Q$47:$Q$1256,MATCH(V1270,$AG$47:$AG$1255,0)),"")', None, None, None, '=IF(W1270="","","тн/шт")',
     '=IFERROR(INDEX($R$47:$R$1256,MATCH(V1270,$AG$47:$AG$1255,0)),"")', None, None, '=IFERROR(AB1270*AC1270,"")'],
    [None, 'Аварии по вине ПО ТКРС', None, None, None, 'Аварии по вине 3-их лиц', None, None, None, 'Браки по вине БНД',
     None, None, None, 'ДК 50/50', None, None, None, 'по ДК вина БНД', None, None, None, 'по ДК вина 3-их лиц', None,
     None, None, 'Прочия химия', None, None, None, None, None],
    [None, 'Час', None,
                                                                 '=SUMIF(AE47:AE1256,"Авария ПО ТКРС",AA47:AA1256)+SUMIF(AE47:AE1256,"Авария ПО ТКРС",AB47:AB1256)+SUMIF(AE47:AE1256,"Авария ПО ТКРС",AC47:AC1256)+SUMIF(AE47:AE1256,"Авария ПО ТКРС",AD47:AD1256)',
                                                                 None, 'Час', None,
                                                                 '=SUMIF(AE47:AE1256,"Авария 3-их лиц",Z47:Z1256)',
                                                                 None, 'Час', None,
                                                                 '=SUMIF(AE47:AE1256,"Брак БНД",Z47:Z1256)', None,
                                                                 'Общий', None,
                                                                 '=SUMIF(AE47:AE1256,"ДК 50/50",AA47:AA1256)+SUMIF(AE47:AE1256,"ДК 50/50",AB47:AB1256)+SUMIF(AE47:AE1256,"ДК 50/50",AC47:AC1256)+SUMIF(AE47:AE1256,"ДК 50/50",AD47:AD1256)+SUMIF(AE47:AE1256,"ДК 50/50",Z47:Z1256)',
                                                                 None, 'Час', None,
                                                                 '=SUMIF(AE47:AE1256,"ДК вина БНД",Z47:Z1256)', None,
                                                                 'Час', None,
                                                                 '=SUMIF(AE47:AE1256,"ДК вина 3-их лиц",Z47:Z1256)',
                                                                 None, 'приапр', '=IF(W1272="","","тн/шт")',
                                                                 '=SUMIF($Q47:$Q1253,Z1271,$R47:$R1253)', None, None,
                                                                 '=AB1272*AC1272'],
    [None, 'Аварии по вине БНД', None, None, None, 'Браки по вине ПО ТКРС', None, None, None, 'Браки по вине 3-их лиц',
     None, None, None, 'ПО ТКРС', None,
     '=SUMIF(AE47:AE1256,"ДК 50/50",AA47:AA1256)+SUMIF(AE47:AE1256,"ДК 50/50",AB47:AB1256)+SUMIF(AE47:AE1256,"ДК 50/50",AC47:AC1256)+SUMIF(AE47:AE1256,"ДК 50/50",AD47:AD1256)',
     None, 'по ДК вина ПО ТКРС', None, None, None, 'ГТС от БНД', None, None, None, 'ГТС от ПО ТКРС', None, None, None,
     None, None],
    [None, 'Час', None, '=SUMIF(AE47:AE1256,"Авария БНД",Z47:Z1256)', None, 'Час', None,
                   '=SUMIF(AE47:AE1256,"Брак ПО ТКРС",AA47:AA1256)+SUMIF(AE47:AE1256,"Брак ПО ТКРС",AB47:AB1256)+SUMIF(AE47:AE1256,"Брак ПО ТКРС",AC47:AC1256)+SUMIF(AE47:AE1256,"Брак ПО ТКРС",AD47:AD1256)',
                   None, 'Час', None, '=SUMIF(AE47:AE1256,"Брак 3-их лиц",Z47:Z1256)', None, 'БНД', None,
                   '=SUMIF(AE47:AE1256,"ДК 50/50",Z47:Z1256)', None, 'Час', None,
                   '=SUMIF(AE47:AE1256,"ДК вина ПО ТКРС",AA47:AA1256)+SUMIF(AE47:AE1256,"ДК вина ПО ТКРС",AB47:AB1256)+SUMIF(AE47:AE1256,"ДК вина ПО ТКРС",AC47:AC1256)+SUMIF(AE47:AE1256,"ДК вина ПО ТКРС",AD47:AD1256)',
                   None, 'Час', None, '=SUMIF(AE47:AE1256,"ГТС от БНД",Z47:Z1256)', None, 'Час', None,
                   '=SUMIF(AE47:AE1256,"ГТС от ПО ТКРС",AA47:AA1256)+SUMIF(AE47:AE1256,"ГТС от ПО ТКРС",AB47:AB1256)+SUMIF(AE47:AE1256,"ГТС от ПО ТКРС",AC47:AC1256)+SUMIF(AE47:AE1256,"ГТС от ПО ТКРС",AD47:AD1256)+SUMIF(AE47:AE1256,"ГТС от ПО ТКРС",Z47:Z1256)',
                   None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Календарное время, час', None, None, None, None, None, None, None, '=H19', None,
     'Технологические жидкости образованные в процессе ремонта', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None],
    [None, 'Фактическое  время,  час', None, None, None, None, None, None, None, '=ROUND(J1276/24*22,2)', None,
     'Наименование технологической операции', None, None, None, None, None, 'Объем образованной жидкости, м3', None,
     None, None, 'Место утилизации ', None, None, None, None, None, 'Объем утилизированной жидкости, м3', None, None,
     None],
    [None, 'Нормативное время, час', None, None, None, None, None, None, None, '=Y1257', None, 'Разрядка', None, None,
     None, None, None, '=SUMIF(D47:D1256,L1278,K47:K1256)', None, None, None, 'коллектор', None, None, None, None, None,
     0, None, None, None],
    [None, 'Снятые объемы, час', None, None, None, None, None, None, None, '=AA1257+AB1257+AC1257+AD1257', None,
     'Свабирование', None, None, None, None, None, '=SUMIF(E47:E1256,L1279,N47:N1256)', None, None, None, 'коллектор',
     None, None, None, None, None, 0, None, None, None],
    [None, 'Время к оплате (после проверки)', None, None, None, None, None, None, None, '=Z1257', None, 'Подъем с ПСЮ',
     None, None, None, None, None, '=SUMIFS(N47:N1256,D47:D1256,"спо",F47:F1256,"*Осложнение при подъеме*")', None,
     None, None, 'коллектор', None, None, None, None, None, 0, None, None, None],
    [None, ' в. т.ч Завышения (норматива), час', None, None, None, None, None, None, None,
     '=ROUND(SUM(AA1257:AD1257)-D1272-H1274-P1273-T1274-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AD$47:$AD$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AC$47:$AC$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AB$47:$AB$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$285,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$286,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$287,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$288,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$290,$AA$47:$AA$1254)-SUMIF($T$47:$T$1254,ЦИКЛ!$C$291,$AA$47:$AA$1254),2)',
     None, 'Спуск с вытеснением', None, None, None, None, None,
     '=SUMIFS(N47:N1256,D47:D1256,"спо",F47:F1256,"*Осложнение при спуске*")', None, None, None, 'коллектор', None,
     None, None, None, None, 0, None, None, None],
    [None, ' в. т.ч Сдельно (Факт),  час', None, None, None, None, None, None, None,
     '=ROUND(SUMIF($S$47:$S$1254,ЦИКЛ!U18,$Z$47:$Z$1254),2)', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, '           Повременно,  час', None, None, None, None, None, None, None, '=J1280-J1282', None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'К оплате  технологическое ожидание, час', None, None, None, None, None, None, None,
     '=SUMIF($T$47:$T$1254,ЦИКЛ!C285,$Z$47:$Z$1254)', None, 'Итого', None, None, None, None, None, '=SUM(R1278:U1283)',
     None, None, None, None, None, None, None, None, None, '=SUM(AB1278:AE1283)', None, None, None],
    [None, 'К оплате  простои Заказчика , час', None, None, None, None, None, None, None,
     '=SUMIF($T$47:$T$1254,ЦИКЛ!C287,$Z$47:$Z$1254)', None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None],
    [None, 'К оплате  метео, час', None, None, None, None, None, None, None,
     '=SUMIF($T$47:$T$1254,ЦИКЛ!C286,$Z$47:$Z$1254)', None,
     'Учёт жидкости с ПНТЖ ( или полученной от подрядчика по глушению)', None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Простои Подрядчика, час', None, None, None, None, None, None, None,
     '=SUMIF($T$47:$T$1254,ЦИКЛ!C288,$Y$47:$Y$1254)', None, 'Пункт набора (наименование внешней подрядной организации)',
     None, None, None, None, None, None, None, 'Объем, м3', None, 'Расход (Утилизация)', None, None, None, 'Объем, м3',
     None, 'Примечание', None, None, None],
    [None, 'Опережающее глушение скважины по графику движения ', None, None, None, None, None, None, None,
     '=ROUND(SUMIF($T$47:$T$1255,ЦИКЛ!C291,$Z$47:$Z$1254),2)', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Глушение скважины по согласованию с Заказчиком', None, None, None, None, None, None, None,
     '=ROUND(SUMIF($T$47:$T$1255,ЦИКЛ!C289,$Z$47:$Z$1254),2)', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Глушение скважины без согласования с Заказчиком', None, None, None, None, None, None, None,
     '=ROUND(SUMIF($T$47:$T$1255,ЦИКЛ!C290,$Z$47:$Z$1254),2)', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'КПВ', None, None, None, None, None, None, None, '=ROUND(J1292/J1277,2)', None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Итого к оплате,час', None, None, None, None, None, None, None, '=ROUND(J1280+J1284+J1285+J1286,2)', None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None],
    [None, 'К оплате по ставке простоев бр/час', None, None, None, None, None, None, None, '=J1284+J1285', None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'К оплате по ставке метео бр/час', None, None, None, None, None, None, None, '=J1286', None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'К оплате по ставке бр/час', None, None, None, None, None, None, None, '=J1280', None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, 'Итого', None, None, None, None, None, None,
     None, '=SUM(T1288:U1295)', None, None, None, None, None, '=SUM(Z1288:AA1295)', None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Штрафы', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Отвественный ', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'УСРСиСТ', None, None, None, None, None, None, None, None, None, None, None, 'ПТО РУДНГ', None, None, None,
     None, None, None, None, None, 'ОТКРСиО', None, None, None, None, None, None, None, None],
    [None, 'Причина', None, None, None, None, None, None, None, None, None, 'понижающий Кпон/руб', None, 'Причина',
     None, None, None, None, None, None, 'понижающий Кпон/руб', None, 'Причина', None, None, None, None, None, None,
     'понижающий Кпон/руб', None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    ['Итого по отделам', None, '=ROUND(IF(AF1303=0,IF(AF1304=0,AF1305,AF1304),AF1303),2)', None, None, None, None, None,
     None, None, None, None, None, '=ROUND(IF(AG1303=0,IF(AG1304=0,AG1305,AG1304),AG1303),2)', None, None, None, None,
     None, None, None, None, '=ROUND(IF(AI1303=0,IF(AI1304=0,AI1305,AI1304),AI1303),2)', None, None, None, None, None,
     None, None, None],
    ['Итого БНД', None, '=ROUND(IF(AF1310=1,AF1311,AF1310),2)', None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, 'Скважина:', None, None, None, '=D11', None, 'ВИД РЕМОНТА', None, '=N23', None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, 'Шифр', '=X23', 'дата начала', None, None, '=H7'],
    [None, 'ЦДНГ', None, None, None, '=D7', None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, 'дата завершения', None, None,
     '=IF(H8>H10,H8,IF(H10>H12,H10,IF(H12>H14,H12,IF(H14>H16,H14,IF(H16>H18,H16,IF(H18>H8,H18))))))'],
    [None, None, 'Причины невыполнения первоначальных мероприятий по план-заказу, отклонений план/факта КР и вида ГТМ',
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Замечания,выявленные в процессе приемки скважины', None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'По факту преждевременного отказа ГНО', None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Скважина:', None, None, '=F1313', None, 'ВИД РЕМОНТА', None, '=J1313', None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, 'Шифр', '=AA1313', 'дата начала', None, None,
     '=AE1313'],
    [None, None, 'ЦДНГ', None, None, '=F1314', None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, 'дата завершения', None, None, '=AE1314'],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Подписи сторон', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'От Заказчика:', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, 'От Подрядчика:', None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Представитель ПТО РУДНГ       ', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, 'Начальник ЦТКРС         ', None, None, None, None, None,
     '___________________________________   "________" _______ 202__ г.', None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Представитель Службы Главного геолога', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, 'Гл. технолог(геолог)  ', None, None, None, None, None,
     '___________________________________   "________" _______ 202__ г.', None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Ведущий супевайзер УСРСиСТ      ', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, 'Инженер по ОиНТ   ', None, None, None, None, None,
     '________Поленова Е.Н_________   "__03__" __07__ 2024_ г.', None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Начальник ЦДНГ', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, 'Мастер бригады №                                       ', None,
     '=IF(U9="",U7,IF(U11="",U9,IF(U13="",U11,IF(U15="",U13,IF(U17="",U15,U17)))))',
     '=IF(U10="",U8,IF(U12="",U10,IF(U14="",U12,IF(U16="",U14,IF(U18="",U16,U18)))))', None, None, None, None, None,
     '_______________________ "___" ___ 202__ г.', None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Вед.геолог ЦДНГ', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'параметры работы (или причины не запуска), заполняет вед.геолог  ЦДНГ (Qж,Qн,%; Qж/Pатм):', None,
     None, None, None, None, None, None, None, None, None, 'м3 /  атм', None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None],
    [None, None, 'Вед.инженер ЦДНГ', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'МОЛ (УСО ГНО) ООО "Башнефть-Добыча"', None, None, None, None, None, None,
     '_____________________________________   "________" _______ 202__ г.', None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'подпись, ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'Вед.инженер ОТКРСиО" ООО Башнефть-Добыча"(по ЭДО)', None, None, None, None, None, None, 0, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
    [None, None, 'ФИО', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None, None, None, None, None, None]]

# import openpyxl

# Загружаем файл Excel
from openpyxl.worksheet.datavalidation import DataValidation
from normir.files_with_list import operation_list
# wb = openpyxl.load_workbook('Книга1.xlsx')
# ws2 = wb['Лист1']

# Список работ
# work_list = [['', '', ' АКТ                                                       ', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'v.2024.2/2', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', 'на сдачу скважины из капитального ремонта', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'бр/час руб.', '', '№ договора', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '9609,18 - КРС; 9128,72 - ПРС; ', '', 'БНД/у/8/53/24/БУР', ''], ['', 'Регион', '', 'Чекмагушевский регион', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'Исполнитель: ', '', '', 'ООО "Ойл-сервис"', '', '', '', '', '', '', '', '', ''], ['', 'ЦДНГ', '', 'ЧЦДНГ 01', '', 'Начало работ:', '', '20.07.2024 22:00 ', '', 'уникальный номер ремонта', '', '', 'причины разделения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '66', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', 'оттяжки ПА', 'макс, в тн (с оттяжками)', '40', 'Подъемник по нормативу на СПО', ''], ['', 'Месторождение:', '', 'Манчаровское', '', 'Окончание работ:', '', '31.07.2024  6:00 ', '', '222222222222222222', '', '', '', '', '', '', '=D11', '', 'Мастер:', '', '', '', 'ГКШ-1500 / ГКШ 300', '111', 'АПРС-40', '', '', '', '', 'АПРС-40 (Оснастка 3×4)', ''], ['', 'Площадь:', '', 'Тамьяновская', '', 'Начало работ:', '', '', '', 'уникальный номер ремонта', '', '', 'причины объединения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '66', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', '', 'макс в тн (РЭ без оттяжек)', '40', 'Подъемник по нормативу на СПО', ''], ['', 'Куст:', '', '', '', 'Окончание работ:', '', '', '', '', '', '', '', '', '', '', '', '', 'Мастер:', '', '', '', 'ГКШ-1500 / ГКШ 300', '', '', '', '', '', '', '', ''], ['', '№ Скважины:', '', '486', '', 'Начало работ:', '', '', '', 'уникальный номер ремонта', '', '', 'причины объединения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', '', '', '', 'Подъемник по нормативу на СПО', ''], ['', 'Признак отказности по отчётности ОРМФ', '', '', 'нет', 'Окончание работ:', '', '', '', '', '', '', '', '', '', '', '', '', 'Мастер:', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Причины отказов', '', '', '', 'Начало работ:', '', '', '', 'уникальный номер ремонта', '', '', 'причины объединения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', '', '', '', 'Подъемник по нормативу на СПО', ''], ['', '', '', '', '', 'Окончание работ:', '', '', '', '', '', '', '', '', '', '', '', '', 'Мастер:', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', 'Начало работ:', '', '', '', 'уникальный номер ремонта', '', '', 'причины объединения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', '', '', '', 'Подъемник по нормативу на СПО', ''], ['', 'Дата отказа', '', '', '', 'Окончание работ:', '', '', '', '', '', '', '', '', '', '', '', '', 'Мастер:', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Дата окон. предыдущего ремонта', '', '', '', 'Начало работ:', '', '', '', 'уникальный номер ремонта', '', '', 'причины объединения', '', 'примечание', '', '№ Скважина', '', 'Бригада:', '', '', '', 'Вид ГКШ', 'гос № ПА', 'Подъемник по факту', '', '', 'факт макс. в тн', '20', 'Подъемник по нормативу на СПО', ''], ['', 'ННО', '', '=IF((D16-D17)<=0,"-",(D16-D17))', '', 'Окончание работ:', '', '', '', '', '', '', '', '', '', '', '', '', 'Мастер:', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '=IF(D18<365,"ДК","-")', '', '', '', 'Время  календ.:', '', '=(H8-H7)*24+(H10-H9)*24+(H12-H11)*24+(H14-H13)*24+(H16-H15)*24+(H18-H17)*24', '', 'оплата по УИН:', '', '', '=J8', '', '№ Скважина', '', '=D11', '', 'Дизельная эл.станция', '', '', '', 'нет', 'Работа МЭС в часах', '', '', '', '=IF(W19="да",J1292-J1287,"0")', '', '', ''], ['', 'запуск в работу при бригаде в конце ремонта', '', '', '', 'Категория скважины', '', '', '=MIN(G24:K24)', 'ЗАВЕРШЕНИЕ РЕМОНТА', '', '', 'да', '', 'инвентарный №', '', '128784', '', '', '', '', '', '', 'Ст-ть часа работы МЭС, руб.', '', '', '', '', '', '', ''], ['', 'Да', '', '', '', 'ПЛАН', 'по Pпл', '', 'по H2S', 'по газовому фактору', '', 'ВИД РЕМОНТА ', '', '=VLOOKUP(X21,ЦИКЛ!AE3:AX245,2,0)', '', '', '', '', '', '', '', '', 'Шифр \nКР/ТР', 'КР4-3', 'Вид ГТМ (плановый)', '', 'прочие', '', 'Вид фонда', '', ''], ['', 'Причины не запуска', '', '', '', '', '1', '', '2', '3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', 'ФАКТ', 'по Pпл', '', 'по H2S', 'по газовому фактору', '', '', '', '=VLOOKUP(X23,ЦИКЛ!AE3:AX245,2,0)', '', '', '', '', '', '', '', '', '', 'КР4-3', 'Вид ГТМ (фактический)', '', 'прочие', '', '', '', ''], ['', 'Примечание', '', '', '', '', '1', '', '2', '3', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', 'Изменение категории скважины с пункта', '', '', '', '', '', 'ВИД РЕМОНТА  \nпри ТР 4-8', '', '', '', '', '', '', '', '', '', '', 'Хоз.процессы', '', '', '', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'О С Н О В Н Ы Е    Д А Н Н Ы Е', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'ДАННЫЕ ПО СКВАЖИНЕ', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Искусственный забой', '', '', '', '1530', '', '', 'данные по УА до ремонта', '', '', '', '', '', 'данные по УА после ремонта', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Текущий забой до ремонта', '', '', '', '205', '', '', 'типоразмер', '', '', 'сост.', 'зав.№', '', 'типоразмер', '', '', 'сост.', 'зав.№', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Текущий забой после ремонта', '', '', '', '', '', '', 'AH1-65х210', '', '', 'удовл', '', '', 'AH1-65х210', '', '', 'удовл', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'П О Д З Е М Н О Е  О Б О Р У Д О В А Н И Е', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', 'Тип поднятого оборудования, марка, мощность напора /\n номер насоса', '', '', '', '', '', 'ЭЦН', '', '', '', '', 'Тип спущенного оборудования, марка, мощность напора / \nномер насоса', '', '', '', 'ШГН', '', '', '', 'Кол-во крепежных поясов, шт.', '', '', '', '', '', '', 'Переводник, №', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'Тип протекторов', '', '', '', '', '', '', 'Переводник, №', '', '', ''], ['', 'Глубина спуска, м.', '', '', '', '', '', '', '', '', '', '', 'Глубина спуска, м.', '', '', '', '', '', '', '', 'Кол-во протекто-ров', '', '60мм', '', '', '', '', 'Подвес. патрубок,№', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '73мм', '', '', '', '', 'Противополетка', '', '', ''], ['', 'Доп.оборудование поднято ', '', '', '', '', '', '', '', '', '', '', 'Длина спущенного кабеля, м.', '', '', '', '', '', '', '', 'Тип сбивного клапана', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'Тип обратного клапана', '', '', '', '', '', '', '', '', '', ''], ['', ' интервал установки, м', '', '', '', '', '', '', '', '', '', '', 'Доп.оборудование   после ТКРС                                                         интервал установки', '', '', '', '', '', '', '', 'Доп.оборудование   после ТКРС                                                         интервал установки', '', '', '', '', '', '', 'противополётное оборудование', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['О б ъ е м    в ы п о л н е н н ы х    р а б о т', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['№', 'Дата', '', 'Цикл ремонта', 'Разбивка цикла ремонта', 'Вид работ', '', '', '', '', '', '', '', '', 'Акт', 'Химия', 'реагенты', 'тн/шт', 'Источник норм: ', 'Простои/глушение', 'Ед.изм.', 'Кол-во', 'Норма на\nедин.(час)', 'Коэф.', 'Время по норме, \nчас.', 'Время к оплате', 'Исключенное Время', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', 'ОТКРСиО', 'УСРСиСТ', 'ПТО РУДНГ', 'ОГТМ/ОР', 'Причина снятия'], ['Объём работ по основному плану', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'Переезд', '', 'ПЗР к переезду (инструктаж (в том числе инструктаж по ТБ), уточнение маршрута транспортировки, монтаж-демонтаж крана, выбор площадки под загрузку-разгрузку, маневрирование техники на кусту, построение техники в колонну)', '', '', '', '', '', '', '', '', '', '', '', '', '§302разд1', '', 'опер.', '1', '1.167', '1', '=V47*W47*X47', '=Y47-AA47-AB47-AC47-AD47', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'Переезд', '', 'Переезд cо скв скв №6322 Юсуповского м-я  на скв 486 Тамьяновская 60км (6 единиц спец.техники (полное звено))', '', '', '', '', '', '', '', '', '', '', '', '', '§302разд1', '', 'км', '60', '0.028', '1', '=V48*W48*X48', '=Y48-AA48-AB48-AC48-AD48', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'Переезд', '', 'Отцепка-прицепка  (2 вагона, инструменталка, ЕДК, аквтонаматыватель, беспилотник)', '', '', '', '', '', '', '', '', '', '', '', '', '§304разд1', '', 'раз', '5', '0.067', '1', '=V49*W49*X49', '=Y49-AA49-AB49-AC49-AD49', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Расставить оборудование на кусту скважины согласно схемы ', '', '', '', '', '', '', '', '', '', '', '', '', '§16разд.1', '', 'раз', '5', '0.09', '1', '=V50*W50*X50', '=Y50-AA50-AB50-AC50-AD50', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж лестниц и перильных ограждений на вагонах и техемкостях ', '', '', '', '', '', '', '', '', '', '', '', '', '§86разд.1', '', 'шт', '5', '0.17', '1', '=V51*W51*X51', '=Y51-AA51-AB51-AC51-AD51', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Разгрузка оборудования', '', '', '', '', '', '', '', '', '', '', '', '', '§299разд.1', '', 'шт', '1', '1.58', '1', '=V52*W52*X52', '=Y52-AA52-AB52-AC52-AD52', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Установка  заземления (2 куль.будки, доливная, мостки, площадка,2 щита, ПА)(с испытанием )', '', '', '', '', '', '', '', '', '', '', '', '', '§33разд.1', '', 'шт', '10', '0.1', '1', '=V53*W53*X53', '=Y53-AA53-AB53-AC53-AD53', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'Тех.операции', '', 'Сборка линии долива', '', '', '', '', '', '', '', '', '', '', '', '', '§18разд.1', '', 'шт', '1', '0.22', '1', '=V54*W54*X54', '=Y54-AA54-AB54-AC54-AD54', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Размотать электрокабель и подключить оборудование к электросети ', '', '', '', '', '', '', '', '', '', '', '', '', '§34разд.1', '', 'шт', '4', '0.12', '1', '=V55*W55*X55', '=Y55-AA55-AB55-AC55-AD55', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Поднести подставки и подставить их под электрокабель  (1шт.)', '', '', '', '', '', '', '', '', '', '', '', '', '§35разд.1', '', 'раз', '10', '0.030000000000000002', '1', '=V56*W56*X56', '=Y56-AA56-AB56-AC56-AD56', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Подключение прожектора', '', '', '', '', '', '', '', '', '', '', '', '', '§36разд.1', '', 'шт', '1', '0.08', '1', '=V57*W57*X57', '=Y57-AA57-AB57-AC57-AD57', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж/демонтаж и проверка работы видеорегистратора', '', '', '', '', '', '', '', '', '', '', '', '', '§38 разд.1', '', 'раз', '1', '0.2', '1', '=V58*W58*X58', '=Y58-AA58-AB58-AC58-AD58', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж ИВЭ-50', '', '', '', '', '', '', '', '', '', '', '', '', '§25разд.1', '', 'раз', '1', '0.17', '1', '=V59*W59*X59', '=Y59-AA59-AB59-AC59-AD59', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж и демонтаж автокрана ', '', '', '', '', '', '', '', '', '', '', '', '', '§32разд.1', '', 'шт', '1', '0.26666666666666666', '1', '=V60*W60*X60', '=Y60-AA60-AB60-AC60-AD60', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Разгрузка и раскатка труб на мостках ', '', '', '', '', '', '', '', '', '', '', '', '', '§39разд.1', '', 'шт', '100', '0.008', '1', '=V61*W61*X61', '=Y61-AA61-AB61-AC61-AD61', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж  дополнительных стеллажей для труб (облегченная конструкция)', '', '', '', '', '', '', '', '', '', '', '', '', '§57разд.1', '', 'раз', '3', '0.0867', '1', '=V62*W62*X62', '=Y62-AA62-AB62-AC62-AD62', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж и демонтаж передвижного приемного мостка', '', '', '', '', '', '', '', '', '', '', '', '', '§58разд.1', '', 'шт', '1', '0.92', '1', '=V63*W63*X63', '=Y63-AA63-AB63-AC63-AD63', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'Тех.операции', '', 'Установить заглушку на коллектор', '', '', '', '', '', '', '', '', '', '', '', '', '§28разд.1', '', 'шт', '1', '0.1', '1', '=V64*W64*X64', '=Y64-AA64-AB64-AC64-AD64', '', '', '', '', ''], ['=ROW()-ROW($A$46)', '20.07.2024', '', 'ПР.перед.ремонтом', '', 'Монтаж подъемника АПРС-40', '', '', '', '', '', '', '', '', '', '', '', '', '§62разд.1', '', 'раз', '1', '1.467', '1', '=V65*W65*X65', '=Y65-AA65-AB65-AC65-AD65', '', '', '', '', ''], ['=ROW()-ROW($A$46)', 'КРС', '', 'ПР.перед.ремонтом', '', 'Работа приёмной комиссии', '', '', '', '', '', '', '', '', '', '', '', '', '§302разд.1', '', 'раз', '1', '1.75', '1', '=V66*W66*X66', '=Y66-AA66-AB66-AC66-AD66', '', '', '', '', ''], ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '=V67*W67*X67', '=Y67-AA67-AB67-AC67-AD67', '', '', '', '', '']]
#
#
#
# stop_str = len(work_list)
# ind_ins = 2 # Начало вставки данных с 2-й строки
# oper_list = '"{}"'.format(','.join(operation_list))
# print( len(oper_list))
#
#
#
#
# for i in range(1, stop_str + 1):
#     for j in range(1, len(work_list[0]) + 1): # Изменено на len(work_list[0]) + 1
#         cell = ws2.cell(row=i, column=j)
#         if cell and str(cell) != str(work_list[i - 1][j - 1]):
#             if work_list[i - 1][j - 1]:
#                 cell.value = work_list[i - 1][j - 1]
#
#                 if cell.value in operation_list:
#                     # Добавляем валидацию данных к ячейке
#                     rule = DataValidation(type="list", formula1=oper_list, allow_blank=True)
#                     ws2.add_data_validation(rule)
#                     rule.ranges = cell.coordinate
#
#
#
#
# ws2.merge_cells(start_row=54, start_column=2, end_row=54, end_column=3)
#
# # ws2.merge_cells(start_row=54, start_column=4, end_row=54, end_column=5)
# # ws2.merge_cells(start_row=64, start_column=4, end_row=64, end_column=5)
#
# # Сохраняем файл
# wb.save('your_excel_file.xlsx')


# fname = '1234.xlsx'
# wb2 = load_workbook(fname, keep_vba=True)
# name_list = wb2.sheetnames
# # print(name_list)
# ws2 = wb2.active
# lifting_norm_nkt = {}
if fname:
    normir_list = []
    for row_ind, row in enumerate(ws2.iter_rows(max_col=31, min_row=693, max_row=725)):
        lst = []
        for col_index, col in enumerate(row):
           lst.append(col.value)
        normir_list.append(lst)
print(normir_list)
#         if row_ind == 0:
#             lifting = row[0]
#             lifting_1 = row[8]
#         if row_ind == 17:
#             lifting_norm_nkt.setdefault(lifting.value, {}).setdefault('19', row[3].value)
#             lifting_norm_nkt.setdefault(lifting.value, {}).setdefault('раздел', row[5].value)
#             lifting_norm_nkt.setdefault(lifting_1.value, {}).setdefault('19', row[11].value)
#             lifting_norm_nkt.setdefault(lifting_1.value, {}).setdefault('раздел', row[13].value)
#
#         elif row_ind == 18:
#             lifting_norm_nkt.setdefault(lifting.value, {}).setdefault('22', row[3].value)
#             lifting_norm_nkt.setdefault(lifting.value, {}).setdefault('раздел', row[5].value)
#             lifting_norm_nkt.setdefault(lifting_1.value, {}).setdefault('25', row[11].value)
#             lifting_norm_nkt.setdefault(lifting_1.value, {}).setdefault('раздел', row[13].value)
# print(lifting_norm_nkt)

# if fname:
#     normir_list = []
#     for row_ind, row in enumerate(ws2.iter_rows(max_col=18, min_row=235, max_row=235)):
#         lst = []
#
#         if row[0].value:
#             if 'спуск' in row[0].value:
#                 for i in ['АЗИНМАШ-37А (Оснастка 2×3)', 'АПРС-32 (Оснастка 2×3)', 'АПРС-40 (Оснастка 2×3)', 'АПРС-40 (Оснастка 3×4)', 'АПРС-50 (Оснастка 3×4)', 'АПР60/80 (Оснастка 3×4)', 'БАРС60/80 (Оснастка 3×4)', 'УПТ-32 (Оснастка 3×4)']:
#
#                     lifting_norm_nkt.setdefault(i, {}).\
#                         setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                     lifting_norm_nkt.setdefault(i, {}). \
#                         setdefault(row[1].value, {}).setdefault('7.6-8.5', row[4].value)
#                     lifting_norm_nkt.setdefault(i, {}). \
#                         setdefault(row[1].value, {}).setdefault('8.6-9.5', row[5].value)
#                     lifting_norm_nkt.setdefault(i, {}). \
#                         setdefault(row[1].value, {}).setdefault('9.6-10.5', row[6].value)
#                     lifting_norm_nkt.setdefault(i, {}). \
#                         setdefault(row[1].value, {}).setdefault('10.6-11.5', row[7].value)
#                     # lifting_norm_nkt.setdefault(i, {}). \
#                     #     setdefault(row[1].value, {}).setdefault(row[2].value, {}).setdefault('11.6-12.5', (row[8].value, row[14].value))
#                     lifting_norm_nkt.setdefault(i, {}). \
#                         setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
# print(lifting_norm_nkt)
# if fname:
#     normir_list = []
#     for row_ind, row in enumerate(ws2.iter_rows(max_col=18, min_row=3, max_row=148)):
#         lst = []
#
#         if row[0].value:
#             if 'спуск' in str(row[0].value) and 10 < row_ind < 51:
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АЗИНМАШ-37А (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-32 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 2×3)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#             elif 'спуск' in str(row[0].value) and 51 < row_ind < 69:
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-40 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#             elif 'спуск' in str(row[0].value) and 69 < row_ind < 87:
#
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПРС-50 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#             elif 'спуск' in str(row[0].value) and 87 < row_ind < 105:
#
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('АПР60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('БАРС60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#
#             elif 'спуск' in str(row[0].value) and 105 < row_ind < 126:
#
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПА-60/80 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#             elif 'спуск' in str(row[0].value) and 126 < row_ind < 147:
#
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('6.5-7.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('7.6-8.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('8.6-9.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('9.6-10.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('10.6-11.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('11.6-12.5', row[3].value)
#                 lifting_norm_nkt.setdefault('УПТ-32 (Оснастка 3×4)', {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[15].value)
#
# print(lifting_norm_nkt)


# fname = '1234.xlsx'
# wb2 = load_workbook(fname, keep_vba=True)
# name_list = wb2.sheetnames
# # print(name_list)
# ws2 = wb2.active
# lifting_norm_nkt = {}
# if fname:
#     normir_list = []
#     for row_ind, row in enumerate(ws2.iter_rows(max_col=18, min_row=209, max_row=228)):
#         lst = []
#         if row[0].value:
#             if 'спуск' not in row[0].value and 'ПА' not in row[0].value \
#                     and 'ЭЦН' not in row[0].value and 'Штанги' not in row[0].value and 'СБТ' not in row[0].value:
#
#                 lifting_norm_nkt.setdefault(row[0].value, {}).\
#                     setdefault(row[1].value, {}).setdefault(row[2].value, (row[3].value, row[4].value))
#
#                 lifting_norm_nkt.setdefault(row[0].value, {}). \
#                     setdefault(row[1].value, {}).setdefault('раздел', row[5].value)
#
#                 lifting_norm_nkt.setdefault(row[8].value, {}).\
#                     setdefault(row[9].value, {}).setdefault(row[10].value, (row[11].value, row[12].value))
#
#                 lifting_norm_nkt.setdefault(row[8].value, {}). \
#                     setdefault(row[9].value, {}).setdefault('раздел', row[13].value)
#


#
# {'АЗИНМАШ-37А (Оснастка 2×3)': {19: {'III': (0.0116, 423), 'раздел': 'п.1.1 раздел 2'},
#                                 22: {'III': (0.0133, 317), 'раздел': 'п.1.1 раздел 2'},
#                                 25: {'III': (0.0133, 242), 'раздел': 'п.1.1 раздел 2'}},
#  'АПРС-40 (Оснастка 3×4)': {19: {'III': (0.015, 429), 'раздел': 'п.1.1 раздел 2'},
#                             22: {'III': (0.0166, 347), 'раздел': 'п.1.1 раздел 2'},
#                             25: {'III': (0.0166, 307), 'раздел': 'п.1.1 раздел 2'}},
#  'АПРС-50 (Оснастка 3×4)': {19: {'III': (0.0116, 438), 'раздел': 'п.1.1 раздел 2'},
#                             22: {'III': (0.0133, 426), 'раздел': 'п.1.1 раздел 2'},
#                             25: {'III': (0.0133, 326), 'раздел': 'п.1.1 раздел 2'}}}

#
#
# print(lifting_norm_nkt)


#
#
# fname = '23 Гордеевское КР4-3 (Ойл-С, бр.139) 01.06.2023.xlsm'
# wb2 = Workbook()
# ws2 = wb2.get_sheet_by_name('Sheet')
# if fname:
#     wb_summary = load_workbook(fname, keep_vba=True)
#
#     name_list = wb_summary.sheetnames
#     for name in name_list:
#         ws_summar = wb_summary[name]
#         title_list = ws_summar.title
#
#         normir_list = []
#         for row_ind, row in enumerate(ws_summar.iter_rows(max_col=55)):
#             list = []
#
#             for col, value in enumerate(row):
#                 list.append(value.value)
#             normir_list.append(list)
#
#         wb2.create_sheet(title_list)
#
#         print(title_list)
#
#         for i in range(len(normir_list)):
#             print(i)
#             for j in range(len(normir_list[i])):
#                 cell = ws2.cell(row=i+1, column=j+1)
#                 cell.value = normir_list[i][j]
#
# wb2.save('1234.xlsx')
