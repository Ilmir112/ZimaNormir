# -*- coding: utf-8 -*-

import os
import sys
import socket
import psutil

import re
import threading

from openpyxl.reader.excel import load_workbook
from PyQt5.QtWidgets import QApplication, QMainWindow, QMenu, QMenuBar, QAction, QTableWidget, \
    QLineEdit, QFileDialog, QToolBar, QPushButton, QMessageBox, QTabWidget, QTableWidgetItem
from PyQt5 import QtCore, QtWidgets

from openpyxl.utils import get_column_letter

from log_files.log import logger, QPlainTextEditLogger
import well_data

from PyQt5.QtCore import Qt, QObject, QThread, pyqtSlot, pyqtSignal

from normir.alone_oreration import count_row_height
from users.login_users import LoginWindow

from openpyxl import load_workbook
from PyQt5.QtGui import QColor

class OpenTemplateNorm(QThread):
    finished = pyqtSignal()

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        print("Поток запущен")
        while not self.isInterruptionRequested():
            # Попытка открыть Excel файл
            try:
                well_data.workbook = load_workbook(self.file_path, keep_vba=True)
                well_data.sheet = well_data.workbook.active

                print(f"Файл '{self.file_path}' успешно загружен!")
                self.finished.emit()  # Уведомление о завершении работы

                # Здесь можно добавить код для обработки данных в файле
                break  # Завершаем цикл, так как файл успешно загружен

            except Exception as esk:
                print(f"Ошибка при открытии файла: {esk}")
                self.finished.emit()  # Уведомление о завершении работы
                break  # Завершаем цикл при ошибке

        print("Поток завершен")

    @staticmethod
    def open_excel():
        return well_data.workbook, well_data.sheet

    def start_loading(self):
        self.start()  # Запустить поток

    def stop(self):
        self.requestInterruption()  # Запрос на прерывание потока
        self.wait()  # Ждем завершения потока
        print("Поток успешно завершен")




class UncaughtExceptions(QObject):
    _exception_caught = pyqtSignal(object)

    def __init__(self):
        super().__init__()

    @pyqtSlot(object)
    def handleException(self, ex):
        try:
            logger.critical(f"{well_data.well_number._value} {well_data.well_area._value} Критическая ошибка: {ex}")
        except:
            logger.critical(f"{well_data.well_number} {well_data.well_area} Критическая ошибка: {ex}")


class MyWindow(QMainWindow):

    def __init__(self):
        super().__init__()

        self.initUI()

        self.login_window = None
        self.new_window = None
        self.raid_window = None
        self.leakage_window = None
        self.correct_window = None
        self.acid_windowPaker = None
        self.work_window = None
        self.signatures_window = None
        self.acid_windowPaker2 = None
        self.rir_window = None
        self.data_window = None
        self.filter_widgets = []
        self.table_class = None
        self.table_juming = None
        self.resize(1400, 800)

        self.perforation_correct_window2 = None
        self.ws = None
        self.ins_ind = None
        self.perforation_list = []
        self.dict_perforation_project = {}

        self.ins_ind_border = None
        self.work_plan = 0
        self.table_widget = None
        self.table_pvr = None

        # threading.Timer(2.0, self.close_splash).start()

        self.log_widget = QPlainTextEditLogger(self)
        logger.addHandler(self.log_widget)
        self.setCentralWidget(self.log_widget.widget)

        # Обработка критических ошибок
        self.excepthook = UncaughtExceptions()
        self.excepthook._exception_caught.connect(self.excepthook.handleException)

        # # # Запускаем обработчик исключений в отдельном потоке
        # self.thread = QThread()
        # self.excepthook.moveToThread(self.thread)
        # # self.thread.started.connect(self.excepthook.handleException)
        # self.thread.start()

    @staticmethod
    def check_process():
        count_zima = 0
        for proc in psutil.process_iter():
            if proc.name() == 'ZIMA.exe':
                count_zima += count_zima
        if count_zima > 1:
            return True  # Процесс найден

        return False  # Процесс не найден

    @staticmethod
    def close_process():
        for proc in psutil.process_iter():
            if proc.name() == 'ZIMA.exe':
                proc.terminate()  # Принудительное завершение

    @staticmethod
    def show_confirmation():
        reply = QMessageBox.question(None, 'Закрыть Zima?',
                                     'Приложение Zima.exe работает. Вы хотите Перезапустить его?',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if reply == QMessageBox.Yes:
            MyWindow.close_process()

    @staticmethod
    def check_connection(host, port=5432):
        """Проверяет соединение с удаленным хостом по указанному порту."""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(2)  # Устанавливаем таймаут 2 секунды
            sock.connect((host, port))
            sock.close()
            return True
        except socket.error:
            return False

    # Остальная часть кода...

    def initUI(self):

        self.table_widget = None

        self.workbook, self.sheet = None, None

        self.setWindowTitle("ZIMA-Normir")
        self.setGeometry(200, 100, 800, 800)

        self.createMenuBar()
        self.le = QLineEdit()

        self.toolbar = QToolBar()
        self.addToolBar(self.toolbar)

        self.saveFileButton = QPushButton("Сохранить проект")
        self.saveFileButton.clicked.connect(self.save_to_krs)
        self.toolbar.addWidget(self.saveFileButton)

        self.correct_curator_Button = QPushButton("Скорректировать куратора")
        self.correct_curator_Button.clicked.connect(self.correct_curator)
        self.toolbar.addWidget(self.correct_curator_Button)

        self.closeFileButton = QPushButton("Закрыть проект")
        self.closeFileButton.clicked.connect(self.close_file)
        self.toolbar.addWidget(self.closeFileButton)

    def correct_curator(self):
        from normir.curators import SelectCurator

        if self.new_window is None:
            self.new_window = SelectCurator()
            # WellCondition.leakage_window.setGeometry(200, 400, 300, 400)
            self.set_modal_window(self.new_window)
            MyWindow.pause_app()
            well_data.pause = True
            self.new_window = None  # Discard reference.
        else:
            self.new_window.close()  # Close window.
            self.new_window = None  # Discard reference.

    def createMenuBar(self):
        self.menuBar = QMenuBar(self)
        self.setMenuBar(self.menuBar)
        self.fileMenu = QMenu('&Файл', self)

        self.signatories = QMenu('&Подписанты ', self)
        self.menuBar.addMenu(self.fileMenu)
        # self.menuBar.addMenu(self.application_geophysical)
        # self.menuBar.addMenu(self.classifierMenu)
        self.menuBar.addMenu(self.signatories)

        self.create_file = self.fileMenu.addMenu('&Создать')
        self.create_file_normir = self.fileMenu.addMenu('&Нормирование')
        self.create_file_normir_new = self.create_file_normir.addAction('Новый', self.action_norm_clicked)

        # self.open_file = self.fileMenu.addAction('Открыть', self.action_clicked)
        # self.save_file = self.fileMenu.addAction('Сохранить', self.action_clicked)
        # self.save_file_as = self.fileMenu.addAction('Сохранить как', self.action_clicked)

    @QtCore.pyqtSlot()
    def action_norm_clicked(self):
        from open_pz import CreatePZ

        from normir.normir_excel import normir_excel_dict
        from normir.main_normir import NormirWindow
        from data_base.work_with_base import insert_data_new_excel_file
        action = self.sender()

        if action == self.create_file_normir_new and self.table_widget is None:

            self.work_plan = 'normir_new'
            self.tableWidgetOpenNormir(self.work_plan)
            self.fname, _ = QtWidgets.QFileDialog.getOpenFileName(self, 'Выберите файл', '.',
                                                                  "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)")
            if self.fname:

                try:
                    self.read_pz(self.fname)
                    well_data.pause = True

                    self.file_path = 'property_excel/template_normir_new.xlsm'

                    # Создаем экземпляр класса и запускаем загрузку
                    excel_loader = OpenTemplateNorm(self.file_path)
                    excel_loader.start_loading()

                    read_pz = CreatePZ(self.wb, self.ws, self.data_window, self.perforation_correct_window2)

                    sheet = read_pz.open_excel_file(self.ws, self.work_plan)

                    data, rowHeights, colWidth, boundaries_dict = \
                        normir_excel_dict['data'], normir_excel_dict['rowHeights'], \
                        normir_excel_dict['colWidth'], normir_excel_dict['merged_cells']

                    self.ws = insert_data_new_excel_file(data, rowHeights, colWidth, boundaries_dict)
                    self.copy_norm(self.ws, self.table_widget, self.work_plan)

                    self.norm_window = NormirWindow(self.table_widget, self.work_plan)
                    # self.rir_window.setGeometry(200, 400, 100, 200)
                    self.norm_window.show()

                    well_data.pause = True
                    self.pause_app()


                except FileNotFoundError as f:
                    QMessageBox.warning(self, 'Ошибка', f'Ошибка при прочтении файла {f}')

    @QtCore.pyqtSlot()
    def tableWidgetOpenNormir(self, work_plan):

        if self.table_widget is None:
            # Создание объекта TabWidget
            self.tabWidget = QTabWidget()
            self.table_widget = QTableWidget()

            self.table_widget.setContextMenuPolicy(QtCore.Qt.ContextMenuPolicy.CustomContextMenu)
            self.table_widget.customContextMenuRequested.connect(self.openContextMenuNormir)
            self.setCentralWidget(self.tabWidget)
            self.model = self.table_widget.model()

            # Этот сигнал испускается всякий раз, когда ячейка в таблице нажата.
            # Указанная строка и столбец - это ячейка, которая была нажата.
            self.table_widget.cellPressed[int, int].connect(self.clickedRowColumn)
            self.tabWidget.addTab(self.table_widget, 'Нормирование')

    def saveFileDialog(self, wb2, full_path):

        try:
            file_name, _ = QFileDialog.getSaveFileName(self, "Save excel-file",
                                                       f"{full_path}", "Excel Files (*.xlsm)")
            if file_name:
                wb2.save(file_name)
        except Exception as e:
            mes = QMessageBox.critical(self, 'Ошибка',
                                       f'файл под таким именем открыт, закройте его: {type(e).__name__}\n\n{str(e)}')
            return
        # try:
        #     # Создаем объект Excel
        #     excel = win32com.client.Dispatch("Excel.Application")
        #     # Открываем файл
        #     workbook = excel.Workbooks.Open(file_name)
        #     # Выбираем активный лист
        #     worksheet = workbook.ActiveSheet
        #
        #     # Назначаем область печати с колонок B до L
        #     worksheet.PageSetup.PrintArea = "B:L"
        #
        # except Exception as e:
        #     print(f"Ошибка при работе с Excel: {type(e).__name__}\n\n{str(e)}")

    def save_to_krs(self):

        merged_cells = []

        if self.table_widget:
            work_list = []
            row_count = self.table_widget.rowCount()
            for row in range(row_count+1):
                row_lst = []
                # self.ins_ind_border += 1
                for column in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(row, column)
                    if item:
                        if (self.table_widget.rowSpan(row, column) > 1 or self.table_widget.columnSpan(row, column) > 1) and row >= 46:
                            merged_cells.append((row, column))

                        if self.check_str_isdigit(item.text()):
                            row_lst.append(item.text().replace(',', '.'))
                        else:
                            row_lst.append(item.text())
                    else:
                        row_lst.append("")
                if row >= 46 and ('роработка' in row_lst[7] or 'Закачка в НКТ' in row_lst[7] or
                                  'Вымывание шара' in row_lst[7] or 'Продавка ГР' in row_lst[7]
                                  or 'Разбуривание на первые' in row_lst[7]
                                  or 'Разбуривание на каждые' in row_lst[7]):
                    row_lst[24] = f'AС{row + 1}'
                    row_lst[25] = f'=Y{row + 1}-AA{row + 1}-AB{row + 1}-AC{row + 1}-AD{row + 1}'
                elif row >= 46 and 'ПЗР+Отсыпка' in row_lst[7]:
                    row_lst[24] = '=((1*W1167)+5+(((2.6/400)*0.75)*V1167)+2+1+(1.6*W1167)+(((2.6/400)*0.75)*V1167)+1+39+27)/60'
                    row_lst[25] = f'=Y{row + 1}-AA{row + 1}-AB{row + 1}-AC{row + 1}-AD{row + 1}'
                elif row >= 46 and 'Заполнить колонны труб водой для проверки работы' in row_lst[7]:
                    row_lst[24] = '=ROUNDUP(SUM((V984*0.00058)+0.06),2)'
                    row_lst[25] = '=ROUNDUP(Y984-AA984-AB984-AC984-AD984,2)'
                elif row >= 46:
                    a = row_lst
                    row_lst[24] = f'=V{row+1}*W{row+1}*X{row+1}'
                    row_lst[25] = f'=Y{row+1}-AA{row+1}-AB{row+1}-AC{row+1}-AD{row+1}'
                    b = row_lst
                work_list.append(row_lst)

            merged_cells_dict = {}
            n = 205
            # ws2.merge_cells(start_column=value[0], start_row=value[1],
            #                 end_column=value[2], end_row=value[3])
            # # print(f' индекс объ {ins_ind}')
            for row in merged_cells:
                if row[0] >= 46:
                    merged_cells_dict.setdefault(row[0], []).append(row[1])
            merge_dict2 = {}
            for key, original_list in merged_cells_dict.items():
                result = []
                current_sublist = []

                for num in original_list:
                    if not current_sublist or (num - current_sublist[-1] <= 1):
                        current_sublist.append(num)
                    else:
                        result.append(current_sublist)
                        current_sublist = [num]

                if current_sublist:  # Добавляем последнюю подпоследовательность, если она не пустая
                    result.append(current_sublist)

                for k in result:
                    merge_dict2[n] = (min(k) + 1, key +1, max(k) + 1, key + 1)
                    n += 1


            count_row_height(well_data.sheet, work_list, merge_dict2)

            well_data.itog_ind_min = 46
            well_data.itog_ind_max = len(work_list)

            # self.wb2.save('1234.xlsm')
            # print(f' длина {len(work_list)}')
            # CreatePZ.add_itog(self, ws2, self.table_widget.rowCount() + 1, self.work_plan)

            # try:
            # for row_ind, row in enumerate(self.ws.iter_rows(values_only=True)):
            #     if 15 < row_ind < 100:
            #         if all(cell in [None, ''] for cell in row) \
            #                 and ('Интервалы темпа' not in str(ws2.cell(row=row_ind, column=2).value) \
            #                      and 'Замечания к эксплуатационному периоду' not in str(
            #                     ws2.cell(row=row_ind, column=2).value) \
            #                      and 'Замечания к эксплуатационному периоду' not in str(
            #                     ws2.cell(row=row_ind - 2, column=2).value)):
            #             # print(row_ind, ('Интервалы темпа' not in str(ws2.cell(row=row_ind, column=2).value)),
            #             #       str(ws2.cell(row=row_ind, column=2).value))
            #             ws2.row_dimensions[row_ind + 1].hidden = True

            # ws2.print_area = f'B1:L{self.table_widget.rowCount() + 45}'
            # ws2.page_setup.fitToPage = True
            # ws2.page_setup.fitToHeight = False
            # ws2.page_setup.fitToWidth = True
            # ws2.print_options.horizontalCentered = True
            # # зададим размер листа
            # ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            # # содержимое по ширине страницы
            # ws2.sheet_properties.pageSetUpPr.fitToPage = True
            # ws2.page_setup.fitToHeight = False

            filenames = f"{well_data.well_number._value} {well_data.well_area._value} " \
                        f"{well_data.type_kr.split(' ')[0]} БР№{well_data.brigade_number} " \
                        f"{well_data.date_end}"
            full_path = filenames

            # # Перед сохранением установите режим расчета
            # wb2.calculation.calcMode = "auto"

            if well_data.workbook:
                # self.wb2.close()
                self.saveFileDialog(well_data.workbook, full_path)
                # self.wb.save(f'{full_path}.xlsm')
                print(f"Table data saved to Excel {full_path} {well_data.number_dp}")

    def check_str_isdigit(self, string):

        # Паттерн для проверки: допустимы только цифры, точка и запятая
        pattern = r'^[\d.,]+$'

        # Проверка строки на соответствие паттерну
        if re.match(pattern, string):
            return True
        else:
            return False

    def close_file(self):
        from find import ProtectedIsNonNone, ProtectedIsDigit

        temp_folder = r'C:\Windows\Temp'

        try:
            for filename in os.listdir(temp_folder):
                file_path = os.path.join(temp_folder, filename)
                # Удаляем только файлы, а не директории
                if os.path.isfile(file_path):
                    os.remove(file_path)

        except Exception as e:
            QMessageBox.critical(window, "Ошибка", f"Не удалось очистить папку с временными файлами: {e}")

        if not self.table_widget is None:
            self.table_widget.clear()
            self.table_widget.resizeColumnsToContents()
            self.table_widget = None
            self.tabWidget = None
            well_data.column_head_m = ''
            well_data.date_drilling_cancel = ''
            well_data.date_drilling_run = ''
            well_data.wellhead_fittings = ''
            well_data.dict_perforation_short = {}
            well_data.plast_work_short = []
            self.table_widget = None
            well_data.normOfTime = 0
            well_data.gipsInWell = False
            well_data.grp_plan = False
            well_data.bottom = 0
            well_data.nkt_opressTrue = False
            well_data.bottomhole_drill = ProtectedIsNonNone(0)
            well_data.open_trunk_well = False
            well_data.normOfTime = 0
            well_data.lift_ecn_can = False
            well_data.pause = True
            well_data.check_data_in_pz = []
            well_data.sucker_rod_none = True
            well_data.curator = '0'
            well_data.lift_ecn_can_addition = False
            well_data.column_passability = False
            well_data.column_additional_passability = False
            well_data.template_depth = 0
            well_data.gnkt_number = 0
            well_data.gnkt_length = 0
            well_data.diametr_length = 0
            well_data.iznos = 0
            well_data.pipe_mileage = 0
            well_data.pipe_fatigue = 0
            well_data.pvo = 0
            well_data.previous_well = 0
            well_data.b_plan = 0
            well_data.pipes_ind = ProtectedIsDigit(0)
            well_data.sucker_rod_ind = ProtectedIsDigit(0)
            well_data.expected_Q = 0
            well_data.expected_P = 0
            well_data.plast_select = ''
            well_data.dict_perforation = {}
            well_data.dict_perforation_project = {}
            well_data.itog_ind_min = 0
            well_data.kat_pvo = 2
            well_data.gaz_f_pr = []
            well_data.paker_layout = 0
            well_data.cat_P_P = []
            well_data.column_direction_diametr = ProtectedIsNonNone('не корректно')
            well_data.column_direction_wall_thickness = ProtectedIsNonNone('не корректно')
            well_data.column_direction_lenght = ProtectedIsNonNone('не корректно')
            well_data.column_conductor_diametr = ProtectedIsNonNone('не корректно')
            well_data.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
            well_data.column_conductor_lenght = ProtectedIsNonNone('не корректно')
            well_data.column_additional_diametr = ProtectedIsNonNone('не корректно')
            well_data.column_additional_wall_thickness = ProtectedIsNonNone('не корректно')
            well_data.head_column_additional = ProtectedIsNonNone('не корректно')
            well_data.shoe_column_additional = ProtectedIsNonNone('не корректно')
            well_data.column_diametr = ProtectedIsNonNone('не корректно')
            well_data.column_wall_thickness = ProtectedIsNonNone('не корректно')
            well_data.shoe_column = ProtectedIsNonNone('не корректно')
            well_data.bottomhole_artificial = ProtectedIsNonNone('не корректно')
            well_data.max_expected_pressure = ProtectedIsNonNone('не корректно')
            well_data.head_column_additional = ProtectedIsNonNone('не корректно')
            well_data.leakiness_Count = 0
            well_data.bur_rastvor = ''
            well_data.data, well_data.rowHeights, well_data.colWidth, well_data.boundaries_dict = '', '', '', ''
            well_data.data_in_base = False
            well_data.well_volume_in_PZ = []
            well_data.expected_pick_up = {}
            well_data.current_bottom = 0
            well_data.emergency_bottom = well_data.current_bottom
            well_data.fluid_work = 0
            well_data.groove_diameter = ''
            well_data.static_level = ProtectedIsNonNone('не корректно')
            well_data.dinamic_level = ProtectedIsNonNone('не корректно')
            well_data.work_perforations_approved = False
            well_data.dict_leakiness = {}
            well_data.leakiness = False
            well_data.emergency_well = False
            well_data.angle_data = []
            well_data.emergency_count = 0
            well_data.skm_interval = []
            well_data.work_perforations = []
            well_data.work_perforations_dict = {}
            well_data.paker_do = {"do": 0, "posle": 0}
            well_data.column_additional = False
            well_data.well_number = ProtectedIsNonNone('')
            well_data.well_area = ProtectedIsNonNone('')
            well_data.values = []
            well_data.dop_work_list = None
            well_data.depth_fond_paker_do = {"do": 0, "posle": 0}
            well_data.paker2_do = {"do": 0, "posle": 0}
            well_data.depth_fond_paker2_do = {"do": 0, "posle": 0}
            well_data.perforation_roof = 50000
            well_data.data_x_min = 0
            well_data.perforation_sole = 0
            well_data.dict_pump_SHGN = {"do": '0', "posle": '0'}
            well_data.dict_pump_ECN = {"do": '0', "posle": '0'}
            well_data.dict_pump_SHGN_h = {"do": '0', "posle": '0'}
            well_data.dict_pump_ECN_h = {"do": '0', "posle": '0'}
            well_data.dict_pump = {"do": '0', "posle": '0'}
            well_data.leakiness_interval = []
            well_data.dict_pump_h = {"do": 0, "posle": 0}
            well_data.ins_ind = 0
            well_data.ins_ind2 = 0
            well_data.image_data = []
            well_data.current_bottom2 = 5000
            well_data.len_razdel_1 = 0
            well_data.count_template = 0
            well_data.data_well_is_True = False
            well_data.cat_P_1 = []
            well_data.countAcid = 0
            well_data.first_pressure = ProtectedIsDigit(0)
            well_data.swabTypeComboIndex = 1
            well_data.swab_true_edit_type = 1
            well_data.data_x_max = ProtectedIsDigit(0)
            well_data.drilling_interval = []
            well_data.max_angle = 0
            well_data.pakerTwoSKO = False
            well_data.privyazkaSKO = 0
            well_data.h2s_pr = []
            well_data.cat_h2s_list = []
            well_data.dict_perforation_short = {}
            well_data.h2s_mg = []
            well_data.lift_key = 0
            well_data.max_admissible_pressure = ProtectedIsNonNone(0)
            well_data.region = ''
            well_data.forPaker_list = False
            well_data.dict_nkt = {}
            well_data.dict_nkt_po = {}
            well_data.data_well_max = ProtectedIsNonNone(0)
            well_data.data_pvr_max = ProtectedIsNonNone(0)
            well_data.dict_sucker_rod = {}
            well_data.dict_sucker_rod_po = {}
            well_data.row_expected = []
            well_data.rowHeights = []
            well_data.plast_project = []
            well_data.plast_work = []
            well_data.leakiness_Count = 0
            well_data.plast_all = []
            well_data.condition_of_wells = ProtectedIsNonNone(0)
            well_data.cat_well_min = ProtectedIsNonNone(0)
            well_data.bvo = False
            well_data.old_version = False
            well_data.image_list = []
            well_data.problemWithEk = False
            well_data.problemWithEk_depth = well_data.current_bottom
            well_data.problemWithEk_diametr = 220
            # path = f"{well_data.path_image}/imageFiles/image_work"[1:]
            #
            # for file in os.listdir(path):
            #     file_path = os.path.join(path, file)
            #     if os.path.isfile(file_path):
            #         os.remove(file_path)

            mes = QMessageBox.information(self, 'Обновление', 'Данные обнулены')

    def on_finished(self):
        print("Работа с файлом Excel завершена.")

    def openContextMenuNormir(self, position):
        context_menu = QMenu(self)
        action_menu = context_menu.addMenu("вид работ")

        simple_technological_menu = QAction('Технологический простой')
        context_menu.addAction(simple_technological_menu)
        simple_technological_menu.triggered.connect(self.simple_technological_work)

        actual_work_menu = QAction('Фактические работы')
        context_menu.addAction(actual_work_menu)
        actual_work_menu.triggered.connect(self.actual_work)


        relocation_menu = QAction('Переезд')
        action_menu.addAction(relocation_menu)
        relocation_menu.triggered.connect(self.relocation_menu)

        jamming_menu = QAction('Глушение скважины')
        action_menu.addAction(jamming_menu)
        jamming_menu.triggered.connect(self.jamming_menu_work)

        lifting_gno_menu = action_menu.addMenu('Подьем ГНО')

        lifting_paker_menu = QAction('Подьем НКТ')
        lifting_gno_menu.addAction(lifting_paker_menu)
        lifting_paker_menu.triggered.connect(self.lifting_paker_menu)

        lifting_shgn_menu = QAction('Подьем штанг')
        lifting_gno_menu.addAction(lifting_shgn_menu)
        lifting_shgn_menu.triggered.connect(self.lifting_shgn_menu)

        spo_template_without_skm_menu = QAction('СПО ПСШ, СПО шаблона, воронки, пера')
        action_menu.addAction(spo_template_without_skm_menu)
        spo_template_without_skm_menu.triggered.connect(self.template_without_skm_action)

        rod_head_action = QAction('СПО Штанголовки')
        action_menu.addAction(rod_head_action)
        rod_head_action.triggered.connect(self.rod_head_action_def)

        raid_action = QAction('СПО Райбера')
        action_menu.addAction(raid_action)
        raid_action.triggered.connect(self.raid_action_def)

        drilling_action = QAction('СПО Долото фрез')
        action_menu.addAction(drilling_action)
        drilling_action.triggered.connect(self.drilling_action_def)

        spo_paker_action = QAction('СПО пакер')
        action_menu.addAction( spo_paker_action )
        spo_paker_action.triggered.connect(self.spo_paker_action)

        pipe_perforation_action = QAction('СПО Трубного перфоратора')
        action_menu.addAction(pipe_perforation_action)
        pipe_perforation_action.triggered.connect(self.pipe_perforation_action)

        rir_with_action = QAction('РИР на пере')
        action_menu.addAction(rir_with_action)
        rir_with_action.triggered.connect(self.rir_with_action)

        work_of_third_parties_action = QAction('Работа сторонних организаций при спущенных НКТ')
        action_menu.addAction(work_of_third_parties_action)
        work_of_third_parties_action.triggered.connect(self.work_of_third_parties_action)

        work_of_third_parties_without_action = QAction('Работа сторонних организаций без НКТ')
        action_menu.addAction(work_of_third_parties_without_action)
        work_of_third_parties_without_action.triggered.connect(self.work_of_third_parties_without_action)

        injection_reagents_action = QAction('Кислота, СКВ, растворитель, незамерз жидкость, блок пачка')
        action_menu.addAction(injection_reagents_action)
        injection_reagents_action.triggered.connect(self.injection_reagents_action)

        alone_menu = action_menu.addMenu('одиночные операции')

        loading_work_action = QAction('Погрузка, выгрузка НКТ или штанг')
        alone_menu.addAction(loading_work_action)
        loading_work_action.triggered.connect(self.loading_work)

        mkp_action = QAction('Копка шахты')
        alone_menu.addAction(mkp_action)
        mkp_action.triggered.connect(self.earthwork_work)

        descent_gno_menu = action_menu.addMenu('Спуск ГНО')

        descent_paker_menu = QAction('спуск НКТ')
        descent_gno_menu.addAction(descent_paker_menu)
        descent_paker_menu.triggered.connect(self.descent_paker_menu)

        descent_shgn_menu = QAction('Спуск штанг')
        descent_gno_menu.addAction(descent_shgn_menu)
        descent_shgn_menu.triggered.connect(self.descent_shgn_menu)

        del_menu = context_menu.addMenu('удаление строки')
        deleteString_action = QAction("Удалить строку", self)
        del_menu.addAction(deleteString_action)
        deleteString_action.triggered.connect(self.deleteString)
        emptyString_action = QAction("добавить пустую строку", self)
        del_menu.addAction(emptyString_action)
        emptyString_action.triggered.connect(self.emptyString)

        context_menu.exec_(self.mapToGlobal(position))
    def earthwork_work(self):
        from normir.earthworks import Earthwor_Window
        if self.raid_window is None:
            self.raid_window = Earthwor_Window(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)

            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def loading_work(self):
        from normir.loading_tubing import LoadingWork
        if self.raid_window is None:
            self.raid_window = LoadingWork(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)

            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def rir_with_action(self):
        from normir.rir_with_pero import RirWithPero
        if self.raid_window is None:
            self.raid_window = RirWithPero(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)

            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def pipe_perforation_action(self):
        from normir.perforation_tubing import PipePerforator
        if self.raid_window is None:
            self.raid_window = PipePerforator(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)

            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def lifting_paker_menu(self):
        from normir.lifting_gno import LiftingWindow
        if self.raid_window is None:
            self.raid_window = LiftingWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None


    def descent_paker_menu(self):
        from normir.descent_gno import DescentGnoWindow
        if self.raid_window is None:
            self.raid_window = DescentGnoWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None


    def close_raid_window(self, event):
        # Вызов обработчика закрытия окна `TemplateWithoutSKM`
        event.accept()  # Разрешаем стандартное закрытие

        # Сброс ссылки на окно
        self.raid_window = None
    def template_without_skm_action(self):
        from normir.template_without_skm import TemplateWithoutSKM

        if self.raid_window is None:
            self.raid_window = TemplateWithoutSKM(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def lifting_shgn_menu(self):
        from normir.lifting_shgn import LiftingShgnWindow
        if self.raid_window is None:
            self.raid_window = LiftingShgnWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def descent_shgn_menu(self):

        from normir.descent_gno import DescentGnoWindow
        if self.raid_window is None:
            self.raid_window = DescentGnoWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def rod_head_action_def(self):
        from normir.rod_head_work import LiftingRodHeadWindow
        if self.raid_window is None:
            self.raid_window = LiftingRodHeadWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def spo_paker_action(self):
        from normir.spo_pakera import SpoPakerAction
        if self.raid_window is None:
            self.raid_window = SpoPakerAction(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def raid_action_def(self):
        from normir.raider_work import RaidWork
        if self.raid_window is None:
            self.raid_window = RaidWork(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def drilling_action_def(self):
        from normir.drilling_work import DrillingWork
        if self.raid_window is None:
            self.raid_window = DrillingWork(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def work_of_third_parties_action(self):
        from normir.work_of_third_parties import WorkOfThirdPaties
        if self.raid_window is None:
            self.raid_window = WorkOfThirdPaties(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None
    def work_of_third_parties_without_action(self):
        from normir.work_of_third_parties_without_nkt import WorkOfThirdPaties
        if self.raid_window is None:
            self.raid_window = WorkOfThirdPaties(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None


    def injection_reagents_action(self):
        from normir.work_of_third_parties_without_nkt import WorkOfThirdPaties
        if self.raid_window is None:
            self.raid_window = WorkOfThirdPaties(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None


    def jamming_menu_work(self):
        from normir.jamming_well import JammingWindow
        if self.raid_window is None:
            self.raid_window = JammingWindow(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def simple_technological_work(self):
        from normir.simple_technological import SimpleWork
        if self.raid_window is None:
            self.raid_window = SimpleWork(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def actual_work(self):
        from normir.relocation_brigade import Relocation_Window
        if self.raid_window is None:
            self.raid_window = Relocation_Window(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    def relocation_menu(self):
        from normir.relocation_brigade import Relocation_Window
        if self.raid_window is None:
            self.raid_window = Relocation_Window(well_data.ins_ind, self.table_widget)
            # self.raid_window.setGeometry(200, 400, 300, 400)

            self.set_modal_window(self.raid_window)
            well_data.pause = True
            self.pause_app()
            well_data.pause = True
            self.raid_window = None
        else:
            self.raid_window.close()  # Close window.
            self.raid_window = None

    @staticmethod
    def set_modal_window(window):
        # Установка модальности окна
        window.setWindowModality(Qt.ApplicationModal)
        window.show()

    def clickedRowColumn(self, r, c):

        self.ins_ind = r + 1
        well_data.ins_ind = r + 1
        # print(r, well_data.count_row_well)
        if r > well_data.count_row_well and 'gnkt' not in self.work_plan:
            data = self.read_clicked_mouse_data(r)

    def read_clicked_mouse_data(self, row):
        pass

        # row = row - well_data.count_row_well
        # # print(well_data.column_diametr._value)
        # data = well_data.data_list
        #
        # well_data.current_bottom = data[row][1]
        # well_data.dict_perforation = json.loads(data[row][2])
        # # print(f' строка {well_data.dict_perforation}')
        #
        # well_data.plast_all = json.loads(data[row][3])
        # well_data.plast_work = json.loads(data[row][4])
        # well_data.dict_leakiness = json.loads(data[row][5])
        # well_data.column_additional = data[row][6]
        #
        # well_data.fluid_work = data[row][7]
        # well_data.template_depth, well_data.template_lenght, well_data.template_depth_addition, well_data.template_lenght_addition = json.loads(
        #     data[row][11])
        # well_data.skm_interval = json.loads(data[row][12])
        #
        # well_data.problemWithEk_depth = data[row][13]
        # well_data.problemWithEk_diametr = data[row][14]

        # print(well_data.skm_interval)

    @staticmethod
    def pause_app():
        while well_data.pause is True:
            QtCore.QCoreApplication.instance().processEvents()

    def read_pz(self, fname):
        self.wb = load_workbook(fname, data_only=True)
        name_list = self.wb.sheetnames
        self.ws = self.wb.active

    def deleteString(self):
        selected_ranges = self.table_widget.selectedRanges()
        selected_rows = []

        if self.ins_ind > well_data.count_row_well:
            # Получение индексов выбранных строк
            for selected_range in selected_ranges:
                top_row = selected_range.topRow()
                bottom_row = selected_range.bottomRow()

                for row in range(top_row, bottom_row + 1):
                    if row not in selected_rows:
                        selected_rows.append(row)

            # Удаление выбранных строк в обратном порядке
            selected_rows.sort(reverse=True)
            # print(selected_rows)
            for row in selected_rows:
                self.table_widget.removeRow(row)


    def emptyString(self):
        if self.ins_ind > well_data.count_row_well:
            ryber_work_list = [
                [None, None, None, None, None, None, None, None, None, None, None, None]]
            self.populate_row(self.ins_ind, ryber_work_list, self.table_widget)

    def populate_row(self, ins_ind, work_list, table_widget, work_plan='krs'):
        if work_list:
            for i, row_data in enumerate(work_list):
                row = ins_ind + i
                table_widget.insertRow(row)

                for column, data in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(data))
                    item.setFlags(item.flags() | Qt.ItemIsEditable)

                    if not data is None:
                        table_widget.setItem(row, column, item)
                    else:
                        table_widget.setItem(row, column, QtWidgets.QTableWidgetItem(str('')))

                    if column == 5:
                        if not data is None:
                            text = data
                            table_widget.setRowHeight(row, int(len(text)/2))
                table_widget.setSpan(i + ins_ind, 1, 1, 2)
                if any(['Подъем НКТ' in row_str for row_index, row_str in enumerate(row_data) if type(row_str) == str]) :
                    table_widget.setSpan(i + ins_ind, 5, 1, 5)
                elif any(['Демонтаж ЭЦН' in row_str or 'Демонтаж УЭЦН' in row_str
                          for row_index, row_str in enumerate(row_data) if type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 2)
                    table_widget.setSpan(i + ins_ind, 7, 1, 3)
                    table_widget.setSpan(i + ins_ind, 11, 1, 2)

                elif any(['Крезол' in row_str for row_index, row_str in enumerate(row_data) if type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 3)
                    table_widget.setSpan(i + ins_ind, 9, 1, 2)
                    table_widget.setSpan(i + ins_ind, 11, 1, 2)

                elif any(['Спуск НКТ' in row_str for row_index, row_str in enumerate(row_data) if type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 6)
                elif any(['что бурили' in row_str for row_index, row_str in enumerate(row_data) if type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 5)
                elif any(['Осложнение при подъеме' in row_str for row_index, row_str in enumerate(row_data) if
                          type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 7)
                elif any(['ПЗР к глушению ' in row_str for row_index, row_str in enumerate(row_data) if
                          type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 3)
                    table_widget.setSpan(i + ins_ind, 9, 1, 5)
                elif any(['Смена объема ' in row_str for row_index, row_str in enumerate(row_data) if
                          type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 5)
                    table_widget.setSpan(i + ins_ind, 12, 1, 2)
                elif any(['причины рязрядки:' in row_str for row_index, row_str in enumerate(row_data) if
                          type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 4)
                    table_widget.setSpan(i + ins_ind, 12, 1, 2)

                elif any(['Подъем штанг' in row_str for row_str in row_data if type(row_str) == str]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 8)

                elif all([row_str == None for row_index, row_str in enumerate(row_data) if 6 < row_index < 14]):
                    table_widget.setSpan(i + ins_ind, 5, 1, 9)
                elif 'классификация простоя' in row_data:
                    table_widget.setSpan(i + ins_ind, 5, 1, 5)
                    table_widget.setSpan(i + ins_ind, 10, 1, 2)
                    table_widget.setSpan(i + ins_ind, 12, 1, 2)

    def set_cell_color(self, item, color):

        # Преобразуем цвет из hex в RGB
        rgb_color = (int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16))
        item.setBackground(QColor(*rgb_color))

    def copy_norm(self, sheet, table_widget, work_plan='normir_new', count_col=31, list_page=1):

        rows = 46
        merged_cells = sheet.merged_cells
        table_widget.setRowCount(rows)
        well_data.count_row_well = table_widget.rowCount()


        table_widget.setColumnCount(count_col)

        rowHeights_exit = [sheet.row_dimensions[i + 1].height if sheet.row_dimensions[i + 1].height is not None else 18
                           for i in range(rows+1)]

        colWidth = [sheet.column_dimensions[get_column_letter(col_ind + 1)].width * 4 for col_ind in range(rows)]
        colors = []
        data = []
        for row in range(1, rows + 1):
            row_colors = []
            row_data = []
            if row > 1 and row < rows - 1:
                try:
                    table_widget.setRowHeight(row, int(rowHeights_exit[row]))
                except:
                    pass
            for col in range(1, count_col + 1):

                cell = sheet.cell(row=row, column=col)
                item = QTableWidgetItem(str(cell.value))
                row_data.append(cell.value)
                if cell.fill and cell.fill.fill_type == 'solid':
                    color_cell = cell.fill.start_color.index
                    self.set_cell_color(item, color_cell)
                    row_colors.append(cell.fill.start_color.index)
                else:
                    row_colors.append('FFFFFF')

                if not cell.value is None:

                    # Проверяем, является ли текущая ячейка объединенной
                    for merged_cell in merged_cells:
                        if merged_cell.coord == 'A46:AE46':

                            a = merged_cell

                        if row in range(merged_cell.min_row, merged_cell.max_row + 1) and \
                                col in range(merged_cell.min_col, merged_cell.max_col + 1):
                            # Устанавливаем количество объединяемых строк и столбцов для текущей ячейки
                            table_widget.setSpan(row - 1, col - 1,
                                                 merged_cell.max_row - merged_cell.min_row + 1,
                                                 merged_cell.max_col - merged_cell.min_col + 1)

                else:
                    item = QTableWidgetItem("")
            data.append(row_data)
            colors.append(row_colors)

        table_widget.setSpan(46, 1, 1, 31)


        for column in range(table_widget.columnCount()):
            table_widget.setColumnWidth(column, int(colWidth[column]))




if __name__ == "__main__":
    # app3 = QApplication(sys.argv)

    app = QApplication(sys.argv)
    # MyWindow.delete_files()

    if MyWindow.check_process():
        MyWindow.show_confirmation()

    try:
        well_data.connect_in_base = MyWindow.check_connection(well_data.host_krs)
        if well_data.connect_in_base is False:
            mes = QMessageBox.information(None, 'Проверка соединения',
                                          'Проверка показало что с облаком соединения нет, '
                                          'будет использована локальная база данных')
        MyWindow.login_window = LoginWindow()
        MyWindow.login_window.show()
        MyWindow.pause_app()
        well_data.pause = False
    except Exception as e:
        mes = QMessageBox.warning(None, 'КРИТИЧЕСКАЯ ОШИБКА',
                                  f'Критическая ошибка, смотри в лог {type(e).__name__}\n\n{str(e)}')

    # if well_data.connect_in_base:
    #     app2 = UpdateChecker()
    #     app2.check_version()
    #     if app2.window_close == True:
    #         MyWindow.set_modal_window(None, app2)
    #         well_data.pause = True
    #         MyWindow.pause_app()
    #         well_data.pause = False
    #         app2.close()

    window = MyWindow()
    window.show()
    # screen_geometry = QApplication.desktop().availableGeometry()
    # window_width = int(screen_geometry.width() * 0.9)
    # window_height = int(screen_geometry.height() * 0.9)
    #
    # window.setGeometry(0, 0, window_width, window_height)
    #

    sys.exit(app.exec_())
