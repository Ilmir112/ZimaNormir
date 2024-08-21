import json
import os
import re
import sqlite3
from io import BytesIO

import openpyxl
import psycopg2
import well_data
import base64


from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QLineEdit, QHeaderView, QVBoxLayout, QMainWindow, QWidget, \
    QTableWidget

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.utils import get_column_letter, range_boundaries


from PIL import Image
from main import MyWindow



class Classifier_well(QMainWindow):
    number_well = None

    def __init__(self, costumer, region, classifier_well, parent=None):

        super(Classifier_well, self).__init__()
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.table_class = QTableWidget()
        self.region = region
        self.costumer = costumer
        self.number_well = None
        if well_data.well_number:
            self.number_well = well_data.well_number._value

        self.setCentralWidget(self.table_class)
        self.model = self.table_class.model()
        if classifier_well == 'classifier_well':
            self.open_to_sqlite_class_well(costumer, region)
        elif classifier_well == 'damping':
            self.open_to_sqlite_without_juming(costumer, region)

    def open_to_sqlite_without_juming(self, costumer, region):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()

        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        data = self.get_data_from_db(region)

        self.table_class.setColumnCount(len(data[0]))
        self.table_class.setRowCount(len(data))
        self.table_class.setCellWidget(0, 0, self.edit_well_number)
        for row in range(len(data)):
            for col in range(len(data[row])):
                item = QTableWidgetItem(str(data[row][col]))
                self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(['номер скважины', 'площадь', 'Текущий квартал'])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def open_to_sqlite_class_well(self, costumer, region):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()
        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter_class)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        self.edit_well_area = QLineEdit()
        self.edit_well_area.setPlaceholderText("Ввести площадь для фильтрации")
        self.edit_well_area.textChanged.connect(self.filter_class_area)
        layout.addWidget(self.edit_well_area)
        region = f'{region}_классификатор'
        # print(region)
        data = self.get_data_from_class_well_db(region)
        # print(data)

        self.table_class.setColumnCount(len(data[0]))
        self.table_class.setRowCount(len(data))
        self.table_class.setCellWidget(0, 1, self.edit_well_number)
        self.table_class.setCellWidget(0, 2, self.edit_well_area)
        for row in range(len(data)):
            for col in range(len(data[row])):
                if col in [5, 6, 10, 11, 13]:
                    if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                        item = QTableWidgetItem(str(round(float(data[row][col]), 1)))
                    else:
                        item = QTableWidgetItem(str(data[row][col]))
                elif col in [9]:
                    if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                        item = QTableWidgetItem(str(round(float(data[row][col]), 7)))
                    else:
                        item = QTableWidgetItem(str(data[row][col]))

                else:
                    item = QTableWidgetItem(str(data[row][col]))
                self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(
            ['ЦДНГ', 'номер скважины', 'площадь', 'Месторождение', 'Категория \n по Рпл',
             'Ргд', 'Рпл', 'Дата замера', 'категория \nH2S', 'H2S-%', "H2S-мг/л",
             "H2S-мг/м3", 'Категория по газу', "Газовый фактор", "версия от"])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def get_data_from_db(self, region):
        if well_data.connect_in_base:

            # Параметры подключения к PostgreSQL
            try:
                # Создание подключения к базе данных PostgreSQL
                conn = psycopg2.connect(**well_data.postgres_params_classif)

                # Выполнение SQL-запроса для получения данных
                with conn.cursor() as cur:
                    cur.execute(f"""
                        SELECT well_number, deposit_area, today
                        FROM {region};
                    """)
                    data = cur.fetchall()

            except psycopg2.Error as e:
                 QMessageBox.warning(self, 'Ошибка', f'Ошибка подключения к базе данных: {type(e).__name__}\n\n{str(e)}')

            finally:
                if conn:
                    conn.close()
        else:
            # try:
            db_path = connect_to_db('databaseclassification.db', '')
            # Создание подключения к базе данных SQLite
            conn = sqlite3.connect(db_path)

            # Выполнение SQL-запроса для получения данных
            cursor = conn.cursor()
            cursor.execute(f"""
                SELECT well_number, deposit_area, today
                FROM {region};
            """)
            data = cursor.fetchall()

            # except sqlite3.Error as e:
            #      QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')
            #
            # finally:
            #     # Закрыть курсор и соединение
            #     if cursor:
            #         cursor.close()
            #     if conn:
            #         conn.close()

        return data

    def get_data_from_class_well_db(self, region):
        # Параметры подключения к PostgreSQL
        if well_data.connect_in_base:
            try:
                # Создание подключения к базе данных PostgreSQL
                conn = psycopg2.connect(**well_data.postgres_params_classif)

                # Выполнение SQL-запроса для получения данных
                with conn.cursor() as cur:
                    cur.execute(f"""
                        SELECT cdng, well_number, deposit_area, oilfield, categoty_pressure,
                               pressure_Ppl, pressure_Gst, date_measurement, categoty_h2s,
                               h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor, today
                        FROM {region};
                    """)
                    data = cur.fetchall()

            except psycopg2.Error as e:
                # Выведите сообщение об ошибке
                QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')

                return []
            finally:
                if conn:
                    conn.close()
        else:
            try:
                db_path = connect_to_db('databaseclassification.db', '')
                # Создание подключения к базе данных SQLite
                conn = sqlite3.connect(db_path)

                # Выполнение SQL-запроса для получения данных
                cursor = conn.cursor()
                cursor.execute(f"""
                    SELECT cdng, well_number, deposit_area, oilfield, categoty_pressure,
                               pressure_Ppl, pressure_Gst, date_measurement, categoty_h2s,
                               h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor, today
                    FROM {region};
                """)
                data = cursor.fetchall()

            except sqlite3.Error as e:
                # Выведите сообщение об ошибке
                QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')

                return
            finally:
                # Закрыть курсор и соединение
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()

        return data

    def export_to_sqlite_without_juming(self, fname, costumer, region):
        # Загрузка файла Excel
        wb = load_workbook(fname)
        ws = wb.active

        # Получение данных из Excel и запись их в базу данных
        for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            for col, value in enumerate(row):
                if not value is None and col <= 18:
                    # print(value)
                    if 'туймазин' in str(value).lower():
                        check_param = 'ТГМ'
                    elif 'ишимбай' in str(value).lower():
                        check_param = 'ИГМ'
                    elif 'чекмагуш' in str(value).lower():
                        check_param = 'ЧГМ'
                    elif 'красно' in str(value).lower():
                        check_param = 'КГМ'
                    elif 'арлан' in str(value).lower():
                        check_param = 'АГМ'
                    if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                            value) or '01.10.' in str(value):

                        version_year = re.findall(r'[0-9.]', str(value))
                        version_year = ''.join(version_year)
                        if version_year[-1] == '.':
                            version_year = version_year[:-1]
            if index_row > 18:
                break

        if well_data.connect_in_base:
            try:
                # Подключение к базе данных
                conn = psycopg2.connect(**well_data.postgres_params_classif)
                cursor = conn.cursor()
                region_list = ['ЧГМ', 'АГМ', 'ТГМ', 'ИГМ', 'КГМ', ]

                for region_name in region_list:
                    if region_name == region:
                        # # Удаление всех данных из таблицы
                        # cursor.execute("DROP TABLE my_table")
                        #
                        # Запрос на удаление таблицы
                        cursor.execute(f"DROP TABLE IF EXISTS {region_name}")

                        # Создание таблицы в базе данных
                        cursor.execute(f'CREATE TABLE IF NOT EXISTS {region_name}'
                                       f'(well_number TEXT,'
                                       f'deposit_area TEXT, '
                                       f'today TEXT,'
                                       f'region TEXT,'
                                       f'costumer TEXT)')

                        # print(region_name, version_year)
                        # print(check_param)
                        if check_param in region_name:
                            QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                      f'регион выбрано корректно  {region_name}')
                            try:
                                # Получение данных из Excel и запись их в базу данных
                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                    if 'ПЕРЕЧЕНЬ' in row:
                                        check_file = True
                                    if 'Скважина' in row:
                                        area_row = index_row + 2
                                        for col, value in enumerate(row):
                                            if not value is None and col <= 20:
                                                if 'Скважина' == value:
                                                    well_column = col
                                                elif 'Площадь' == value:
                                                    area_column = col

                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                    if index_row > area_row:

                                        well_number = row[well_column]
                                        area_well = row[area_column]

                                        if well_number:
                                            cursor.execute(
                                                f"INSERT INTO {region_name} (well_number, deposit_area, today, region, costumer) "
                                                f"VALUES (%s, %s, %s, %s,%s)",
                                                (well_number, area_well, version_year, region_name, costumer))

                                mes = QMessageBox.information(self, 'данные обновлены', 'Данные обновлены')
                            except:
                                 QMessageBox.warning(self, 'ОШИБКА', 'Выбран файл с не корректными данными')

                        else:
                             QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                      f'в Данном перечне отсутствую скважины {region_name}')

                # Сохранение изменений
                conn.commit()

            except psycopg2.Error as e:
                # Выведите сообщение об ошибке
                 QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')
            finally:
                # Закройте курсор и соединение
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
        else:
            try:
                db_path = connect_to_db('databaseclassification.db', '')
                # Создание подключения к базе данных SQLite
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                region_list = ['ЧГМ', 'АГМ', 'ТГМ', 'ИГМ', 'КГМ']

                for region_name in region_list:
                    if region_name == region:
                        # Удаление таблицы, если она существует
                        cursor.execute(f"DROP TABLE IF EXISTS {region_name}")

                        # Создание таблицы
                        cursor.execute(
                            f'CREATE TABLE IF NOT EXISTS {region_name} ('
                            f'well_number TEXT, '
                            f'deposit_area TEXT, '
                            f'today TEXT, '
                            f'region TEXT, '
                            f'costumer TEXT)'
                        )

                        if check_param in region_name:
                            QMessageBox.warning(None, 'ВНИМАНИЕ ОШИБКА',
                                                f'регион выбрано корректно  {region_name}')
                            try:
                                # Получение данных из Excel и запись в базу данных
                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                    if 'ПЕРЕЧЕНЬ' in row:
                                        check_file = True
                                    if 'Скважина' in row:
                                        area_row = index_row + 2
                                        for col, value in enumerate(row):
                                            if not value is None and col <= 20:
                                                if 'Скважина' == value:
                                                    well_column = col
                                                elif 'Площадь' == value:
                                                    area_column = col

                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                    if index_row > area_row:
                                        well_number = row[well_column]
                                        area_well = row[area_column]

                                        if well_number:
                                            cursor.execute(
                                                f"INSERT INTO {region_name} (well_number, deposit_area, today, region, costumer) "
                                                f"VALUES (?, ?, ?, ?, ?)",
                                                (well_number, area_well, version_year, region_name, costumer)
                                            )

                                QMessageBox.information(None, 'данные обновлены', 'Данные обновлены')
                            except Exception as e:
                                QMessageBox.warning(None, 'ОШИБКА', f'Выбран файл с не корректными данными: {type(e).__name__}\n\n{str(e)}')

                        else:
                            QMessageBox.warning(None, 'ВНИМАНИЕ ОШИБКА',
                                                f'в Данном перечне отсутствую скважины {region_name}')

                        # Сохранение изменений
                        conn.commit()

            except sqlite3.Error as e:
                # Выведите сообщение об ошибке
                QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных: /n {type(e).__name__}\n\n{str(e)}')
            finally:
                # Закройте курсор и соединение
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()

    def filter(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(0, 1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class_area(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def export_to_sqlite_class_well(self, fname, costumer, region):
        # Параметры подключения к PostgreSQL

        region_list = ['ЧГМ_классификатор', 'АГМ_классификатор', 'ТГМ_классификатор', 'ИГМ_классификатор',
                       'КГМ_классификатор']

        # Загрузка файла Excel
        wb = load_workbook(fname)
        ws = wb.active

        # Определение столбцов
        well_column, cdng, area_column, oilfield, categoty_pressure = None, None, None, None, None
        pressure_Gst, date_measurement, pressure_Ppl, categoty_h2s = None, None, None, None
        h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor = None, None, None, None, None
        area_row = 0
        check_file = False
        check_param = None

        for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if 'Классификация' in row:
                check_file = True
            elif any(['фонд' in str(value).lower() for col, value in enumerate(row)]):
                for col, value in enumerate(row):
                    if 'туймазин' in str(value).lower():
                        check_param = 'ТГМ'
                    if 'ишимбай' in str(value).lower():
                        check_param = 'ИГМ'
                    if 'чекмагуш' in str(value).lower():
                        check_param = 'ЧГМ'
                    if 'красно' in str(value).lower():
                        check_param = 'КГМ'
                    if 'арлан' in str(value).lower():
                        check_param = 'АГМ'

            elif 'Скважина' in row:
                area_row = index_row + 2
                for col, value in enumerate(row):
                    if not value is None and col <= 20:
                        if 'Скважина' == value:
                            well_column = col
                        elif 'Цех' == value:
                            cdng = col
                        elif 'Площадь' == value:
                            area_column = col
                        elif 'Месторождение' == value:
                            oilfield = col
                        elif 'Пластовое давление' == value:
                            categoty_pressure = col
                            pressure_Gst = col + 1
                            date_measurement = col + 2
                            pressure_Ppl = col + 3
                        elif 'содержание сероводорода' in str(value).lower():
                            categoty_h2s = col
                            h2s_pr = col + 1
                            h2s_mg_l = col + 2
                            h2s_mg_m = col + 3
                        elif 'Газовый фактор' == value:
                            categoty_gf = col
                            gas_factor = col + 1

        if well_data.connect_in_base:
            try:
                # Создание подключения к базе данных PostgreSQL
                conn = psycopg2.connect(**well_data.postgres_params_classif)
                cursor = conn.cursor()

                for region_name in region_list:
                    if region in region_name:
                        # Удаление всех данных из таблицы (опционально)
                        cursor.execute(f"DROP TABLE IF EXISTS {region_name};")

                        # Создание таблицы, если она не существует
                        cursor.execute(f"""
                            CREATE TABLE IF NOT EXISTS {region_name} (
                                ID SERIAL PRIMARY KEY NOT NULL,
                                cdng TEXT,
                                well_number TEXT,
                                deposit_area TEXT,
                                oilfield TEXT,
                                categoty_pressure TEXT,
                                pressure_Ppl TEXT,
                                pressure_Gst TEXT,
                                date_measurement TEXT,
                                categoty_h2s TEXT,
                                h2s_pr TEXT,
                                h2s_mg_l TEXT,
                                h2s_mg_m TEXT,
                                categoty_gf TEXT,
                                gas_factor TEXT,
                                today TEXT,
                                region TEXT,
                                costumer TEXT
                            );
                        """)

                        if check_param in region_name:
                            QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                      f'регион выбрано корректно  {region_name}')

                            try:
                                # Вставка данных в таблицу

                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                    if index_row < area_row and check_file:
                                        for col, value in enumerate(row):
                                            if not value is None and col <= 18:
                                                if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                                        value) or '01.10.' in str(value):
                                                    version_year = re.findall(r'[0-9.]', str(value))
                                                    version_year = ''.join(version_year)
                                                    if version_year[-1] == '.':
                                                        version_year = version_year[:-1]
                                    elif index_row > area_row and check_file:
                                        well_number = row[well_column]
                                        area_well = row[area_column]
                                        oilfield_str = row[oilfield]

                                        for col, value in enumerate(row):
                                            if not value is None and col <= 18:
                                                if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                                        value) or '01.10.' in str(value):
                                                    version_year = re.findall(r'[0-9.]', str(value))
                                                    version_year = ''.join(version_year)
                                                    if version_year[-1] == '.':
                                                        version_year = version_year[:-1]

                                        if well_number:
                                            cursor.execute(f"""
                                                INSERT INTO {region_name} (
                                                    cdng, well_number, deposit_area, oilfield,
                                                    categoty_pressure, pressure_Ppl, pressure_Gst, date_measurement,
                                                    categoty_h2s, h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor,
                                                    today, region, costumer
                                                )
                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                            """, (
                                                row[cdng], well_number, area_well, oilfield_str, row[categoty_pressure],
                                                row[pressure_Ppl], row[pressure_Gst], row[date_measurement],
                                                row[categoty_h2s],
                                                row[h2s_pr], row[h2s_mg_l], row[h2s_mg_m], row[categoty_gf],
                                                row[gas_factor],
                                                version_year, region, costumer
                                            ))
                            except:
                                 QMessageBox.warning(self, 'ОШИБКА', 'Выбран файл с не корректными данными')

                        else:
                             QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                      f'в Данном перечне отсутствую скважины {region_name}')
                        conn.commit()
                mes = QMessageBox.information(self, 'Успешно', 'Классификатор успешно обновлен')

            except (psycopg2.Error, Exception) as e:
                # Выведите сообщение об ошибке
                 QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')
            finally:
                # Закройте курсор и соединение
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()
        else:
            try:
                db_path = connect_to_db('databaseclassification.db', '')
                # Создание подключения к базе данных SQLite
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                for region_name in region_list:
                    if region in region_name:
                        # Удаление таблицы, если она существует
                        cursor.execute(f"DROP TABLE IF EXISTS {region_name}")

                        # Создание таблицы
                        cursor.execute(
                            f'CREATE TABLE IF NOT EXISTS {region_name} ('
                            f'ID INTEGER PRIMARY KEY AUTOINCREMENT, '
                            f'cdng TEXT, '
                            f'well_number TEXT, '
                            f'deposit_area TEXT, '
                            f'oilfield TEXT, '
                            f'categoty_pressure TEXT, '
                            f'pressure_Ppl TEXT, '
                            f'pressure_Gst TEXT, '
                            f'date_measurement TEXT, '
                            f'categoty_h2s TEXT, '
                            f'h2s_pr TEXT, '
                            f'h2s_mg_l TEXT, '
                            f'h2s_mg_m TEXT, '
                            f'categoty_gf TEXT, '
                            f'gas_factor TEXT, '
                            f'today TEXT, '
                            f'region TEXT, '
                            f'costumer TEXT'
                            f')'
                        )
                        # Получение года из даты (добавлен re.findall)
                        version_year = None

                        if check_param in region_name:
                            QMessageBox.warning(None, 'ВНИМАНИЕ ОШИБКА',
                                                f'регион выбрано корректно  {region_name}')
                            try:
                                # Вставка данных в таблицу
                                for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):

                                    if index_row < area_row and check_file:
                                        for col, value in enumerate(row):
                                            if not value is None and col <= 18:
                                                if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                                        value) or '01.10.' in str(value):
                                                    version_year = re.findall(r'[0-9.]', str(value))
                                                    version_year = ''.join(version_year)
                                                    if version_year[-1] == '.':
                                                        version_year = version_year[:-1]
                                    elif index_row > area_row and check_file:
                                        well_number = row[well_column]
                                        area_well = row[area_column]
                                        oilfield_str = row[oilfield]

                                        # Вставка данных, если год найден
                                        if well_number:
                                            cursor.execute(
                                                f"INSERT INTO {region_name} (cdng, well_number, deposit_area, oilfield, categoty_pressure, pressure_Ppl, pressure_Gst, date_measurement, categoty_h2s, h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor, today, region, costumer) "
                                                f"VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                                                (
                                                    row[cdng], well_number, area_well, oilfield_str,
                                                    row[categoty_pressure],
                                                    row[pressure_Ppl], row[pressure_Gst], row[date_measurement],
                                                    row[categoty_h2s],
                                                    row[h2s_pr], row[h2s_mg_l], row[h2s_mg_m], row[categoty_gf],
                                                    row[gas_factor],
                                                    version_year, region, costumer
                                                )
                                            )

                                # Сохранение изменений
                                conn.commit()
                                QMessageBox.information(None, 'данные обновлены', 'Данные обновлены')

                            except Exception as e:
                                QMessageBox.warning(None, 'ОШИБКА', f'Выбран файл с не корректными данными: {type(e).__name__}\n\n{str(e)}')

                        else:
                            QMessageBox.warning(None, 'ВНИМАНИЕ ОШИБКА',
                                                f'в Данном перечне отсутствую скважины {region_name}')



            except sqlite3.Error as e:
                # Выведите сообщение об ошибке
                QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных : {type(e).__name__}\n\n{str(e)}')
            finally:
                # Закройте курсор и соединение
                if cursor:
                    cursor.close()
                if conn:
                    conn.close()


def insert_database_well_data(well_number, well_area, contractor, costumer, data_well_dict, excel, work_plan):

    data_well = json.dumps(data_well_dict, ensure_ascii=False)
    excel_json = json.dumps(excel, ensure_ascii=False)
    date_today = datetime.now()
    # print(row, well_data.count_row_well)
    if 'dop_plan' in work_plan:
        work_plan_str = f'ДП№{well_data.number_dp}'
    elif 'krs' in work_plan or 'plan' in work_plan:
        work_plan_str = 'ПР'
    if well_data.connect_in_base:
        try:

            conn = psycopg2.connect(**well_data.postgres_params_data_well)
            cursor = conn.cursor()
            # Проверка наличия строки с заданными параметрами
            cursor.execute("""
                   SELECT EXISTS (
                       SELECT 1 
                       FROM wells
                       WHERE well_number = %s AND area_well = %s AND contractor = %s AND costumer = %s AND work_plan = %s
                   ), today -- Добавляем contractor в SELECT
               FROM wells 
               WHERE well_number = %s AND area_well = %s AND contractor = %s AND costumer = %s AND work_plan = %s
               """, (
                str(well_number), well_area, contractor, costumer, work_plan_str, str(well_number), well_area, contractor, costumer, work_plan_str))


            row_exists = cursor.fetchone()

            if row_exists:
                row_exists, date_in_base = row_exists
                reply = QMessageBox.question(None, 'Строка найдена',
                                             f'Строка с {well_number} {well_area} {work_plan} уже существует от {date_in_base}. '
                                             f'Обновить данные?')
                if reply == QMessageBox.Yes:
                    create_database_well_db(well_data.work_plan, well_data.number_dp)
                    try:
                        cursor.execute("""
                                        UPDATE wells
                                        SET data_well =%s, today =%s, excel_json =%s, work_plan=%s, geolog =%s                                                                 
                                        WHERE well_number =%s AND area_well =%s AND contractor = %s AND costumer =%s AND work_plan =%s
                                    """, (
                            data_well, date_today, excel_json,  work_plan_str, well_data.user[1],
                            str(well_number), well_area, contractor, costumer, work_plan_str))

                        QMessageBox.information(None, 'Успешно', 'Данные обновлены')
                    except (Exception, psycopg2.Error) as error:
                        QMessageBox.critical(None, 'Ошибка', f'Ошибка при обновлении данных: {error}')
            else:
                create_database_well_db(well_data.work_plan, well_data.number_dp)

                # Подготовленные данные для вставки
                data_values = (str(well_number), well_area,
                               data_well, date_today, excel_json, contractor, well_data.costumer, work_plan_str, well_data.user[1])

                # Подготовленный запрос для вставки данных с параметрами
                query = f"INSERT INTO wells VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)"

                # Выполнение запроса с использованием параметров
                cursor.execute(query, data_values)

                mes = QMessageBox.information(None, 'база данных',
                                              f'Скважина {well_data.well_number._value} добавлена в базу данных c excel файлами')

            # Сохранить изменения и закрыть соединение
            conn.commit()
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()


        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
             QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных  well_data {type(e).__name__}\n\n{str(e)}')
    else:
        try:

            # Формируем полный путь к файлу базы данных
            db_path = connect_to_db('well_data.db', 'data_base_well/')

            conn = sqlite3.connect(f'{db_path}')
            cursor = conn.cursor()
            # Проверка наличия строки с заданными параметрами
            cursor.execute("""
                   SELECT EXISTS (
                       SELECT 1 
                       FROM wells
                       WHERE well_number = ? AND area_well = ? AND contractor = ? AND costumer = ? AND work_plan = ?
                   ), today -- Добавляем contractor в SELECT
               FROM wells 
               WHERE well_number = ? AND area_well = ? AND contractor = ? AND costumer = ? AND work_plan = ?
               """, (
                str(well_number), well_area, contractor, costumer, work_plan_str, str(well_number), well_area, contractor, costumer, work_plan_str))

            row_exists = cursor.fetchone()

            if row_exists:
                row_exists, date_in_base = row_exists
                reply = QMessageBox.question(None, 'Строка найдена',
                                             f'Строка с {well_number} {well_area} уже существует от {date_in_base}. '
                                             f'Обновить данные?')
                if reply == QMessageBox.Yes:
                    create_database_well_db(well_data.work_plan, well_data.number_dp)
                    try:
                        cursor.execute("""
                                        UPDATE wells
                                        SET data_well = ?, today = ?, excel_json = ?, work_plan = ?, geolog = ?                                                                      
                                        WHERE well_number = ? AND area_well = ? 
                                        AND contractor = ? AND costumer = ? AND work_plan = ? AND geolog = ?
                                    """, (
                            data_well, date_today, excel_json,  work_plan_str, well_data.user[1],
                        well_number, well_area, contractor, costumer, work_plan, well_data.user[1]))

                        QMessageBox.information(None, 'Успешно', 'Данные в обновлены обновлены')
                    except sqlite3.Error as error:
                        QMessageBox.critical(None, 'Ошибка', f'Ошибка при обновлении данных: {error}')
            else:
                create_database_well_db(well_data.work_plan, well_data.number_dp)

                # Подготовленные данные для вставки
                data_values = (str(well_number), well_area,
                               data_well, date_today, excel_json, contractor, well_data.costumer, work_plan_str,
                               well_data.user[1])

                # Подготовленный запрос для вставки данных с параметрами
                query = """INSERT INTO wells(well_number, area_well, data_well, today, excel_json, 
                contractor, costumer, work_plan, geolog) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);"""

                # Выполнение запроса с использованием параметров
                cursor.execute(query, data_values)

                QMessageBox.information(None, 'база данных', f'Скважина {well_data.well_number._value} '
                                                                   f'добавлена в базу данных welldata')

            # Сохранить изменения и закрыть соединение
            conn.commit()
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()

        except sqlite3.Error as e:
            # Выведите сообщение об ошибке
             QMessageBox.warning(None, 'Ошибка', f'Ошибка подключения к базе данных hg{type(e).__name__}\n\n{str(e)}')


def connect_to_db(name_base, folder_base):
    # Получаем текущий каталог приложения
    current_dir = os.path.dirname(__file__)

    # Определяем путь к папке с базой данных
    db_folder = os.path.join(current_dir, folder_base)

    # Формируем полный путь к файлу базы данных
    db_path = os.path.join(db_folder, name_base)

    return db_path



def check_in_database_well_data(number_well, area_well, work_plan):
    work_plan = 'krs'
    if well_data.connect_in_base:
        try:
            conn = psycopg2.connect(**well_data.postgres_params_data_well)
            cursor = conn.cursor()

            cursor.execute("SELECT data_well FROM wells WHERE well_number = %s AND area_well = %s "
                           "AND contractor = %s AND costumer = %s",
                           (str(number_well), area_well, well_data.contractor, well_data.costumer))

            data_well = cursor.fetchone()
            if data_well:
                return data_well
            else:
                return False, data_well

        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
             QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных, Скважина не добавлена в базу')
    else:
        try:
            db_path = connect_to_db('well_data.db', 'data_base_well/')

            conn = sqlite3.connect(f'{db_path}')
            cursor = conn.cursor()
            a = well_data.contractor

            cursor.execute("SELECT data_well FROM wells WHERE well_number = ? AND area_well = ? "
                           "AND contractor = ? AND costumer = ?",
                           (str(number_well), area_well, well_data.contractor, well_data.costumer))

            data_well = cursor.fetchone()
            if data_well:
                return data_well
            else:
                return False, data_well

        except sqlite3.Error as e:
            # Выведите сообщение об ошибке
             QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных, Скважина не добавлена в базу')


def excel_in_json(sheet):
    index_end_copy = 46

    data = {}
    for row_index, row in enumerate(sheet.iter_rows()):
        row_data = []
        if all(cell == None for cell in row[:32]) is False:
            if any([cell.value == "ИТОГО:" for cell in row[:4]]):
                index_end_copy = row_index

                break
            for cell in row[:32]:
                # Получение значения и стилей
                value = cell.value

                font = cell.font
                fill = cell.fill
                # Преобразуем RGB в строковый формат
                rgb_string = f"RGB({fill.fgColor.rgb})"

                borders = cell.border
                alignment = cell.alignment

                row_data.append({
                    'value': value,
                    'font': {
                        'name': font.name,
                        'size': font.size,
                        'bold': font.bold,
                        'italic': font.italic
                    },
                    'fill': {
                        'color': rgb_string
                    },
                    'borders': {
                        'left': borders.left.style,
                        'right': borders.right.style,
                        'top': borders.top.style,
                        'bottom': borders.bottom.style
                    },
                    'alignment': {
                        'horizontal': alignment.horizontal,
                        'vertical': alignment.vertical,
                        'wrap_text': alignment.wrap_text
                    },
                })
                data[row[0].row] = row_data

    data['image'] = well_data.image_data
    rowHeights = [sheet.row_dimensions[i + 1].height for i in range(sheet.max_row) if i <= index_end_copy]
    # rowHeights = [sheet.row_dimensions[i + 1].height for i in range(sheet.max_row)if i <= 46]
    colWidth = [sheet.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 80)] + [None]
    boundaries_dict = {}

    for ind, _range in enumerate(sheet.merged_cells.ranges):
        if range_boundaries(str(_range))[1] < index_end_copy:
            boundaries_dict[ind] = range_boundaries(str(_range))



    data_excel = {'data': data, 'rowHeights': rowHeights, 'colWidth': colWidth, 'merged_cells': boundaries_dict}

    return data_excel


def insert_data_well_dop_plan(data_well):
    from well_data import ProtectedIsDigit, ProtectedIsNonNone

    well_data_dict = json.loads(data_well)


    well_data.column_direction_diametr = ProtectedIsDigit(well_data_dict["направление"]["диаметр"])
    well_data.column_direction_wall_thickness = ProtectedIsDigit(well_data_dict["направление"]["толщина стенки"])
    well_data.column_direction_lenght = ProtectedIsDigit(well_data_dict["направление"]["башмак"])
    well_data.level_cement_direction = ProtectedIsDigit(well_data_dict["направление"]["цемент"])
    well_data.column_conductor_diametr = ProtectedIsDigit(well_data_dict["кондуктор"]["диаметр"])
    well_data.column_conductor_wall_thicknes = ProtectedIsDigit(well_data_dict["кондуктор"]["толщина стенки"])
    well_data.column_conductor_lenght = ProtectedIsDigit(well_data_dict["кондуктор"]["башмак"])
    well_data.level_cement_conductor = ProtectedIsDigit(well_data_dict["кондуктор"]["цемент"])
    well_data.column_diametr = ProtectedIsDigit(well_data_dict["ЭК"]["диаметр"])
    well_data.column_wall_thickness = ProtectedIsDigit(well_data_dict["ЭК"]["толщина стенки"])
    well_data.shoe_column = ProtectedIsDigit(well_data_dict["ЭК"]["башмак"])
    well_data.column_additional = well_data_dict["допколонна"]["наличие"]
    well_data.column_additional_diametr = ProtectedIsDigit(well_data_dict["допколонна"]["диаметр"])
    well_data.column_additional_wall_thickness = ProtectedIsDigit(well_data_dict["допколонна"]["толщина стенки"])
    well_data.shoe_column_additional = ProtectedIsDigit(well_data_dict["допколонна"]["башмак"])
    well_data.head_column_additional = ProtectedIsDigit(well_data_dict["допколонна"]["голова"])
    well_data.curator = well_data_dict["куратор"]
    well_data.dict_pump_SHGN = well_data_dict["оборудование"]["ШГН"]["тип"]
    well_data.dict_pump_SHGN_h = well_data_dict["оборудование"]["ШГН"]["глубина "]
    well_data.dict_pump_ECN = well_data_dict["оборудование"]["ЭЦН"]["тип"]
    well_data.dict_pump_ECN_h = well_data_dict["оборудование"]["ЭЦН"]["глубина "]
    well_data.paker_do = well_data_dict["оборудование"]["пакер"]["тип"]
    well_data.depth_fond_paker_do = well_data_dict["оборудование"]["пакер"]["глубина "]
    well_data.paker2_do = well_data_dict["оборудование"]["пакер2"]["тип"]
    well_data.depth_fond_paker2_do = well_data_dict["оборудование"]["пакер2"]["глубина "]
    well_data.static_level = ProtectedIsDigit(well_data_dict["статика"])
    well_data.dinamic_level = ProtectedIsDigit(well_data_dict["динамика"])
    well_data.dict_nkt_po = well_data_dict["НКТ"]
    well_data.dict_sucker_rod_po = well_data_dict["штанги"]
    well_data.Qoil = well_data_dict['ожидаемые']['нефть']
    well_data.Qwater = well_data_dict['ожидаемые']['вода']
    well_data.proc_water = well_data_dict['ожидаемые']['обводненность']
    well_data.expected_P = well_data_dict['ожидаемые']['давление']
    well_data.expected_Q = well_data_dict['ожидаемые']['приемистость']

    well_data.bottomhole_drill = ProtectedIsDigit(well_data_dict['данные']['пробуренный забой'])
    well_data.bottomhole_artificial = ProtectedIsDigit(well_data_dict['данные']['искусственный забой'])
    well_data.max_angle = ProtectedIsDigit(well_data_dict['данные']['максимальный угол'])
    well_data.max_angle_H = ProtectedIsDigit(well_data_dict['данные']['глубина'])
    well_data.max_expected_pressure = ProtectedIsDigit(well_data_dict['данные']['максимальное ожидаемое давление'])
    well_data.max_admissible_pressure = ProtectedIsDigit(well_data_dict['данные']['максимальное допустимое давление'])

    well_data.curator = well_data_dict['куратор']
    well_data.region = well_data_dict['регион']
    well_data.cdng = ProtectedIsNonNone(well_data_dict['ЦДНГ'])

    well_data.data_well_dict = well_data_dict
    mes = QMessageBox.information(None, 'Данные с базы', "Данные вставлены из базы данных")

    definition_plast_work(None)

def read_database_gnkt(contractor, gnkt_number):
    try:
        # Подключение к базе данных
        conn = psycopg2.connect(**well_data.postgres_conn_gnkt)

        if 'ойл-сервис' in contractor.lower():
            contractor = 'oil_service'
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM КГМ WHERE today =(%s)", (gnkt_number, '1963'))
        print(f' база данных открыта')
        result = cursor.fetchone()
    except psycopg2.Error as e:
        # Выведите сообщение об ошибке
         QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных')
    finally:
        # Закройте курсор и соединение
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    # print(result)


def create_database_well_db(work_plan, number_dp):
    if well_data.connect_in_base:
        try:
            conn = psycopg2.connect(**well_data.postgres_conn_work_well)
            cursor = conn.cursor()
            if number_dp == 0:
                number_dp = ''
            if 'Ойл' in well_data.contractor:
                contractor = 'ОЙЛ'
            elif 'РН' in well_data.contractor:
                contractor = 'РН'

            if work_plan in ['krs', 'plan_change']:
                work_plan = 'krs'

            # Создаем таблицу для хранения данных
            number = json.dumps(
                str(well_data.well_number._value) + " " + well_data.well_area._value + " " + work_plan + str(
                    number_dp) + ' ' + contractor,
                ensure_ascii=False)

            # Попытка удалить таблицу, если она существует
            cursor.execute(f'DROP TABLE IF EXISTS {number}')

            cursor.execute(f'CREATE TABLE IF NOT EXISTS {number}'
                           f'(index_row INTEGER,'
                           f'current_bottom FLOAT,'
                           f'perforation TEXT, '
                           f'plast_all TEXT, '
                           f'plast_work TEXT, '
                           f'leakage TEXT,'
                           f'column_additional TEXT,'
                           f'fluid TEXT,'
                           f'category_pressuar TEXT,'
                           f'category_h2s TEXT,'
                           f'category_gf TEXT,'
                           f'template_depth TEXT,'
                           f'skm_list TEXT,'
                           f'problemWithEk_depth FLOAT,'
                           f'problemWithEk_diametr FLOAT,'
                           f'today DATE)')
            date_create = datetime.now()
            for index, data in enumerate(well_data.data_list):
                current_bottom = data[1]
                dict_perforation_json = data[2]
                plast_all = data[3]
                plast_work = data[4]
                dict_leakiness = data[5]
                column_additional = data[6]
                fluid_work = data[7]
                template_depth = data[11]

                skm_interval = data[12]
                problemWithEk_depth = data[13]
                problemWithEk_diametr = data[14]


                # Подготовленные данные для вставки (пример)
                data_values = (index, current_bottom, dict_perforation_json, plast_all, plast_work,
                               dict_leakiness, column_additional, fluid_work, well_data.category_pressuar,
                               well_data.category_h2s, well_data.category_gf, template_depth, skm_interval,
                               problemWithEk_depth, problemWithEk_diametr, date_create)

                # Подготовленный запрос для вставки данных с параметрами
                query = f"INSERT INTO {number} VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

                # Выполнение запроса с использованием параметров
                cursor.execute(query, data_values)

            # Сохранить изменения и закрыть соединение
            conn.commit()
        # Закройте курсор и соединение

        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
             QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных, Скважина не добавлена в базу')
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()


    else:
        """Добавляет данные о скважине в SQLite базу данных."""

        try:
            # Формируем полный путь к файлу базы данных
            db_path = connect_to_db('databaseWell.db', 'data_base_well')
            conn = sqlite3.connect(db_path)
            cursor = conn.cursor()

            if number_dp == 0:
                number_dp = ''

            if 'Ойл' in well_data.contractor:
                contractor = 'ОЙЛ'
            elif 'РН' in well_data.contractor:
                contractor = 'РН'

            if work_plan in ['krs', 'plan_change']:
                work_plan = 'krs'

            # Создаем таблицу для хранения данных
            number = json.dumps(
                str(well_data.well_number._value) + " " + well_data.well_area._value + " " + work_plan + str(
                    number_dp) + ' ' + contractor,
                ensure_ascii=False
            )

            # Попытка удалить таблицу, если она существует
            cursor.execute(f'DROP TABLE IF EXISTS {number}')

            cursor.execute(
                f'CREATE TABLE IF NOT EXISTS {number} ('
                f'index_row INTEGER, '
                f'current_bottom FLOAT, '
                f'perforation TEXT, '
                f'plast_all TEXT, '
                f'plast_work TEXT, '
                f'leakage TEXT, '
                f'column_additional TEXT, '
                f'fluid TEXT, '
                f'category_pressuar TEXT, '
                f'category_h2s TEXT, '
                f'category_gf TEXT, '
                f'template_depth TEXT, '
                f'skm_list TEXT, '
                f'problemWithEk_depth FLOAT, '
                f'problemWithEk_diametr FLOAT,'
                f'today DATE)'
            )

            for index, data in enumerate(well_data.data_list):
                current_bottom = data[1]
                dict_perforation_json = data[2]
                plast_all = data[3]
                plast_work = data[4]
                dict_leakiness = data[5]
                column_additional = data[6]
                fluid_work = data[7]
                template_depth = data[11]
                skm_interval = data[12]
                problemWithEk_depth = data[13]
                problemWithEk_diametr = data[14]
                date_create = datetime.now().strftime("%d.%m.%Y")
                # Подготовленные данные для вставки
                data_values = (
                    index,
                    current_bottom,
                    dict_perforation_json,
                    plast_all,
                    plast_work,
                    dict_leakiness,
                    column_additional,
                    fluid_work,
                    well_data.category_pressuar,
                    well_data.category_h2s,
                    well_data.category_gf,
                    template_depth,
                    skm_interval,
                    problemWithEk_depth,
                    problemWithEk_diametr,
                    date_create
                )

                # Подготовленный запрос для вставки данных
                query = f"INSERT INTO {number} VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"

                # Выполнение запроса с использованием параметров
                cursor.execute(query, data_values)

            # Сохранить изменения и закрыть соединение
            conn.commit()


        except sqlite3.Error as e:
            # Выведите сообщение об ошибке
            QMessageBox.warning(None, 'Ошибка', f'Ошибка добавления скважины в базу данных: {type(e).__name__}\n\n{str(e)}')
        finally:
            if cursor:
                cursor.close()
            if conn:
                conn.close()

def get_table_creation_time(conn, table_name):
    if well_data.connect_in_base:
        with conn.cursor() as cur:
            # Выполняем SQL-запрос для получения всех записей из таблицы
            cur.execute(f"SELECT * FROM \"{table_name}\"")

            try:

                # Получаем результаты запроса в виде списка кортежей
                rows = cur.fetchone()[15]
                return f' от {rows}'
            except:
                return ''
    else:
        cur = conn.cursor()
        # Выполняем SQL-запрос для получения всех записей из таблицы
        cur.execute(f"SELECT * FROM '{table_name}'")

        try:
            # Получаем результаты запроса в виде списка кортежей
            rows = cur.fetchall()[0][15]

            return f' от {rows}'
        except:
            return ''


def round_cell(data):
    try:
        if data is None or type(data) is datetime:
            return data
        float_item = float(data)
        if float_item.is_integer():
            return int(float_item)
        else:
            return round(float_item, 4)
    except ValueError:
        return data


def insert_data_new_excel_file(data, rowHeights, colWidth, boundaries_dict):
    wb_new = openpyxl.Workbook()
    sheet_new = wb_new.active

    for row_index, row_data in data.items():
        if row_index != 'image':
            for col_index, cell_data in enumerate(row_data, 1):
                cell = sheet_new.cell(row=int(row_index), column=int(col_index))
                if cell_data:
                    cell.value = round_cell(cell_data['value'])

    row_max = sheet_new.max_row

    for key, value in boundaries_dict.items():
        if value[1] <= row_max:
            sheet_new.merge_cells(start_column=value[0], start_row=value[1],
                                  end_column=value[2], end_row=value[3])

    # Восстановление данных и стилей из словаря
    for row_index, row_data in data.items():
        if row_index != 'image':
            for col_index, cell_data in enumerate(row_data, 1):
                cell = sheet_new.cell(row=int(row_index), column=int(col_index))

                # Получение строки RGB из JSON
                rgb_string = cell_data['fill']['color']
                # Извлекаем шестнадцатеричный код цвета
                hex_color = rgb_string[4:-1]

                if hex_color != '00000000':

                    try:
                        color = Color(rgb=hex_color)

                        # Создание объекта заливки
                        fill = PatternFill(patternType='solid', fgColor=color)
                        cell.fill = fill
                    except:
                        pass
                cell.font = Font(name=cell_data['font']['name'], size=cell_data['font']['size'],
                                 bold=cell_data['font']['bold'], italic=cell_data['font']['italic'])

                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style=cell_data['borders']['left']),
                                                     right=openpyxl.styles.Side(style=cell_data['borders']['right']),
                                                     top=openpyxl.styles.Side(style=cell_data['borders']['top']),
                                                     bottom=openpyxl.styles.Side(style=cell_data['borders']['bottom']))

                wrap_true = cell_data['alignment']['wrap_text']

                cell.alignment = openpyxl.styles.Alignment(horizontal=cell_data['alignment']['horizontal'],
                                                           vertical=cell_data['alignment']['vertical'],
                                                           wrap_text=wrap_true)
    # wb_new.save('1234.xlsx')


    try:
        well_data.image_data = data['image']
                    # Добавьте обработку ошибки, например, пропуск изображения или запись информации об ошибке в лог
    except ValueError as e:
        print(f"Ошибка при вставке изображения: {type(e).__name__}\n\n{str(e)}")


    for col in range(13):
        sheet_new.column_dimensions[get_column_letter(col + 1)].width = colWidth[col]
    index_delete = 47

    for index_row, row in enumerate(sheet_new.iter_rows()):
        # Копирование высоты строки
        if any(['Наименование работ' in str(col.value) for col in row[:13]]) and well_data.work_plan not in ['plan_change']:
            index_delete = index_row+2
            well_data.ins_ind2 = index_row +2

        elif any(['ПЛАН РАБОТ' in str(col.value).upper() for col in row[:4]]) and well_data.work_plan not in ['plan_change']:
            sheet_new.cell(row=index_row + 1, column=2).value = f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {well_data.number_dp}'

        elif any(['ИТОГО:' in str(col.value).upper() for col in row[:4]]) and well_data.work_plan in ['plan_change']:
            index_delete = index_row + 2
            well_data.ins_ind2 = index_row + 2

        elif all([col is None for col in row[:13]]):
            sheet_new.row_dimensions[index_row].hidden = True
        try:
            sheet_new.row_dimensions[index_row].height = rowHeights[index_row - 1]
        except:
            pass
    if well_data.work_plan not in ['plan_change']:
        sheet_new.delete_rows(index_delete, sheet_new.max_row - index_delete )

    return sheet_new


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    # app.setStyleSheet()
    window = Classifier_well()
    window.show()
    sys.exit(app.exec_())
